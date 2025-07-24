#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Autocoder Core

从原始autocoder.py提取的核心功能，保持所有函数不变，只修改配置接口
"""

import os
import sys
import re
import copy
import json
import logging
import openpyxl
import pandas as pd
import numpy as np
from collections import deque
from dataclasses import dataclass
from typing import List, Dict, Any, Optional
from pathlib import Path

g_version = '1.2.5'


# 公共接口
__all__ = [
    'AutocoderConfig',
    'init_autocoder',
    'run_autocoder',
    'reset_autocoder',
    'convert_excel_to_code',
    'safe_log'
]

# 配置类
@dataclass
class AutocoderConfig:
    """Autocoder配置类"""
    debug_level: str = 'info'
    language: str = 'english'
    reg_short_description: bool = True
    mask_style: str = 'nxp'
    input_file: str = ''
    output_dir: str = './converted_markdown_files/'
    sysinfo_json: str = ''

# 全局变量 - 这些将通过init_autocoder函数设置
g_debug_level = 'info'
g_mask_style = 'nxp5777m'
g_language = 'english'
g_input_excel_file_path = ''
g_trans_flag = False
g_source_code_root_path = '.\\build-header-files\\English\\'
g_trans_enum_item_description = False
g_reg_short_description = True
g_gen_file_path = './converted_markdown_files/'

g_pinmap_filename = ''
g_register_style = ''
g_alias = ''
g_variable_name_set = set()
g_baseinfo_dict = {}
g_bits_variable_dict = {}
g_treeview_dict = {}
g_reglist_shortname_key_dict = {}
g_sysinfo_dict = {}
g_cfg_macro_dict = {}
g_ar_macro_dict = {}
#----------------------------
g_reserved_reg_tree = []
g_spread_reg_tree = []
g_reg_tree = []

#定义全局变量，用来预存储baseinfo，reglist等excel表单的内容
g_product_prefix = []
g_module_name = []
g_base_address = []
g_module_no_set = set()
g_space_size = 0
g_l1_node_number = 0

g_functions_dictlist = []
g_dev_functions_dictlist = []
g_mcal_functions_dictlist = []
g_ar_functions_dictlist = []
g_dev_enum_dictlist = []
g_mcal_enum_dictlist = []
g_ar_enum_dictlist = []
g_mcal_type_macro_dictlist = []
g_reglist_dictlist = []
g_dev_structure_dictlist = []
g_mcal_structure_dictlist = []
g_ar_structure_dictlist = []
g_ar_cfg_interfaces_dictlist = []
g_ar_type_definitions_dictlist = []
g_ar_apis_dictlist = []
g_reg_apis_dictlist = []
g_ar_error_codes_dictlist = []

#--------------------全局常量
g_general_types_h_flag = False
g_notes_format_str = '\n *\t\t\t\t\t'
g_c_file_start_str ='''#ifdef __cplusplus
extern "C"{
#endif
'''
g_c_file_end_str ='''#ifdef __cplusplus
}
#endif
'''
#----------------------------

# 日志将在init_autocoder函数中配置
logger = None

def safe_log(level: str, message: str):
    """
    安全的日志记录函数，处理logger为None的情况
    
    Args:
        level: 日志级别 ('debug', 'info', 'warning', 'error', 'critical')
        message: 日志消息
    """
    if logger is not None:
        getattr(logger, level.lower())(message)
    else:
        # 如果logger未初始化，使用print作为后备
        print(f"[{level.upper()}] {message}")

def description_2_enumitem_in_reglist_table(str_desc):
    # 判断str_desc是否有enum value的定义，pattern为行首的'Values:', 匹配不到直接返回
    match = re.search(r'Values:', str_desc, flags=re.MULTILINE)
    if not match:
        logger.debug(f'description_2_enumitem_in_reglist_table no need shaping.\n')
        return None

    # 去除Values:\n以及之前的部分
    str_desc = re.sub(r'.*?Values:\n', '', str_desc, flags=re.DOTALL)
    # 去除:Value After Reset以及之后的部分
    str_desc = re.sub(r'Value After Reset:.*', '', str_desc, flags=re.DOTALL)
    #去掉空行
    str_desc = re.sub(r'\n\s*\n', '\n', str_desc)
    logger.info(f'description_2_enumitem_in_reglist_table 输入 {g_register_style}: \n{str_desc}')
    if g_register_style == 'nxp':
        str_desc = nxp_description_2_enumitem_in_reglist_table(str_desc)

    if g_register_style == 'infineon':
        str_desc = infineon_description_2_enumitem_in_reglist_table(str_desc)

    if g_register_style == 'arkuart':
        str_desc = arkuart_description_2_enumitem_in_reglist_table(str_desc)

    str_desc = str_desc.strip()
    logger.info(f'description_2_enumitem_in_reglist_table 结果 {g_register_style}: \n{str_desc}')
    return str_desc

def split_paragraph_by_binary(text):  
    # 正则表达式模式，匹配以二进制数（0或1）开始的段落，包括该段落直到下一个二进制数或文本结束  
    pattern = r'(?m)^([01]{1,4})\s+([\s\S]*?)(?=\n^[01]{1,4}\s|\Z)' 
      
    # 使用re.finditer找到所有匹配项  
    matches = re.finditer(pattern, text, flags=re.MULTILINE| re.DOTALL)  
      
    # 从匹配项中提取并构建结果列表  
    full_parts = [(match.group(1), match.group(2).lstrip()) for match in matches]  

    parts = []
    for binary, content in full_parts:  
        hex_str = hex(int(binary, 2))
        hex_value = hex_str.replace('0x', '')  
        # content截取r'Note:|NOTE:'之前的部分
        content = re.split(r'Note:|NOTE:', content, 1, re.IGNORECASE)[0] # re.IGNORECASE使匹配忽略大小写
        content = content.replace('\n', ' ')+ ' '
        # 去掉尾部空格
        content = content.strip()  
        part = f'{hex_str}::{hex_value}::{content}'  
        parts.append(part)  

    return parts  

def nxp_extract_range_from_text(text):
    # 匹配格式为: 1 to 255 xxxxxx
    match = re.search(r'^(\d+) to (\d+)(.*)$', text, flags=re.MULTILINE|re.DOTALL)
    if match:
        logger.info(f'nxp_extract_range_from_text: \n{text}')
        # 将匹配到的字符串转换为整数
        m, n, tail = int(match.group(1)), int(match.group(2)), match.group(3).strip()
        new_text = ""  # 初始化新文本变量
        tail = tail.split('Note:')[0]
        # 用循环，填充m和n之间的数字，构建新的字符串，"m tail\n......n tail"
        # for i in range(m, n + 1):  # 注意这里是 n + 1，以确保包含上限n
        if n>15:
            n = 15
        for i in range(m, n+1): #超过1111B就截断
            binary_i = bin(i)[2:].rjust(4, '0')
            new_text += f'{binary_i} {tail}\n'
        logger.info(f'nxp_extract_range_from_text: \n{new_text}')  # 这里我使用了print代替logger.info，以便直接输出结果
        return new_text
    else:
        #如果没有找到匹配项，直接返回text
        return text
    
# 寄存器description风格转换函数。将nxp的寄存器desciption风格转换成arkuart的风格
def nxp_description_2_enumitem_in_reglist_table(new_text):  
    # 提取'1 to 255 'pattern开头的特殊格式中的数字m和n
    new_text = nxp_extract_range_from_text(new_text)

    parts = split_paragraph_by_binary(new_text)  

    for part in parts:
        logger.info(f'nxp_description_2_enumitem_in_reglist_table:遍历匹配点: \n{part}')

    # 遍历parts中的每个元素加上换行符，重新组合成新字符串
    new_text = '\n'.join(parts)

    # 去掉空行
    new_text = re.sub(r'\n\s*\n', '\n', new_text)

    logger.debug(f'nxp_description_2_enumitem_in_reglist_table: \n{new_text}')
    return new_text 

def infineon_description_2_enumitem_in_reglist_table(new_text):  
    pattern = r'^0x[0-9A-Fa-f]+\s'
    # 使用re.finditer找到所有匹配项
    matches = re.finditer(pattern, new_text, flags=re.MULTILINE)
    # 组装包含匹配数字及后续文本的新字符串列表
    parts = []
    start_index = 0
    for match in matches:
        # 匹配项的起始和当前匹配结束位置
        match_start, match_end = match.span()

        # 从上一个匹配项之后到当前匹配项的数字之前的文本（如果有）
        previous_part = new_text[start_index:match_start]
        if previous_part.strip():  # 如果非空，则添加到parts中（可选，根据需要调整）
            parts.append(previous_part)
        
        # 搜索下一个匹配项的开始位置
        next_match = re.search(pattern, new_text[match_end:], flags=re.MULTILINE)
        if next_match:
            next_match_start = next_match.span()[0] + match_end  # 匹配项的结束位置加上下一个匹配项的起始位置
        else:
            next_match_start = len(new_text)  # 如果没有找到下一个匹配项，则直到文本末尾
        
        # 添加包含当前匹配数字和后续文本的新字符串
        part_with_match = new_text[match_start:next_match_start]
        part_with_match = part_with_match.replace('\n', ' ')
        parts.append(part_with_match)        
        # 更新下一个部分的起始索引
        start_index = next_match_start
    
    new_parts = []
    for part in parts:
        '''# 示例输入文本
        text1 = '0x0       On abc: enable the module clock'
        text2 = '0x1       Off xyd: stop the module clock'
        text3 = '0x2       no the module clock'

        # 示例输出结果
        outtext1 = '0x0       (On request)enable the module clock'
        outtext2 = '0x1       (Off request)stop the module clock'
        outtext3 = '0x2       (2)no the module clock'
        '''
        # 正则表达式，匹配十六进制值hex_value、和remainder。
        # 在remainder里面，如果有冒号，则第一个冒号前为prefix，第一个冒号后为remainder；
        # 如果没有冒号，则prefix为hex_value，remainder不变。
        # 字符串拼接f'{hex_value}       ({prefix}){remainder}'
        pattern = r'^(0x[0-9A-Fa-f]+)\s+(.*)'
        match = re.search(pattern, part)
        if match:
            hex_value, remainder = match.groups()
            # 使用split方法尝试分割字符串，并检查结果列表的长度
            remaindersplit = remainder.split(':',1)
            if len(remaindersplit) == 2:
                # 如果列表长度等于2，说明成功分割，按预期赋值
                prefix, new_remainder = remaindersplit
                if len(prefix) > 16:
                    prefix = hex_value.replace('0x', '')
            elif len(remaindersplit) == 1:
                prefix = hex_value.replace('0x', '')
                new_remainder = remainder

            prefix = prefix.strip()
            new_remainder = new_remainder.strip()
            # 拼接字符串
            part = f'{hex_value}::{prefix}::{new_remainder}' 
            
        new_parts.append(part)   

    # 遍历parts中的每个元素加上换行符，重新组合成新字符串
    new_text = '\n'.join(new_parts)

    # 去掉空行
    new_text = re.sub(r'\n\s*\n', '\n', new_text)

    logger.error(f'infineon_description_2_enumitem_in_reglist_table: \n{new_text}')
    return new_text 

def arkuart_description_2_enumitem_in_reglist_table(new_text):
    # 去除可能存在的额外空格和换行符，并返回结果
    new_text = new_text.strip()
    # 去掉每行开头第一个数字或字母之前的字符，包括空白字符和特殊字符
    new_text = re.sub(r'^[^0-9a-zA-Z]+', '', new_text, flags=re.MULTILINE)
    # 每行0x0 (DISABLED): disable parity 格式转换为：0x0::DISABLED::disable parity.
    new_text = re.sub(r'^(0x[0-9a-fA-F]+)\s*\(([^)]+)\)\s*:\s*(.+)',
            lambda match: f'{match.group(1)}::{match.group(2)}::{match.group(3).strip()}', 
            new_text, flags=re.MULTILINE)
    return new_text 

def source_sheet_data_list_reverse_flag(reg_sheet_dictlist):
    reverse_flag = False
    old_retvalue = 0

    for entry in reg_sheet_dictlist:
        if ':' in entry['Bits']:
            part_mn = re.split(':', entry['Bits'])
            ret_value = int(part_mn[0])
        else:
            ret_value = int(entry['Bits'])
        if old_retvalue > ret_value:
            reverse_flag = True
            break            
        old_retvalue = ret_value

    return reverse_flag

def arkuart_pre_shaping_regsheet(short_name,reg_sheet_dictlist):
    copy_reg_sheet_dictlist = copy.deepcopy(reg_sheet_dictlist)
    for entry in copy_reg_sheet_dictlist:
        entry['Field'] = entry['Name'].replace(' ', '')
        # arkuart的Bits变量x,y替换成数字
        arkuart_style_bits_xy_2_number(entry,short_name)
        ret_value = entry['Bits']
        if ':' in ret_value:
            part_mn = re.split(':', ret_value)
            ret_value = f'[{int(part_mn[0])}:{int(part_mn[1])}]'
        else:
            ret_value = f'[{int(ret_value)}:{int(ret_value)}]'
        entry['Bits'] = ret_value
        # arkuart的'Memory Access'变量替换成Type类型R W R/W        
        entry['Type']= arkuart_style_type_variable_2_rw(entry,short_name)
        if entry['Type'] is None:  # 说明不是变量
            entry['Type'] = entry['MemoryAccess']  # 列名已经清洗过，所以Memory Access字段使用MemoryAccess
        logger.info((f"arkuart_pre_shaping_regsheet. shortname:{short_name}. field:{entry['Field']}. type:{entry['Type']}" ))

    return copy_reg_sheet_dictlist

def infineon_binary_2_hex(match):
    binary = match.group(1)  # 提取匹配到的二进制数
    decimal_value = int(binary, 2)
    hex_value = hex(decimal_value)
    #hex_value_without_0x = hex_value[2:]  # 去掉前缀'0x'
    #hex_str = f'\n{hex_value} ({hex_value_without_0x}):'  # 添加前缀
    return hex_value

def infineon_pre_shaping_regsheet(short_name,reg_sheet_dictlist):
    copy_reg_sheet_dictlist = copy.deepcopy(reg_sheet_dictlist)

    if copy_reg_sheet_dictlist is None:
        logger.info("reg_sheet_dictlist is None, cannot iterate over it.")
        return None

    for item in copy_reg_sheet_dictlist:
        for key, value in item.items():
            if isinstance(value, int) or isinstance(value, float):  # 检查是否为数字
                item[key] = str(value)  # 将数字转换为字符串
            if key in ['Field', 'Bits']:  # 去除换行符
                item[key] = item[key].replace('\n','')
              
    #处理Field为特殊字符串类似'CEViNP (i=0-7)'的寄存器表格，将其拆为CEV0NP-CEV7NP            
    #处理Field为特殊字符串类似'AGSRy (y=0-3)'的寄存器表格，将其拆为AGSR0-AGSR3
    #处理Field为特殊字符串类似'ENx (x=0-31)'的寄存器表格，将其拆为EN0-EN31
    #g_variable_name_set = ['n', 'i', 'x', 'y']

    # 遍历所有的标志，并处理包含它们的寄存器表格
    for flag in g_variable_name_set:
        if contains_character(copy_reg_sheet_dictlist, f'{flag}='):
            copy_reg_sheet_dictlist = split_expression_bits_table(copy_reg_sheet_dictlist, flag)
        shaped_register_dictlist = copy_reg_sheet_dictlist

    #格式化处理bits字段。"0" "1" "3" "2,31:4" 格式化为"[0:0]" "[1:1]" "[3:3]" "[0:2],[31:4]"
    #for row in range (this_sheet.shape[0]):
    for entry in shaped_register_dictlist:
        entry['Bits'] = parse_and_format_bits_in_bitfield(entry['Bits'])
    logger.debug(shaped_register_dictlist) 

    #如果是“[0:2],[31:4]”格式，需要继续格式化处理bits字段。
    #将格式"[0:2],[31:4]"，拆分为多行，并依据bits进行行排序。
    #这样的特殊格式，往往同时有如下特征：Field为特殊字符串'0'，Description包含'Reserved, write 0'关键字
    shaped_reserved_bits_register_dictlist = []
    for entry in shaped_register_dictlist:
        entry['Field'] = entry['Field'].replace(' ','')
        if '0' == entry['Field']:
            if ',' in entry['Bits']:  
                bits_list = entry['Bits'].split(',')  
                # 为每个拆分的Bits创建一个新行，并解析出[m:n]中的n ,同时将Field的值改为reserved+'n',保持其他列的值不变， 
                for bit in bits_list:  
                    logger.debug('bits list: %s'%(bit))
                    temp_entry = entry.copy()  
                    temp_entry['Bits'] = bit.strip()  
                    right_bit = parse_bits_mn(bit.strip())  
                    #temp_entry['BaseBit'] = right_bit  # 添加一个新列来存储[m:n]中的n
                    temp_entry['Field'] = 'reserved_'+str(right_bit)
                    shaped_reserved_bits_register_dictlist.append(temp_entry)  
            else:  
                #entry['BaseBit'] = parse_bits_mn(entry['Bits'])  # 对于非拆分行，直接解析并存储[m:n]中的n  
                entry['Field'] = 'reserved_'+str(parse_bits_mn(entry['Bits']))  # 对于reserver位，将Field值改为reserved+'n'  
                shaped_reserved_bits_register_dictlist.append(entry)
            continue
        pattern = r'([01]{1,4})B '
        #匹配到的位置前插入换行符，并保留0B/1B后的空格
        entry['Description'] = re.sub(pattern, r'\n\1B ', entry['Description'])

        # 在第一个匹配位置之前插入'\nValues:\n'
        match = re.search(pattern, entry['Description'], flags=re.MULTILINE)
        if match:
            start_index = match.start()
            entry['Description'] = entry['Description'][:start_index] + '\nValues:\n' + entry['Description'][start_index:]

        # 匹配到后，将匹配到的二进制数转为十六进制
        entry['Description'] = re.sub(pattern, infineon_binary_2_hex, entry['Description'], flags=re.MULTILINE)
        # 去除空行
        entry['Description'] = re.sub(r'\n\s*\n', '\n', entry['Description'])

        logger.info((f"shortname:{short_name}. field:{entry['Field']}. bits:{entry['Bits']}" ))
        shaped_reserved_bits_register_dictlist.append(entry)
    
    #shaped_reserved_bits_register_dictlist根据entry['Bits']排序
    shaped_reserved_bits_register_dictlist.sort(key=lambda x: parse_bits_mn(x['Bits']))    
    logger.info(f'{short_name:<12}:{shaped_reserved_bits_register_dictlist}')             

    return shaped_reserved_bits_register_dictlist

# 寄存器description风格转换函数。将nxp的寄存器desciption风格转换成arkuart的风格
def nxp_pre_shaping_regsheet(short_name,reg_sheet_dictlist):  
    shaped_register_dictlist = copy.deepcopy(reg_sheet_dictlist)
    shaped_reserved_bits_register_dictlist = []
    for entry in shaped_register_dictlist:
        logger.info(f'nxp_pre_shaping_regsheet:\n{entry}')
        text = entry['Description']
        pattern = r'^([01]{1,4})\s'
        # 在第一个匹配位置之前插入'\nValues:\n'
        match = re.search(pattern, text, flags=re.MULTILINE)
        if match:
            start_index = match.start()
            new_text = text[:start_index] + '\nValues:\n' + text[start_index:]
        else:
            new_text = text

        # 去掉空行
        new_text = re.sub(r'\n\s*\n', '\n', new_text)
        # 去掉首部换行符
        new_text = re.sub(r'^\n', '', new_text)
        entry['Description'] = new_text
        logger.debug(f'nxp_pre_shaping_regsheet: \n{new_text}')
        
        #以第一个换行符为分隔符，分为两部分.以防有多行
        value = entry['Field']
        part_list = value.split('\n',1)
        entry['Field'] = part_list[1].replace('\n', '').replace(' ', '').split('[')[0]  # nxp mpc577m奇葩字段名IE_NVSM[6:1]，去掉[6:1]
        bits_value = part_list[0]
        # m-n格式的bits值转为n:m格式
        if any(char in bits_value for char in ['–', '-']):
            part_mn = re.split(r'[-–]', bits_value)
            bits_value = f'[{int(part_mn[0])}:{int(part_mn[1])}]'
        else:
            bits_value = f'[{int(bits_value)}:{int(bits_value)}]'            
        entry['Bits'] = bits_value
        entry['Type'] = 'rw'
        
        if entry['Field']=='Reserved': # 如果字段名为Reserved
            match = re.search(r'\[.*?:(\d+)\]', bits_value)
            if match:
                m_value = match.group(1)  # 提取m的值
            else:
                m_value = 0  # 如果bits_value中没有m，则默认m为0
            entry['Field'] = 'reserved_'+str(m_value)
        
        # 将entry反序添加到shaped_reserved_bits_register_dictlist
        shaped_reserved_bits_register_dictlist.insert(0, entry)
    logger.info(f'{short_name:<12}:{shaped_reserved_bits_register_dictlist}')
    return shaped_reserved_bits_register_dictlist

def fill_long_name_in_reglist_table(ch_language=False):
    workbook = openpyxl.load_workbook(g_input_excel_file_path)
    colname_short_name = "Short Name"
    colname_long_name = "Long Name"
    if ch_language is True:
        colname_long_name = "Long Name Chinese"

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name != 'RegList':continue

        # 读取第一行以获取列名并初始化数据字典
        column_names = [cell.value for cell in sheet[1]]
        data_dict = {name: [] for name in column_names}  # 正确初始化数据字典

        # 填充数据字典
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
            for col_index, value in enumerate(row):
                column_name = column_names[col_index]
                data_dict[column_name].append(value)

        long_name_index = column_names.index(colname_long_name)

        for i, short_name in enumerate(data_dict[colname_short_name]):
            if 'reserved' in short_name.lower():continue
            if data_dict[colname_long_name][i] is not None:
                logger.debug(f'{short_name} is exist long name: {data_dict[colname_long_name][i]}')
                continue
            short_name = short_name.strip()
            value = g_treeview_dict[short_name][colname_long_name]

            cell = sheet.cell(row=i+2, column=long_name_index+1)
            cell.value = value
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='top')

    workbook.save(g_input_excel_file_path)
    workbook.close()

def fill_width_mask_in_reglist_table():
    workbook = openpyxl.load_workbook(g_input_excel_file_path)
    colname_bits = "Bits"
    colname_width = 'Width'
    colname_mask = 'Mask'

    sheet_name = 'RegList'
    # 获取表'RegList'的sheet,判断'RegList'是否存在，如果不存在则返回
    if 'RegList' not in workbook.sheetnames:
        logger.info(f'table : RegList not found.')
        # 关闭工作簿
        workbook.close()
        return
    sheet = workbook[sheet_name]
    # 读取第一行以获取列名并初始化数据字典
    column_names = [cell.value for cell in sheet[1]]
    data_dict = {name: [] for name in column_names}  # 正确初始化数据字典

    # 填充数据字典
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        for col_index, value in enumerate(row):
            column_name = column_names[col_index]
            data_dict[column_name].append(value)

    if colname_width not in data_dict:
        logger.error(f'Column "{colname_width}" not found in sheet "{sheet_name}"')

    width_index = column_names.index(colname_width)
    mask_index = column_names.index(colname_mask)

    for i, bits_value in enumerate(data_dict[colname_bits]):
        if 'reserved' in data_dict['Field'][i].lower():continue
        if data_dict[colname_width][i] is not None:continue

        logger.debug(f"fill_width_mask_in_reglist_table: {data_dict['Short Name'][i]}:{data_dict['Field'][i]}\n{bits_value}")
        width_value = calculate_bit_width(bits_value)
        mask_value = parse_bits_to_mask(bits_value)

        cell = sheet.cell(row=i+2, column=width_index+1)
        cell.value = width_value
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='top')

        cell = sheet.cell(row=i+2, column=mask_index+1)
        cell.value = hex(mask_value)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='top')

    # 保存文件 
    workbook.save(g_input_excel_file_path)
    workbook.close()

def fill_enum_item_in_reglist_table(sheet_name = 'RegList', colname_enumitem = 'EnumItem'):
    workbook = openpyxl.load_workbook(g_input_excel_file_path)
    colname_desc = "Description"

    # 判断sheet_name是否在workbook内
    if sheet_name not in workbook.sheetnames:
        logger.error(f'Sheet "{sheet_name}" not found.')
        workbook.close()
        return
    sheet = workbook[sheet_name]
    # 读取第一行以获取列名并初始化数据字典
    column_names = [cell.value for cell in sheet[1]]
    data_dict = {name: [] for name in column_names}  # 正确初始化数据字典

    # 填充数据字典
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        for col_index, value in enumerate(row):
            column_name = column_names[col_index]
            data_dict[column_name].append(value)

    if colname_desc not in data_dict:
        logger.error(f'Column "{colname_desc}" not found in sheet "{sheet_name}"')

    enum_item_index = column_names.index(colname_enumitem)

    for i, desc_value in enumerate(data_dict[colname_desc]):
        if pd.isnull(desc_value):continue
        if sheet_name == 'RegList' and 'reserved' in data_dict['Field'][i].lower():continue
        if data_dict[colname_enumitem][i] is not None:continue

        logger.info(f"fill_enum_item_in_reglist_table {sheet_name}: {desc_value}")
        value = description_2_enumitem_in_reglist_table(desc_value)

        cell = sheet.cell(row=i+2, column=enum_item_index+1)
        cell.value = value
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='top')

    # 通过名称激活第一个sheet
    # workbook.active = workbook['Baseinfo']   
    # 保存文件 
    workbook.save(g_input_excel_file_path)
    workbook.close()

def autogen_reg_list_excel(sheet_range=None):
    logger.info(f'Enter autogen_reg_list_excel\n')
    global g_bits_variable_dict,g_treeview_dict
    excel_data = pd.read_excel(g_input_excel_file_path, sheet_name=None)
    if 'BitsVariableSet' in excel_data:
        # 将excel_data['BitsVariableSet']表的列'Variable Name'做key，列'Value'作为key值，生成字典
        g_bits_variable_dict = excel_data['BitsVariableSet'].set_index('Variable Name')['Value'].to_dict()
        logger.info(g_bits_variable_dict)

    if 'TreeView' in excel_data:
        # 将excel_data['TreeView']表生成字典,key为行'Short Name',value为行表，行表中包含该行的每一列数据
        # 'Short Name'需要做个预处理，删除换行符和空格
        excel_data['TreeView']['Short Name'] = excel_data['TreeView']['Short Name'].str.replace('\n', '').str.replace(' ', '')
        g_treeview_dict = excel_data['TreeView'].set_index('Short Name').T.to_dict()
        logger.debug(g_treeview_dict)

    # 打开Excel文件
    workbook = openpyxl.load_workbook(g_input_excel_file_path)

    # 获取表'RegList'的sheet,判断'RegList'是否存在，如果不存在则返回
    if 'RegList' not in workbook.sheetnames:
        logger.info(f'autogen_reg_list_excel.table : RegList not found.')
        # 关闭工作簿
        workbook.close()
        return
    reg_list_sheet = workbook['RegList']
    # 如果reg_list_sheet除了表头，还有其他行，则返回，不做处理
    if reg_list_sheet.max_row > 1:
        logger.info(f'autogen_reg_list_excel.table : RegList has data, skip.')
        # 关闭工作簿
        workbook.close()
        return

    # 遍历g_treeview_dict中的每个sheet
    #not_lookup_sheet_list = ['Baseinfo','PinMap','BitsVariableSet','Functions','ArCfgInterfaces','ArTypeDefinitions','ArApis','RegApis','ArErrorCodes','ArDeviations','ArLimitations','DevEnum','DevStructure','DevFunctions','McalFunctions','ArFunctions','RegList','TreeView','RegSummary']
    for sheet_name in g_treeview_dict.keys():
        sheet = workbook[sheet_name]
        if sheet_range != None and sheet_name != sheet_range:continue
        logger.info(f'autogen_reg_list_excel.table :{sheet_name}')

        # 读取第一行以获取列名并初始化数据字典
        column_names = [cell.value for cell in sheet[1] if cell.value is not None and cell.value != ""]
        # 清洗列名,去掉换行符，去掉空格
        clean_column_names = [name.replace('\n', '').replace(' ', '') for name in column_names]
        data_dict = {name: [] for name in clean_column_names}  # 正确初始化数据字典

        # 填充数据字典
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
            for col_index, value in enumerate(row):
                if pd.isna(value):continue
                if sheet[1][col_index].value is None:
                    continue  # 如果列名为空，跳过该列
                if sheet[1][col_index].value.replace('\n', '').replace(' ', '') not in clean_column_names:
                    continue  # 如果清洗后的列名不在clean_column_names，跳过该列
                column_name = clean_column_names[col_index]                
                # 如果value是字符串，则进行.strip()
                if isinstance(value, str):
                    value = value.strip().replace('■','')
                data_dict[column_name].append(value)

        # data_dict由列表字典，转成字典列表
        source_sheet_data_list = []
        for index, value_list in enumerate(zip(*data_dict.values())):
            source_sheet_data_list.append({clean_name: str(value) for clean_name, value in zip(clean_column_names, value_list)})

        # 将列表source_sheet_data_list反序
        if g_register_style == 'arkuart':
            # 如果寄存器表的顺序是31 30...-->0，则反序
            #if source_sheet_data_list_reverse_flag(source_sheet_data_list):
            source_sheet_data_list.reverse()
            logger.error(f'autogen_reg_list_excel.2 :{source_sheet_data_list}')
            source_sheet_data_list = arkuart_pre_shaping_regsheet(sheet_name,source_sheet_data_list)

        # infineon风格，补全列表source_sheet_data_list：添加reserved行，或者拆分变量字段为多行
        if g_register_style == 'infineon':
            source_sheet_data_list = infineon_pre_shaping_regsheet(sheet_name,source_sheet_data_list)

        # nxp风格，shaping列表source_sheet_data_list
        if g_register_style == 'nxp':
            source_sheet_data_list = nxp_pre_shaping_regsheet(sheet_name,source_sheet_data_list)
            
        for entry in source_sheet_data_list:
                logger.error(entry)
                process_and_update_to_reg_list_sheet(entry, reg_list_sheet, sheet_name)

    # 保存工作簿
    workbook.save(g_input_excel_file_path)
    workbook.close()

def process_and_update_to_reg_list_sheet(entry, reg_list_sheet, sheet_name):
    """
    处理entry中的数据并更新到reg_list_sheet中。
    """
    row_data = []  # 临时存储当前行的非空数据

    # 假设sheet的第一行包含列名，生成列名到列索引的映射
    first_row = reg_list_sheet[1]
    target_col_name_to_index_map = {cell.value: cell.column for cell in first_row}
    logger.info(target_col_name_to_index_map)
           
    # 遍历词典
    for key,value in entry.items():
        # 如果key不是target_col_name_to_index_map的key，则跳过
        if key not in target_col_name_to_index_map:
            continue
        if value is None:
            continue
        # 确保value是字符串类型
        if not isinstance(value, str):
            # 如果value不是字符串，则将其转换为字符串
            value = str(value)

        logger.error(f'process_and_update_to_reg_list_sheet------{key},{value}')

        target_column_index = target_col_name_to_index_map[key]
        # 这里可以存储为元组(value, target_column_index)，如果需要在后面使用索引
        row_data.append((value, target_column_index))
    
    # 如果row_data有数据，将其添加到reg_list_sheet中
    if row_data:
        row_data.append((sheet_name, 1)) # 添加表名到reg_list_sheet第一列中
        logger.info(row_data)
        # 找到reg_list_sheet中下一个可用的行号  
        next_row = reg_list_sheet.max_row + 1
        
        # 根据目标列的索引，设置reg_list_sheet的单元格值
        for value, target_col_index in row_data:
            # 定位目标单元格
            cell = reg_list_sheet.cell(row=next_row, column=target_col_index)                    
            # 仅设置单元格的值，不改变其他属性（如字体、填充、对齐等）
            cell.value = value        
            # 设置单元格自动换行
            cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)

#----------------------------
        # lib functions
#----------------------------
def print_app_info():
    print("Usage:")
    print("-h\n\thelp")
    print("-v\n\tversion")
    print("-f <filename>.xlsx\n\tautogen register header files")
    #print("-t <filename>.xlsx\n\tautogen register header files; and translate excel file to chinese")
    print("autocoder version : %s by yuwenfeng 2024"%(g_version))

# pstr = 'YUxDDiZZyTTx' numbers=[3,2,6,9].  （i,x,y）为特征字符集合，注意这个集合是无序的，
# 请按照特征字符集合在字符串pstr中的出现顺序，依次替换成numbers中的数字，第一次出现特征字集合的时候，替换为numbers的第一个数字；第n次出现特征字集合的时候，替换为numbers的第n个数字。
# 预期结果为YU3DD2ZZ6TT9
def replace_features_with_numbers(pstr, numbers):  
    # 特征字符集合  
    # 全局变量名作为特征字符，初始化时候赋值。g_variable_name_set = {'i', 'x', 'y'}  
    # 初始化计数器，用于追踪feature_set的出现次数  
    feature_count = 0  
      
    # 初始化结果字符串  
    result = []  
      
    # 遍历字符串中的每个字符  
    for char in pstr:  
        # 如果字符是特征字符集合中的一个  
        if char in g_variable_name_set:  
            # 增加feature_set的出现次数  
            feature_count += 1  
            # 使用feature_set的出现次数减1作为索引（因为索引从0开始）  
            index = feature_count - 1  
            # 确保索引不会超出numbers列表的长度  
            index = index % len(numbers)  
            # 替换为特征数字  
            result.append(str(numbers[index]))  
        else:  
            # 如果字符不是特征字符集合中的一个，则直接添加到结果字符串中  
            result.append(char)  
      
    # 将结果列表转换为字符串并返回  
    return ''.join(result)              

def calculate_bit_width(input_str):  
    # 假设输入字符串的格式是"[a:b],[c:d],..."  
    ranges_str = input_str.strip('[]').split('],[')  # 去除外部的方括号，并根据'],['分割字符串  
    bit_widths = []    
    for range_str in ranges_str:  
        # 去除每个范围字符串两侧的空白字符  
        range_str = range_str.strip()            
        # 检查范围字符串是否为空  
        if not range_str:  
            continue    
        # 分割起始和结束数字  
        start_end = range_str.split(':')  
        # 根据冒号的数量判断是单个数字还是一个范围  
        if len(start_end) == 1:  
            # 单个数字，位宽为1  
            bit_width = 1  
        else:  
            # 一个范围，计算位宽  
            start, end = map(int, start_end)  
            bit_width = start - end + 1    
        # 将计算出的位宽添加到列表中  
        bit_widths.append(bit_width)   
    return bit_widths[0]  
  
def parse_and_format_bits_in_bitfield(input_str): 
    logger.debug(f'Enter function {parse_and_format_bits_in_bitfield.__name__}. Parameter input_str:{input_str}')

    # 检查输入字符串是否已经是正确的格式
    if all(part.startswith('[') and part.endswith(']') for part in input_str.split(',')):
        return input_str

    # 分割字符串，同时考虑到换行符和逗号  
    parts = input_str.replace('\n', ',').split(',')        
    # 用于存储格式化后的字符串部分  
    formatted_parts = []        
    # 遍历每个部分  
    for part in parts:  
        # 去除部分两侧可能存在的空白字符  
        part = part.strip()            
        # 检查部分是否为空  
        if not part:  
            continue            
        # 检查部分是否包含冒号  
        if ':' in part:  
            # 如果包含冒号，则直接将其用方括号包围  
            formatted_part = '[' + part + ']'  
        else:  
            # 如果不包含冒号，则将其重复并用冒号分隔，再用方括号包围  
            formatted_part = '[' + part + ':' + part + ']'            
        # 将格式化后的部分添加到列表中  
        formatted_parts.append(formatted_part)        
    # 将格式化后的部分用逗号连接成最终字符串  
    output_str = ','.join(formatted_parts)  
    return output_str  
  
# 解析Bits字段[m:n]并返回n数值  
def parse_bits_mn(bit_str): 
    match = re.match(r'\[(\d+):(\d+)\]', bit_str.strip())  
    if match:  
        return int(match.group(2))  # 返回n  
    else:  
        return None  
    
#计算bit[m:n]对应的掩码，掩码的计算方法是将bit n到bit m都置为1，其他bit置为0。
def parse_bits_to_mask(bit_str): 
    match = re.match(r'\[(\d+):(\d+)\]', bit_str.strip())
    m = int(match.group(1))
    n = int(match.group(2))
    mask = 0
    for i in range(0,m-n+1):
        mask = mask | (1<<i)
    
    return mask

# 将三元运算符转换为python的if-else语句
# 输入'(0==1) ? \"23\" :\"24\"' 输出 '23 if (0==1) else 24'
def convert_ternary_to_python(ternary_str):
    # 如果字符串被引号包着，则去除引号，否则保持不变
    if ternary_str.startswith('"') and ternary_str.endswith('"'):
        ternary_str = ternary_str[1:-1]  # 去除成对的引号
    ternary_str = ternary_str.replace(r'\"', '"').replace(r"\'", "'")
    logger.debug(f'convert_ternary_to_python输入: {ternary_str}')
    
    # 正则表达式匹配三元运算符，考虑不同引号类型和空格，同时处理转义字符
    ternary_pattern = r'\s*\((.*?)\)\s*?\?\s*(["\'])(0x[\dA-Fa-f]+|\d+|.*?)\2\s*:\s*(["\'])(0x[\dA-Fa-f]+|\d+|.*?)\4'
    
    def replace_ternary(match):
        condition = match.group(1).strip()
        true_value = match.group(3).strip()
        false_value = match.group(5).strip()
        # 转换为Python的if-else语句
        ifelse_str = f'{true_value} if {condition} else {false_value}'
        return ifelse_str    
    try:
        # 使用正则表达式替换所有的三元运算符
        python_str = re.sub(ternary_pattern, replace_ternary, ternary_str)
        logger.debug(f'convert_ternary_to_python输出: {python_str}')
        return str(eval(python_str))
    except re.error as e:
        logger.error(f'正则表达式匹配错误: {e}')
        return None

def arkuart_style_bits_range_variable_xy_to_value(text):
    logger.info(f'arkuart_style_bits_range_variable_xy_to_value 输入: {text}')
    # 使用正则表达式匹配并拆分加法和减法部分
    add_sub_pattern = r'([+-])'
    parts = re.split(add_sub_pattern, text.strip())

    # 过滤掉空字符串，并处理每一部分
    cleaned_parts = [part.strip() for part in parts if part.strip()]

    # 区分加法和减法部分
    add_parts = []
    sub_parts = []
    subtraction_flag = False
    for part in cleaned_parts:
        if part == '+':
            subtraction_flag = False
        elif part == '-':
            subtraction_flag = True
        else:
            if subtraction_flag:
                sub_parts.append(part)
            else:
                add_parts.append(part)

    # 将三元运算符转换为Python代码
    python_add_expr_parts = [convert_ternary_to_python(part) for part in add_parts]
    python_add_expr = ' + '.join(python_add_expr_parts)

    # 构建最终的Python表达式
    if sub_parts:
        python_sub_expr = sub_parts[0]  # 假设只有一个减法部分
        python_expr = f'({python_add_expr}) - {python_sub_expr}'
    else:
        python_expr = python_add_expr
    try:
        # 使用 ast.literal_eval 代替 eval 以增加安全性
        result = eval(python_expr)
    except (ValueError, SyntaxError, TypeError) as e:
        # 处理表达式解析异常
        logger.error(f"解析表达式出错: {e}")
        result = None

    logger.info(f'arkuart_style_bits_range_variable_xy_to_value 输出: {python_expr} = {result}')
    return int(result)

# 将arkuart风格中Bits变量x.y转换为数字
def arkuart_style_bits_xy_2_number(entry,shortname):
    # 处理arkuart的entry['Bits']中的变量x,y
    if not g_bits_variable_dict : # 字典没有被添加键值，则采用缺省值取替代bits中的xy变量
        logger.error(f"'BitsVariableSet' table not found in excel_data")
    else:                         # 如果有专门的表'BitsVariableSet'来定义bits中的xy变量，则使用这个表
        # 如果Range Variable[x]或者Range Variable[y]在entry['Description']中，则提取Range Variable[x]:或者Range Variable[y]:后的字符串，直到Range Variable[x]:或者Range Variable[y]:或者'Memory Access:'或者字符串结束为止
        #pattern = re.compile(r'.*Range Variable\[([xy])\]:(.*)', re.IGNORECASE)
        pattern = re.compile(r'Range Variable\[([xy])\]:((?:(?!Range Variable\[[xy]\]:|Memory Access:).)*)', re.IGNORECASE | re.DOTALL)
        # 提取entry['Description']中匹配的字符串
        # 使用 finditer 遍历所有匹配项
        origin_bits = str(entry['Bits'])
        if 'x' in origin_bits or 'y' in origin_bits:
            logger.warning((f"shortname:{shortname}. field:{entry['Field']}. bits:{origin_bits}. -->" ))
        for match in pattern.finditer(entry['Description']):
            # 如果找到匹配项，提取并打印匹配后的字符串部分（即冒号后面的部分）
            xy_name = match.group(1).strip()
            extracted_text = match.group(2).strip()  # 使用 strip() 去除前后的空白字符
            #去掉换行符
            extracted_text = extracted_text.replace('\n','')
            for key, value in g_bits_variable_dict.items():
                if pd.isna(key):continue
                if pd.isna(value):continue
                extracted_text = extracted_text.replace(key, str(value))
            bits_value = arkuart_style_bits_range_variable_xy_to_value(extracted_text)
            if xy_name in {'x','y'}:
                logger.warning((f"{xy_name}:{bits_value}" ))
                entry['Bits'] = entry['Bits'].replace(xy_name,str(bits_value))
      
# 处理excel中的'Memory Access'类型变量
def arkuart_style_type_variable_2_rw(entry,shortname):
    # 处理arkuart的entry['Description']中Memory Access:的变量
    if g_bits_variable_dict : 
        # 定义正则表达式，匹配'Memory Access:'后的所有内容
        pattern = re.compile(r'Memory Access:(.*)$', re.DOTALL)
        for match in pattern.finditer(entry['Description']):
            # 如果找到匹配项，提取并打印匹配后的字符串部分（即冒号后面的部分）
            extracted_text = match.group(1).strip()  # 使用 strip() 去除前后的空白字符
            #去掉换行符
            extracted_text = extracted_text.replace('\n','')
            for key, value in g_bits_variable_dict.items():
                logger.debug(f'arkuart_style_type_variable_2_rw key:{key},value:{value}')
                if pd.isna(key):continue
                extracted_text = extracted_text.replace(key.replace(' ',''), str(value))
            type_value = arkuart_style_memroy_access_variable_to_type_value(extracted_text)
            if type_value:
                logger.info((f"arkuart_style_type_variable_2_rw. shortname:{shortname}. field:{entry['Field']}. Type --> {type_value}" ))
                return type_value
    return None

# 将'(0==1 and 32!=0) ? "read-write" : "read-only"' 计算条件(0==1 && 32!=0)后，返回read-write或者read-only
def convert_to_python(express_str):     
    logger.info(f'convert_to_python 输入: {express_str}')   
      
    # 正则表达式匹配三元运算符  
    ternary_pattern = r'\((.*?)\)\s*?\?\s*(["\'])(.*?)\2\s*:\s*(["\'])(.*?)\4'
    match = re.match(ternary_pattern, express_str)
    if match:      
        condition_str, true_quote, true_value, false_quote, false_value = match.groups()  
        # 去除true_value和false_value两边的引号  
        true_value = true_value.strip(true_quote)  
        false_value = false_value.strip(false_quote)  
          
        # 这里要确保condition_str只包含安全的Python表达式  
        try:  
            condition = eval(condition_str)  
        except Exception as e:  
            logger.error(f'条件表达式计算失败: {e}')  
            return None  
          
        # 计算三元表达式的值  
        result = true_value if condition else false_value  
        return result.replace(' ','')  
    else:      
        logger.warning(f'未找到三元表达式在: {express_str}')  
        return None  

def arkuart_style_memroy_access_variable_to_type_value(text):
    # 如果字符串被引号包着，则去除引号，否则保持不变
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]  # 去除成对的引号
    if text.startswith('{') and text.endswith('}'):
        text = text[1:-1]  # 去除成对的大括号号    
    text = text.replace(r'\"', '"').replace(r"\'", "'")
    # 替换&&为and，因为Python使用and而不是&&  
    text = text.replace('&&', 'and')  
    logger.info(f'arkuart_style_memroy_access_variable_to_type_value 输入: {text}')

    python_expr = convert_to_python(text)
    # 判断是否为None，如果为None，则说明解析失败
    if python_expr is None:
        logger.error(f'ternary_to_python解析失败1: {text}')
        return None
    # 判断是否为字符串，如果不是字符串，则说明解析失败
    if not isinstance(python_expr, str):
        logger.error(f'ternary_to_python解析失败2: {text}')
        return None
    # 判断字符串是否为read-write或者read-only
    if python_expr not in ['read-write', 'read-only', 'write-only']:
        logger.error(f'ternary_to_python解析失败3: {text}')
        return None
    # 判断字符串是否为read-write，如果是则返回1，否则返回0
    if python_expr == 'read-write':
        python_expr = 'R/W'
    elif python_expr == 'write-only':
        python_expr = 'W'
    elif python_expr == 'read-only':   
        python_expr = 'R'
    else:
        python_expr = None

    logger.info(f'arkuart_style_memroy_access_variable_to_type_value 输出: {python_expr} ')
    return python_expr

# 定义一个函数pad_single_digit_numbers，输入一个字符串s，返回一个字符串
def pinmap_pad_single_digit_numbers(s: str) -> str:
    # 使用正则表达式找到末尾的一位数字，并替换为两位数字（后面补9）
    return re.sub(r'(?<!\d)\d(?!\d)$', r'\g<0>9', s)

def pinmap_custom_sort_key(item) -> str:
    # 对Module中的一位数字进行补零处理，并返回处理后的字符串
    padded_module = pinmap_pad_single_digit_numbers(item['Module'])
    return padded_module

def shaping_pinmap_datalist(pinmap_datalist):
    logger.info('Enter function %s' % shaping_pinmap_datalist.__name__)
    for entry in pinmap_datalist:
        logger.debug(entry)    

    all_sheet_dict_list = merge_similar_rows(pinmap_datalist, g_product_prefix,g_module_name)

    #先row['Direction']排序，然后在取值相同的row['Direction']内以row['Module']对all_sheet_dict_list进行排序
    from itertools import groupby
    # 首先，根据 'Port Type' 排序
    all_sheet_dict_list.sort(key=lambda x: x['Port Type'], reverse=True)
    # 然后，对每个 'Port Type' 分组内的元素按 'Module' 排序
    all_sheet_dict_list_sorted = []
    for key, group in groupby(all_sheet_dict_list, key=lambda x: x['Port Type']):
        sorted_group = sorted(group, key=pinmap_custom_sort_key)
        all_sheet_dict_list_sorted.extend(sorted_group)

    for entry in all_sheet_dict_list_sorted:
        logger.debug(entry)    
    return all_sheet_dict_list_sorted

def infineon_get_source_pinmap_datalist():
    logger.info('Enter function %s' % infineon_get_source_pinmap_datalist.__name__)
    alias_set = set(f'{g_alias.upper()}_{module_no}' for module_no in g_module_no_set)
    # 检查g_module_no_set的状态，如果为空或仅包含一个元素0，则将g_alias添加到alias_set中
    if not g_module_no_set or (len(g_module_no_set) == 1 and 0 in g_module_no_set):
        alias_set.add(g_alias.upper())
    logger.info(f'alias_set:{alias_set}')

    # 打开excel表格，跳过前2行，因为第3行是表头
    excel_data = pd.read_excel(g_pinmap_filename, sheet_name=None, skiprows=2, header=0)

    target_datalist = []
    # 遍历所有sheet
    for sheet_name in excel_data:
        logger.error(sheet_name)
        source_dictlist = excel_data[sheet_name].to_dict(orient='records')
        # 如果第一行不是表头，也就是不含有'Ball'，则重新读取该sheet，跳过3行
        if 'Ball' not in source_dictlist[0]:
            #重新读取该sheet，跳过3行
            source_dictlist = pd.read_excel(g_pinmap_filename, sheet_name=sheet_name, skiprows=3, header=0).to_dict(orient='records')

        for i,entry in enumerate(source_dictlist):
            entry['Port'] = str(entry['Symbol'])
            if '.' in entry['Port'] : entry['Port'] = entry['Port'].replace('.','_')
            if '/' in entry['Port'] : entry['Port'] = entry['Port'].replace('/','_')
            if pd.isna(entry['Ball']):
                entry['Ball'] = source_dictlist[i-1]['Ball']
                entry['Port'] = source_dictlist[i-1]['Port']
            if pd.isna(entry['Ctrl.']):
                entry['Ctrl.'] = source_dictlist[i-1]['Ctrl.']
            if pd.isna(entry['Symbol']):continue
            # 使用any函数检查entry['Symbol']是否包含alias_set中的任何别名
            is_module_aliased = any(alias in entry['Symbol'].upper() for alias in alias_set)        
            # 如果entry['Symbol']没有包含任何别名，则跳过当前迭代
            if not is_module_aliased: continue
            

            logger.info(entry)
            if 'I' in entry['Ctrl.']:
                entry['Direction'] = 'IN'
            else:
                entry['Direction'] = 'OUT'

            if entry['Direction'] == 'OUT':
                entry['Port Type'] = f'{g_product_prefix}{g_module_name}_Out'
            elif entry['Direction'] == 'IN':
                entry['Port Type'] = f'{g_product_prefix}{g_module_name}_In'
            else:
                entry['Port Type'] = 'Error'

            logger.info(entry)

            entry['Description'] = sheet_name + '.Ball ' + entry['Ball']
            entry['Module'] = entry['Symbol'].replace(g_module_name.upper(),g_module_name)
            target_datalist.append(entry)

    shaped_datalist = shaping_pinmap_datalist(target_datalist)
    return shaped_datalist

def nxp_get_source_pinmap_datalist():
    logger.info('Enter function %s' % nxp_get_source_pinmap_datalist.__name__)
    alias_set = set(f'{g_alias.upper()}_{module_no}' for module_no in g_module_no_set)
    # 检查g_module_no_set的状态，如果为空或仅包含一个元素0，则将g_alias添加到alias_set中
    if not g_module_no_set or (len(g_module_no_set) == 1 and 0 in g_module_no_set):
        alias_set.add(g_alias.upper())
    logger.info(f'alias_set:{alias_set}')

    # 读取Excel文件，并指定列名所在的行（假设为第7行，索引为6，因为索引从0开始）
    excel_data = pd.read_excel(g_pinmap_filename, sheet_name='IO Signal Table', skiprows=6, header=1)
    source_dictlist = excel_data.to_dict(orient='records')
    target_dictlist = []
    for entry in source_dictlist:
        if pd.isna(entry['Port (JDP)']):continue
        # 使用any函数检查entry['Module']是否包含alias_set中的任何别名
        is_module_aliased = any(alias in entry['Module'].upper() for alias in alias_set)        
        # 如果entry['Module']没有包含任何别名，则跳过当前迭代
        if not is_module_aliased: continue

        entry['Port'] = entry['Port (JDP)'].replace('[','_').replace(']','')
        if 'I' in entry['Direction']:
            entry['Direction'] = 'IN'
        else:
            entry['Direction'] = 'OUT'

        if entry['Direction'] == 'OUT':
            entry['Port Type'] = f'{g_product_prefix}{g_module_name}_Out'
        elif entry['Direction'] == 'IN':
            entry['Port Type'] = f'{g_product_prefix}{g_module_name}_In'
        else:
            entry['Port Type'] = 'Error'

        logger.info(entry)
        target_dictlist.append(entry)
    
    shaped_datalist = shaping_pinmap_datalist(target_dictlist)
    return shaped_datalist

def set_source_pinmap_datalist_entry_to_pinmap_sheet(entry, pinmap_sheet):
    """
    处理entry中的数据并更新到pinmap_sheet中。
    """
    row_data = []  # 临时存储当前行的非空数据

    # 假设sheet的第一行包含列名，生成列名到列索引的映射
    first_row = pinmap_sheet[1]
    target_col_name_to_index_map = {cell.value: cell.column for cell in first_row}
    logger.info(target_col_name_to_index_map)
           
    # 遍历词典
    for key,value in entry.items():
        # 如果key不是target_col_name_to_index_map的key，则跳过
        if key not in target_col_name_to_index_map:
            continue
        if value is None:
            continue
        # 确保value是字符串类型
        if not isinstance(value, str):
            # 如果value不是字符串，则将其转换为字符串
            value = str(value)

        logger.error(f'set_source_pinmap_datalist_entry_to_pinmap_sheet------{key},{value}')

        target_column_index = target_col_name_to_index_map[key]
        # 这里可以存储为元组(value, target_column_index)，如果需要在后面使用索引
        row_data.append((value, target_column_index))
    
    # 如果row_data有数据，将其添加到pinmap_sheet中
    if row_data:
        logger.info(row_data)
        # 找到pinmap_sheet中下一个可用的行号  
        next_row = pinmap_sheet.max_row + 1
        
        # 根据目标列的索引，设置pinmap_sheet的单元格值
        for value, target_col_index in row_data:
            # 定位目标单元格
            cell = pinmap_sheet.cell(row=next_row, column=target_col_index)                    
            # 仅设置单元格的值，不改变其他属性（如字体、填充、对齐等）
            cell.value = value.replace('\n', '')       
            # 设置单元格自动换行
            cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)

def rebuild_pinmap_sheet_in_excel():
    logger.info(f'Enter rebuild_pinmap_sheet_in_excel\n')
    # 打开Excel文件
    workbook = openpyxl.load_workbook(g_input_excel_file_path)

    # 获取表'PinMap'的sheet,判断'PinMap'是否存在，如果不存在则返回
    if 'PinMap' not in workbook.sheetnames:
        logger.info(f'rebuild_pinmap_sheet_in_excel.table : PinMap not found.')
        # 关闭工作簿
        workbook.close()
        return
    pinmap_sheet = workbook['PinMap']
    # 如果pinmap_sheet除了表头，还有其他行，则返回，不做处理
    if pinmap_sheet.max_row > 1:
        logger.info(f'rebuild_pinmap_sheet_in_excel.table : RegList has data, skip.')
        # 关闭工作簿
        workbook.close()
        return

    source_sheet_data_list = []
    # 遍历原始pinmap总表的每个sheet，获取每个sheet的pinmap数据，返回datalist
    if g_register_style == 'infineon':
        source_sheet_data_list = infineon_get_source_pinmap_datalist()     
    if g_register_style == 'nxp':
        source_sheet_data_list = nxp_get_source_pinmap_datalist()     

    for entry in source_sheet_data_list:
        logger.info(entry)
        set_source_pinmap_datalist_entry_to_pinmap_sheet(entry, pinmap_sheet)

    # 保存工作簿
    workbook.save(g_input_excel_file_path)
    workbook.close()

#----------------------------
        # TreeNode class
#----------------------------
# 定义节点类
class TreeNode:
    def __init__(self, name, number, offset_start, size, short_name=None, long_name=None, parent=None):
        self.name = name
        self.number = number 
        self.offset_start = offset_start
        self.size = size
        self.short_name = short_name
        self.long_name = long_name
        self.reg_sheet_dictlist = [] #None  # 字典列表来存储Excel中的寄存器sheet        
        self.original_values    = None  # 保存所有属性的原始值

        self.children = []
        self.parent = parent

        if parent:
            parent.add_child(self)

    def add_child(self, child):
        self.children.append(child)

    def original_node_save(self):
        self.original_values = {'name':self.name, 'number':self.number}

    def traverse_and_save_original_values(self):
        self.original_node_save()  # 保存当前节点的原始值
        logger.info(self.original_values)
        for child in self.children:
            child.traverse_and_save_original_values()  # 递归遍历子节点，保存子节点的原始值

    # 添加一个方法来获取某个属性的原始值
    def get_original_value(self, attribute_name):
        return self.original_values.get(attribute_name, None)

    def insert_before(self, new_node, parent=None):
        # 增加了parent参数，默认为None，用于指定在哪个父节点的children中插入
        if not parent:
            parent = self.parent
        if parent:
            parent.children.insert(parent.children.index(self), new_node)
            new_node.parent = parent

    #获取跟节点
    def get_root(self):
        if self.parent:
            return self.parent.get_root()
        else:
            return self
        
    #获取最接近根节点的祖先
    def get_L1_node(self):
        # 假设get_root是一个无参数方法，返回根节点
        root_node = self.get_root()
        
        # 如果当前节点的父节点是根节点，则返回当前节点（意味着它是L1节点）
        # 否则，我们需要在父节点的上下文中继续寻找L1节点
        return self if self.parent == root_node else self.parent.get_L1_node()

    #定义一个函数，将根节点到当前节点的路径名用'_'连成一个字符串
    def get_path(self):
        path = ''
        while self:
            path = self.name + '_' + path if path else self.name
            self = self.parent
        return path or None  # 如果不存在父节点，返回None    
    
    #定义一个函数，将根节点到当前节点的原始值路径名用'_'连成一个字符串
    def get_original_value_path(self):
        if self.original_values is None:
            return None
        path = ''
        while self:
            path = self.original_values['name'] + '_' + path if path else self.original_values['name']
            self = self.parent
        return path or None  # 如果不存在父节点，返回None    

    def get_original_values_path(self):
        if self.original_values is None:
            return None
        
        # 从当前节点到根节点的名称列表
        names = [self.original_values['name']]
        current_node = self.parent
        while current_node:
            names.append(current_node.original_values['name'])
            current_node = current_node.parent
        
        # 使用join将名称列表转换为字符串，并反转字符串以得到正确的顺序
        path = '_'.join(reversed(names))
        return path if path else None

    def get_number_short_name(self):
        if self.original_values is None:
            return None
        
        # 从当前节点到L2节点(含L2)的number列表
        numbers = []
        if self.original_values['number']>1:
            numbers.append(self.name.replace(self.original_values['name'],''))
        L1_node = self.get_L1_node()
        current_node = self.parent
        while current_node!= L1_node:
            if current_node.original_values['number']>1:
                numbers.append(current_node.name.replace(current_node.original_values['name'],''))
            current_node = current_node.parent  

        #如果列表为空，说明没有variable，不用整形，直接返回short_name
        if not numbers :
            return self.short_name   
           
        # pstr = 'YUxDDiZZyTTx' numbers=[1,2,3,4].  x,y,i为特征字符，按照从左到右出场顺序，依次替换成numbers中的数字，结果为YU1DD2ZZ3TT4
        # 定义一个函数，将pstr中的特征字符替换成numbers中的数字
        merged_result = replace_features_with_numbers(self.short_name,numbers[::-1])
        logger.info(merged_result)
        return merged_result

    def get_L1_node_name(self):    
        l1_node = self.get_L1_node()
        if l1_node is not None:
            return l1_node.name
        else:
            # 这里可以返回一个默认值、抛出一个异常，或采取其他适当的行动
            return None

    def get_L1_node_number(self):    
        l1_node = self.get_L1_node()
        if l1_node is not None:
            return l1_node.name.replace(l1_node.original_values['name'],'')
        else:
            # 这里可以返回一个默认值、抛出一个异常，或采取其他适当的行动
            return None

    def get_absolute_address(self):
        if not self.parent:
            return self.offset_start
        #定义一个函数，获取根节点的offset_start
        def get_root(self):
            root = self
            while root.parent:
                root = root.parent
            return root.offset_start
        
        address = self.offset_start + get_root(self)
        return address

    # 获取节点的本地偏移地址结束地址
    def get_offset_end(self):
        return self.offset_start + self.number*self.size


    # 获取节点的本地偏移地址。计算方法为：当前节点的offset_start-父节点的offset_start
    def get_local_offset_address(self):
        if not self.parent:
            return 0
        return self.offset_start - self.parent.offset_start

    # 获取节点的尾部空间。计算方法为：父节点的(offset_start+size)减去当前节点的(offset_start+size)
    def get_tail_space(self):
        if not self.parent:
            return 0
        return self.parent.offset_start + self.parent.size - (self.offset_start + self.number*self.size)

    # 获取中间节点。排除根节点，排除叶子节点
    def get_non_leaf_nodes_list(self):
        result = []
        queue = deque([self])
        visited = set([id(self)])  # 使用id来标记已访问的节点，避免与根节点的重复比较
        while queue:
            current_node = queue.popleft()
            
            # 如果children属性不存在或为空列表，则认为该节点是叶子节点
            if not hasattr(current_node, 'children') or not current_node.children:
                continue
                
            # 将非叶子节点添加到结果列表中
            result.append(current_node)
            
            # 遍历子节点并将其添加到队列中
            for child in current_node.children:
                if id(child) not in visited:  # 确保不会重复访问节点
                    visited.add(id(child))
                    queue.append(child)

        # 排除根节点
        if result and id(self) == id(result[0]):
            result = result[1:]

        return result
    
    # 获取树的所有节点列表
    def get_all_nodes_list(self):
        result = []
        queue = deque([self])
        
        while queue:
            current_node = queue.popleft()
            result.append(current_node)
            
            # 检查children属性是否存在且不为None
            children = getattr(current_node, 'children', None)
            #将queue.append(child)更改为queue.extend(children)，以便一次性将所有子节点添加到队列中，而不是在循环中逐个添加
            if children:
                queue.extend(children) 
                
        return result     


    def get_non_root_nodes_list(self):
        result = []
        queue = deque([self])
        visited = set([id(self)])  # 使用id来标记已访问的节点，避免与根节点的重复比较
        while queue:
            current_node = queue.popleft()
                            
            # 将节点添加到结果列表中
            result.append(current_node)
            
            # 遍历子节点并将其添加到队列中
            for child in current_node.children:
                if id(child) not in visited:  # 确保不会重复访问节点
                    visited.add(id(child))
                    queue.append(child)

        # 排除根节点
        if result and id(self) == id(result[0]):
            result = result[1:]

        return result
    
    # 获取所有叶子节点
    def get_leaf_nodes(self, node):
        """
        获取给定节点的所有叶子节点。
        
        :param node: 树的节点
        :return: 包含所有叶子节点的列表
        """
        if node is None:
            return []
        
        if not getattr(node, 'children', None):  # 检查node是否有children属性，且children是否为空
            return [node]
        
        leaf_nodes = []
        for child in node.children:
            leaf_nodes.extend(self.get_leaf_nodes(child))  # 使用extend来避免额外的扁平化操作
        return leaf_nodes    

    def copy_subtree(self, new_parent,i,size):
        """递归地拷贝整个子树，并将其附加到新的父节点"""
        address_increase = i * size
        new_node = TreeNode(
            self.name,
            self.number,
            self.offset_start+address_increase,
            self.size,
            self.short_name,
            self.long_name,
            parent=new_parent
        )
        new_node.reg_sheet_dictlist = self.reg_sheet_dictlist
        new_node.original_values = self.original_values
        new_node.children = [child.copy_subtree(new_node,i,size) for child in self.children]
        return new_node

    def insert_spread_node(self):
        if self.number <= 1 or not self.parent or not self.parent.children:
            return
       
        for i in range(self.number):
            new_name = f"{self.name}{i}"
            new_node = TreeNode(
                new_name,
                1,
                self.offset_start + i * self.size,
                self.size,
                self.short_name,
                self.long_name,
                parent=self.parent
            )
            new_node.reg_sheet_dictlist = self.reg_sheet_dictlist
            new_node.original_values = self.original_values
            # 复制子树并赋值给新节点的children属性
            #new_node.children = 
            #address_increase = i * self.size
            [child.copy_subtree(new_node,i,self.size) for child in self.children]

        self.parent.children.remove(self)

    def gen_a_spread_tree(self):
        node = copy.deepcopy(self)

        # 获取L1节点，将L1节点的number改为1.这是因为使用spread tree的reg.h将number下标改为了offset参数，因此名称上不再需要数字下标
        node.children[0].number = 1

        #获取最长的从叶子到根的距离
        def get_longest_path(node):
            if not node.children:
                return 0
            max_length = 0
            for child in node.children:
                length = get_longest_path(child) + 1
                if length > max_length:
                    max_length = length
            return max_length
               
        for i in range(get_longest_path(node)):
            non_root_nodes = node.get_non_root_nodes_list()
                
            for current_node in non_root_nodes:
                current_node.insert_spread_node()
            
        non_root_nodes = node.get_non_root_nodes_list()
        for current_node in non_root_nodes:
            current_node.children.sort(key=lambda child: child.offset_start)
            
        return node
    
    def insert_reserved_node(self):
        logger.debug(f"Before insert, len(self.children) = {len(self.children)}")
        if not self.children:
            return  # 没有子节点，不需要插入新节点

        def generate_reserved_nodes():            
            for i in range(0, len(self.children)):  # 从第二个子节点开始，以避免IndexError
                if i==0 and self.children[0].get_local_offset_address()>0:
                    new_node = TreeNode(
                        f"reserved_0[{self.children[0].get_local_offset_address()}]",
                        1,  # reserved节点的number为1
                        self.offset_start,  # 调整offset_start
                        self.children[0].get_local_offset_address(),  # reserved节点的size
                        'reserved',
                        parent=self  # 显式设置parent为self
                        )
                    yield new_node  # 返回新节点
                    continue                                                                                                            

                current_child = self.children[i]
                previous_child = self.children[i - 1]
                previous_offset_end = previous_child.offset_start + previous_child.size * previous_child.number
                if previous_offset_end < current_child.offset_start:
                    new_node = TreeNode(
                        f"reserved_{(previous_child.get_offset_end()-self.offset_start):X}[{current_child.offset_start - previous_offset_end}]",
                        1,  # reserved节点的number为1
                        previous_offset_end,  # 调整offset_start
                        current_child.offset_start - previous_offset_end,  # reserved节点的size
                        'reserved',
                        parent=self  # 显式设置parent为self
                    )
                    logger.debug(f"After insert, len(self.children) = {len(self.children)}")
                    yield new_node  # 返回新节点
                
        def generate_tail_reserved_nodes():
            if not self.parent:
                return            
            for i in range(0, len(self.children)):  # 从第二个子节点开始，以避免IndexError              
                if i==len(self.children)-1 and self.children[i].get_tail_space()>0:
                    new_node = TreeNode(
                        f"reserved_{(self.children[i].get_offset_end()-self.offset_start):X}[{self.children[i].get_tail_space()}]",
                        1,  # reserved节点的number为1
                        self.children[i].offset_start+self.children[i].size*self.children[i].number,  # 调整offset_start
                        self.children[i].get_tail_space(),  # reserved节点的size
                        'reserved',
                        parent=self  # 显式设置parent为self
                        )
                    yield new_node  # 返回新节点                                                                                                            

        new_children = list(generate_reserved_nodes())
        self.children.sort(key=lambda child: child.offset_start)
        new_children = list(generate_tail_reserved_nodes())

    def gen_a_reserved_tree(self):
        node = copy.deepcopy(self)
        non_leaf_nodes = node.get_non_leaf_nodes_list()
        for current_node in non_leaf_nodes:
            current_node.insert_reserved_node()

        return node

    def print_reg_sheet(self):
        if self.reg_sheet_dictlist is None:
            logger.info("reg_sheet_dictlist is None, cannot iterate over it.")
            return
        if 'reserved' in self.name:
            return
        logger.info(f"current_node.name: {self.name}")
        for entry in self.reg_sheet_dictlist:
            logger.info(entry)

def shaping_g_functions_dictlist(functions_dictlist, register_exist = False):
    if functions_dictlist == []:
        return []
    current_row = functions_dictlist[0]
    row = {}
    shaped_functions_dictlist = []
    for i in range(len(functions_dictlist)):
        current_row = functions_dictlist[i]
        #一个Function对应的多个union,register行，打包成一个列表，作为key 'Member List'的value
        if pd.isna(current_row['Function']):
            row['Member List'].append({'Union':current_row['Union'],'Register':current_row['Register'],'Field':current_row['Field'],'Value':current_row['Value'],'Notes':current_row['Notes']})
        else:
            logger.info(f"shaping_g_functions_dictlist row: {row}")
            if row : shaped_functions_dictlist.append(row)
            row = {}
            logger.debug(f"shaping_g_functions_dictlist current_row: {current_row}")
            row['Function'] = current_row['Function']
            row['Parameters'] = current_row['Parameters'].strip().replace('\n',' ')
            #row['Check Parameters'] = current_row['Check Parameters']
            # 如果row['Parameters']里面有两个或者多个连续空格，替换为一个空格
            row['Parameters'] = re.sub(r' {2,}', ' ', row['Parameters'])
            row['Function Description'] = current_row['Function Description']
            row['Function Description Chinese'] = current_row['Function Description Chinese']
            if 'Autosar Version' in current_row:
                row['Autosar Version'] = current_row['Autosar Version']
            if register_exist is True:
                row['Member List'] = []
                row['Member List'].append({'Union':current_row['Union'],'Register':current_row['Register'],'Field':current_row['Field'],'Value':current_row['Value'],'Notes':current_row['Notes']})
    if row : shaped_functions_dictlist.append(row)
    logger.info(f"shaping_g_functions_dictlist row: {row}")
    return shaped_functions_dictlist

def shaping_g_dev_structure_dictlist(dev_structure_dictlist):
    # 如果表单g_dev_structure_dictlist为空，则直接return
    if dev_structure_dictlist == []:
        return []
    current_row = dev_structure_dictlist[0]
    row = {}
    shaped_dev_structure_dictlist = []
    for i in range(len(dev_structure_dictlist)):
        current_row = dev_structure_dictlist[i]
        #一个Structure对应的多个Member行，打包成一个列表，作为key 'Member List'的value
        if pd.isna(current_row['Structure']):
            row['Member List'].append({'Member Type':current_row['Member Type'],'Member Name':current_row['Member Name'],'Member Description':current_row['Member Description'],'Member Description Chinese':current_row['Member Description Chinese']})
        else:
            if row : shaped_dev_structure_dictlist.append(row)
            row = {}
            logger.info(f"current_row: {current_row}")
            row['Structure'] = current_row['Structure']
            row['Structure Name'] = current_row['Structure Name']
            row['Structure Description'] = current_row['Structure Description']
            row['Structure Description Chinese'] = current_row['Structure Description Chinese']
            row['Member List'] = []
            row['Member List'].append({'Member Type':current_row['Member Type'],'Member Name':current_row['Member Name'],'Member Description':current_row['Member Description'],'Member Description Chinese':current_row['Member Description Chinese']})
    if row : shaped_dev_structure_dictlist.append(row)
    return shaped_dev_structure_dictlist

# 排除所有键的值都是空字符串或None的字典
def is_not_empty_row(row):
    return any(value for value in row.values() if value and value != '')

def read_global_setting_from_excel():
    # 读取Excel文件
    excel_data = pd.read_excel(g_input_excel_file_path, sheet_name=None)
    #将excel表单内容存入全局变量    
    global g_baseinfo_dict,g_functions_dictlist,g_dev_enum_dictlist,g_gen_file_path,g_register_style,g_source_code_root_path,g_pinmap_filename,g_alias
    global g_product_prefix,g_module_name,g_base_address,g_space_size,g_variable_name_set,g_module_no_set,g_bits_variable_dict,g_dev_functions_dictlist,g_dev_structure_dictlist
    global g_mcal_functions_dictlist,g_ar_functions_dictlist,g_mcal_structure_dictlist,g_ar_structure_dictlist,g_mcal_type_macro_dictlist,g_mcal_enum_dictlist,g_ar_enum_dictlist
    if 'Baseinfo' in excel_data:
        # 将excel_data['g_baseinfo_dict']表的列'Info Name'做key，列'Info Value'作为key值，生成字典
        g_baseinfo_dict = excel_data['Baseinfo'].set_index('Info Name')['Info Value'].to_dict()
    logger.info(g_baseinfo_dict)
    g_product_prefix = g_baseinfo_dict['Product Prefix']
    g_module_name = g_baseinfo_dict['Module Name']
    g_base_address = g_baseinfo_dict['Base Address']
    g_space_size = g_baseinfo_dict['Space Size']
    g_register_style = g_baseinfo_dict['Register Style']
    g_alias = g_baseinfo_dict['Alias']

    g_gen_file_path = os.path.join(g_source_code_root_path, g_module_name) + os.sep
    # 如果目录 g_gen_file_path 不存在，则创建
    if not os.path.exists(g_gen_file_path):
        os.makedirs(g_gen_file_path)
    
    pinmap_file = g_baseinfo_dict.get('PinMap File')
    # 首先表PinMap表必须存在, 然后判断文件pinmap_file是否在存在
    if ('PinMap' in excel_data) and (pinmap_file is not None and not pd.isna(pinmap_file)):
        if not os.path.isfile(pinmap_file):
            logger.error(f"pinmap_file:{pinmap_file} does not exist!")
        else:
            g_pinmap_filename = pinmap_file

    # 使用get方法从字典中获取'Module No Set'的值，如果键不存在，则返回空字符串
    module_no_str = g_baseinfo_dict.get('Module No Set', '')
    if pd.notna(module_no_str):
        module_no_str = str(module_no_str)
    #使用get方法从字典中获取'Module No Set'的值，如果键存在，但是内容为nan，则返回空字符串
    if module_no_str and pd.notna(module_no_str):
        g_module_no_set = set(module_no_str.split(','))    
    logger.info(g_baseinfo_dict)    
    logger.info(g_module_no_set)

    # 将'Reg Variable'对应的值按逗号分割，并转换为集合
    if 'Reg Variable' in g_baseinfo_dict and not pd.isna(g_baseinfo_dict['Reg Variable']):
        g_variable_name_set = set(g_baseinfo_dict['Reg Variable'].split(','))  # 先按逗号分割，再转换为集合
    else:
        # 如果'Variable Name'不存在于字典中，可以设置一个空集合或者抛出异常，取决于具体需求
        g_variable_name_set = set()

    if 'BitsVariableSet' in excel_data:
        # 将excel_data['BitsVariableSet']表的列'Variable Name'做key，列'Value'作为key值，生成字典
        g_bits_variable_dict = excel_data['BitsVariableSet'].set_index('Variable Name')['Value'].to_dict()

    # 获取'Functions'表单
    if 'Functions' in excel_data:
        functions_dictlist = excel_data['Functions'].to_dict(orient='records')
        functions_dictlist_filtered = [row for row in functions_dictlist if is_not_empty_row(row)]
        g_functions_dictlist = shaping_g_functions_dictlist(functions_dictlist_filtered, True)
    else:
        logger.error(f"'Functions' table not found in excel_data")
    # 获取'DevEnum'表单
    if 'DevEnum' in excel_data:
        g_dev_enum_dictlist = excel_data['DevEnum'].to_dict(orient='records')
        logger.debug(f'g_dev_enum_dictlist:{g_dev_enum_dictlist}')
    else:
        logger.error(f"'DevEnum' table not found in excel_data")

    # 获取'DevFunctions'表单
    if 'DevFunctions' in excel_data:
        dev_functions_dictlist = excel_data['DevFunctions'].to_dict(orient='records')
        dev_functions_dictlist_filtered = [row for row in dev_functions_dictlist if is_not_empty_row(row)]
        g_dev_functions_dictlist = shaping_g_functions_dictlist(dev_functions_dictlist_filtered)
    else:
        logger.error(f"'DevFunctions' table not found in excel_data")

    # 获取'DevStructure'表单
    if 'DevStructure' in excel_data:
        dev_structure_dictlist = excel_data['DevStructure'].to_dict(orient='records')
        g_dev_structure_dictlist = shaping_g_dev_structure_dictlist(dev_structure_dictlist)
    else:
        logger.error(f"'DevStructure' table not found in excel_data")
        
    # 获取'McalFunctions'表单
    if 'McalFunctions' in excel_data:
        mcal_functions_dictlist = excel_data['McalFunctions'].to_dict(orient='records')
        mcal_functions_dictlist_filtered = [row for row in mcal_functions_dictlist if is_not_empty_row(row)]
        g_mcal_functions_dictlist = shaping_g_functions_dictlist(mcal_functions_dictlist_filtered)
    else:
        logger.error(f"'McalFunctions' table not found in excel_data")

    # 获取'McalEnum'表单
    if 'McalEnum' in excel_data:
        g_mcal_enum_dictlist = excel_data['McalEnum'].to_dict(orient='records')
        logger.debug(f'g_mcal_enum_dictlist:{g_mcal_enum_dictlist}')
    else:
        logger.error(f"'McalEnum' table not found in excel_data")
        
    # 获取'McalTypeMacro'表单
    if 'McalTypeMacro' in excel_data:
        g_mcal_type_macro_dictlist = excel_data['McalTypeMacro'].to_dict(orient='records')
        logger.debug(f'g_mcal_type_macro_dictlist:{g_mcal_type_macro_dictlist}')
    else:
        logger.error(f"'McalTypeMacro' table not found in excel_data")
        
    # 获取'McalStructure'表单
    if 'McalStructure' in excel_data:
        mcal_structure_dictlist = excel_data['McalStructure'].to_dict(orient='records')
        g_mcal_structure_dictlist = shaping_g_dev_structure_dictlist(mcal_structure_dictlist)
    else:
        logger.error(f"'McalStructure' table not found in excel_data")

    # 获取'ArFunctions'表单
    if 'ArFunctions' in excel_data:
        ar_functions_dictlist = excel_data['ArFunctions'].to_dict(orient='records')
        ar_functions_dictlist_filtered = [row for row in ar_functions_dictlist if is_not_empty_row(row)]
        g_ar_functions_dictlist = shaping_g_functions_dictlist(ar_functions_dictlist_filtered)
    else:
        logger.error(f"'ArFunctions' table not found in excel_data")
                
    # 获取'ArEnum'表单
    if 'ArEnum' in excel_data:
        g_ar_enum_dictlist = excel_data['ArEnum'].to_dict(orient='records')
        logger.debug(f'g_ar_enum_dictlist:{g_ar_enum_dictlist}')
    else:
        logger.error(f"'ArEnum' table not found in excel_data")
        
    # 获取'ArStructure'表单
    if 'ArStructure' in excel_data:
        ar_structure_dictlist = excel_data['ArStructure'].to_dict(orient='records')
        g_ar_structure_dictlist = shaping_g_dev_structure_dictlist(ar_structure_dictlist)
    else:
        logger.error(f"'ArStructure' table not found in excel_data")

def get_longname_from_global_dict(short_name):
    key = 'Long Name'
    if g_trans_flag:
        key = 'Long Name Chinese'

    if short_name in g_reglist_shortname_key_dict:
        return g_reglist_shortname_key_dict[short_name][0][key]
    else:
        return None
    
def build_tree_from_excel():
    # 读取Excel文件
    excel_data = pd.read_excel(g_input_excel_file_path, sheet_name=None)
    global g_reglist_dictlist,g_reglist_shortname_key_dict

    # 获取'RegList'表单
    if 'RegList' in excel_data:
        g_reglist_dictlist = excel_data['RegList'].astype(str).to_dict(orient='records')
    else:
        logger.error(f"'RegList' table not found in excel_data")
        return []
        
    # 使用g_reglist_shortname_key_dict字典来组织相同short_name的寄存器信息
    for reg_info in g_reglist_dictlist:
        short_name = reg_info['Short Name']
        if short_name not in g_reglist_shortname_key_dict:
            g_reglist_shortname_key_dict[short_name] = []  # 初始化空列表来存储相同short_name的寄存器信息
        reg_info['Width'] = calculate_reglist_bit_width(reg_info['Bits'])
        g_reglist_shortname_key_dict[short_name].append(reg_info)  # 将寄存器信息添加到对应short_name的列表中
        
    # 获取数据列表,选择需要的列
    # 找出所有以'L'开头的列作为层级列
    level_columns = [col for col in excel_data['TreeView'].columns if col.startswith('L') and col[1:].isdigit()]
    level_columns.sort() # 确保按照L1, L2, L3...的顺序排序
    
    selected_columns = level_columns + ['Short Name', 'Long Name']
    # 如果存在'Long Name Chinese'列，则添加到selected_columns中
    if 'Long Name Chinese' in excel_data['TreeView']:
        selected_columns.append('Long Name Chinese')
    reg_summary_data = excel_data['TreeView'][selected_columns].to_dict(orient='records')

    # 创建根节点
    root = TreeNode(f'{g_product_prefix}', 1,int(g_base_address,16),int(g_space_size,16))
    nodes = {f'{root.name}': root}  # 根节点的键只用名称即可

    # 辅助函数：根据层级名称找到或创建节点
    def find_or_create_node(level_names, current_node):
        global g_l1_node_number
        parent_node = current_node
        path_parts = []  # 用于构建全路径的部分
        
        for level_name in level_names:
            if level_name and pd.notna(level_name):
                # 假设level_name只是一个字符串，而不是包含多个值的字符串
                name, number, offset_start, size = level_name.split(',')
                name = name.strip()
                if name.upper() == g_module_name.upper():
                    g_l1_node_number = int(number)
                short_name = f'{name} object'
                long_name = f'{name} object'
                
                # 添加当前节点名称到路径部分
                path_parts.append(name)
                
                # 创建一个完整路径作为唯一键
                node_path = '_'.join(path_parts)
                
                child_node = nodes.get(node_path)
                if not child_node:
                    child_node = TreeNode(name, int(number), int(offset_start, 16), int(size, 16), short_name, long_name, parent_node)
                    nodes[node_path] = child_node
                
                current_node = child_node
                parent_node = child_node
                
        return current_node

    # 遍历数据行构建树
    for row in reg_summary_data:
        level_names = [row.get(name) for name in level_columns if row.get(name) and pd.notna(row.get(name))]  # 过滤空名称
        short_name = row['Short Name'].replace('\n', '') 
        long_name = get_longname_from_global_dict(short_name)
        # 取倒数第一个非空level_name
        last_level_name = None
        for i in range(len(level_names)-1, -1, -1):
            if pd.notna(level_names[i]):
                last_level_name = level_names[i]
                break

        # 提取不包含最后一个非空level_name的非空部分
        non_empty_level_names = [name for name in level_names if name != last_level_name and pd.notna(name)]

        # 调用函数时使用正确的参数
        if last_level_name:  # 检查是否找到了非空level_name
            current_parent = find_or_create_node(non_empty_level_names, root)  # 传递非空level_name列表作为参数
        else:
            current_parent = root  # 如果未找到非空level_name，则使用根节点作为当前父节点   

        if last_level_name and pd.notna(last_level_name):
            # 解析字符串并提取信息
            name, number, offset_start, size = last_level_name.split(',')
            name = name.strip()
            
            # 构建完整路径
            path_parts = []
            temp_node = current_parent
            while temp_node:
                if temp_node != root:  # 不包括根节点
                    path_parts.insert(0, temp_node.name)
                temp_node = temp_node.parent
            
            path_parts.append(name)
            node_path = '_'.join(path_parts)
            
            # 检查节点是否已经存在
            if node_path in nodes:
                node = nodes[node_path]
            else:
                node = TreeNode(name, int(number), int(offset_start, 16), int(size, 16), short_name, long_name, current_parent)  # 将提取的信息赋给节点属性
                # 遍历g_reglist_dict，查找short_name对应的所有行，加入列表node.reg_sheet_dictlist  
                node.reg_sheet_dictlist = g_reglist_shortname_key_dict[short_name]          
                logger.info(f"node.reg_sheet_dictlist: {node.reg_sheet_dictlist}")
                nodes[node_path] = node
                
            current_parent = node  # 更新当前父节点为新创建的节点

    return root

# 打印树结构（递归方式）  
def print_tree(node, level=0):  
    logger.info('\t' * level + f"{node.name}:{node.number}:{node.offset_start:X}:{node.size:X}:{node.short_name or ''}:{node.original_values or ''}:{node.long_name or ''}")  # 打印叶子节点
    #logger.info('\t' * level + f"original values: { node.original_value}")  # 打印原始值
    if not node.children:
        return
    
    for child in node.children:  
        print_tree(child, level + 1)      

def print_absolute_address(node, level=0):  
    logger.info('\t' * level + f"{node.name}:{node.number}:{node.offset_start:X}:{node.get_absolute_address():X}")  # 打印叶子节点
    if not node.children:
        return
    
    for child in node.children:  
        print_absolute_address(child, level + 1)      

#----------------------------
        # application functions
#----------------------------
def get_copyright_str(filename):
    copyright_str = f"""/**
 *  <#>file {filename}
 *  <#>brief
 *  <#>copyright Copyright (c) 2024 {g_product_prefix} Technologies AG. All rights reserved.
 *
 */\n"""
    return copyright_str

#针对GxARBPR寄存器，根据变量i的取值范围，将存在变量i的Field和Bits列展开，并拆分为i+1行。拆分过程中，仅改变field和bits列的值，其他列值不变。
#举例如下：比如存在一个表
#Field	         Bits	    Type	Description
#PRIOi (i=0-2)	4*i+1:4*i	rw	    "Priority of Request Source i"
#根据变量i拆分完的效果如下
#Field	Bits	    Type	Description
#PRIO0	1:0	        rw	    "Priority of Request Source i"
#PRIO0	5:4	        rw	    "Priority of Request Source i"
#PRIO0	9:8	        rw	    "Priority of Request Source i"
# 原始数据表格
#rawdata = [  {'Field': 'PRIOi (i=0-2)', 'Bits': '4*i+1:4*i', 'Type': 'rw', 'Description': '"Priority of Request Source i'},
#          {'Field': 'CSMi (i=0-2)', 'Bits': '4*i+3', 'Type': 'rw', 'Description': 'Conversion Start Mode of Request Source i'},]

#解析格式字符串，比如4*i/y+1，并根据i/y的值，计算算式的值。这里:后面是可选内容
def parse_and_calculate_expression_in_bits(expression_string, i,expression_variable_flag): 
        #将i或者y替换为str(i) 
        expression = expression_string.replace(expression_variable_flag, str(i)).replace(' ','').replace('\n','')  
        # 使用eval计算表达式的值  
        try:  
            return str(eval(expression)) 
        except ZeroDivisionError:  
            return "Error: Division by zero"  
        except Exception as e:  
            return f"Error: {e}"  
  
#遍历输入数据表格中的每一行。对于每一行，使用正则表达式检查 Field 列是否包含格式化字符串 (i=0-2)或者(y=0-15)。如果匹配成功，从正则表达式中提取 i/y 的起始和结束值，并根据这个范围循环拆分 Field 和 Bits 列。
def split_expression_bits_table(register_dictlist,expression_variable_flag):
    logger.debug('Enter function %s.Parameter %s'%(split_expression_bits_table.__name__,expression_variable_flag))
    # 初始化结果列表
    result = []

    # 遍历原始数据表格中的每一行
    for entry in register_dictlist:
        #将entry中的数字转化为字符串
        entry['Field'] = str(entry['Field']).replace(' ','').replace('\n','')
        #根据格式(i=0-2)或者(y=0-15)，格式化提取i/y的取值范围.
        pattern = rf'({expression_variable_flag}=(\d+)-(\d+))'
        match_field = re.search(pattern, entry['Field'])
        #Field列需要拆分
        if match_field:
            start = int(match_field.group(2))
            end = int(match_field.group(3))
            logger.debug(f'start:{start},end:{end},entry:\n{entry}')
            # 根据i的取值范围拆分Field和Bits列
            for i in range(start,end+1):
                # 创建新的行数据，仅改变Field和Bits列的值
                new_entry = entry.copy()
                #field列，截断i后面的内容，保留i
                new_entry['Field'] = new_entry['Field'].replace(expression_variable_flag, str(i)).split('(')[0].replace(' ','')
                new_entry['Bits'] = new_entry['Bits'].replace(' ','').replace('\n','')
                new_entry['Description'] = new_entry['Description'].replace(' '+expression_variable_flag, ' '+str(i))
                #如果Bits列是格式字符串，比如4*i+1:4*i，则以冒号分界，拆分为前后两个部分。这里:后面是可选内容
                if ':' in new_entry['Bits']:
                    #Bits列是格式字符串，拆分为前后两部分
                    bits_split = new_entry['Bits'].split(':')
                    #计算前后两部分，并拼接
                    new_entry['Bits'] = parse_and_calculate_expression_in_bits(bits_split[0],i,expression_variable_flag) + ':' + parse_and_calculate_expression_in_bits(bits_split[1],i,expression_variable_flag)
                else:
                    # Bits列是普通字符串，直接计算i表达式的结果
                    new_entry['Bits'] = parse_and_calculate_expression_in_bits(new_entry['Bits'],i,expression_variable_flag)
                # 将新的行数据添加到结果列表中
                result.append(new_entry)
        else:
            # 如果不需要拆分，直接添加原始行数据到结果列表中
            result.append(entry)

    return result

#判断一个字符串列表中是否包含特定的字符
def contains_character(dictlist, character):  
    for entry in dictlist:  
        if character in str(entry['Field']): 
            return True  
    return False  

#----------------------------
        # common functions
#----------------------------
def get_short_description(in_desc):
    part1 = in_desc.split('\n')[0]   #取第一行   
    # 使用正则表达式匹配第一个句子（以中英文的,.;!?中任一标点结束）
    match = re.search(r'([^.!?。；！？]+(?=[.!?。；！？]))', part1)
    if match:  
        # 提取第一个句子的内容，不包括结束符号  
        part1 = match.group(1)  
    else:          
        # 如果没有找到匹配（即整个字符串只包含字母），则保持原样
        pass
    # 正则表达式，匹配以1或0开头（整个数字的长度为1、2或3），后紧跟字符'B'的字符串
    pattern = r'(?:([01]{1,4})B) '
    match = re.search(pattern, part1)

    if match:
        # 截取匹配之前的部分
        part1 = part1[:match.start()]

    logger.debug(f'get_short_description: \t{part1}') 
    return part1
    
def calculate_reglist_bit_width(input_str):  
    # 假设输入字符串的格式是"[a:b]" ，提取a和b的值
    match = re.search(r'(\d+):(\d+)', input_str)
    if match:
        end = int(match.group(1))  # 结束位
        start = int(match.group(2))    # 起始位
        bit_width = end - start + 1
    else:
        bit_width = 0
    return bit_width 

# get size,path from reserved_tree_leaf_list
def get_values_from_tree_leaf_list(shortname,tree_leaf_list):
    logger.info('Enter function %s' % get_values_from_tree_leaf_list.__name__)

    for current_node in tree_leaf_list:
        if shortname == current_node.short_name:
            break
        else:
            continue
    return current_node.size,current_node.get_path()

#autogen bit field struct
def autogen_regdef_h_bits_struct(filehandle,reserved_tree_leaf_list):
    logger.info('Enter function %s' % autogen_regdef_h_bits_struct.__name__)

    # short_name作为key遍历词典g_reglist_shortname_key_dict。value为列表
    for short_name,value in g_reglist_shortname_key_dict.items():
        short_name_size,short_name_path = get_values_from_tree_leaf_list(short_name,reserved_tree_leaf_list)
        if short_name_size != 4: continue   #寄存器size不是32，不生成B
        logger.error(f"short_name:{short_name} short_name_size:{short_name_size} short_name_path:{short_name_path}")
        treeview_register_description = ''
        if g_trans_flag is True:
            long_name = value[0]['Long Name Chinese']
            if 'Description Chinese' in g_treeview_dict[short_name] and not pd.isna(g_treeview_dict[short_name]['Description Chinese']):
                treeview_register_description = '\n'+g_treeview_dict[short_name]['Description Chinese']
                treeview_register_description = treeview_register_description.replace('\n','\n *  ')
        else:
            long_name = value[0]['Long Name']
            if 'Description' in g_treeview_dict[short_name] and not pd.isna(g_treeview_dict[short_name]['Description']):
                treeview_register_description = '\n'+g_treeview_dict[short_name]['Description']
                treeview_register_description = treeview_register_description.replace('\n','\n *  ')
        filehandle.write(f"/** <#>brief {long_name}({short_name}) {treeview_register_description}\n */\n")
        filehandle.write(f"typedef struct\n")
        filehandle.write("{\n")

        for entry in value:
            logger.info(entry)
            if g_reg_short_description is True:
                reg_description = get_short_description(entry['Description'])
            else:
                reg_description = entry['Description']
            # 如果字典entry存在'Description Chinese'，则使用中文描述
            if g_trans_enum_item_description is True and 'Description Chinese' in entry:
                reg_description = entry['Description Chinese'].split('存在:')[0] #get_short_description(entry['Description Chinese'])+' '+reg_short_description
            reg_description = reg_description.replace('\n',' ')
            # reg_description连续空格替换为单个空格
            reg_description = re.sub(r'\s+', ' ', reg_description)
            logger.info(f"short_description:{short_name}.{entry['Field']}.({reg_description})")
            type_len = 32
            if short_name_size == 1:
                type_len = 8       
            if short_name_size == 2:
                type_len = 16     
            filehandle.write(f"\t{g_product_prefix}_UReg_{type_len}Bit {entry['Field']:>16}:{entry['Width']};\t/** <#>brief {entry['Bits']:<8} ({entry['Type']}) {reg_description} */\n")

        filehandle.write(f"}}{short_name_path}_Bits;\n\n")
        filehandle.write("\n")

def autogen_regdef_h_union(filehandle,reserved_tree_leaf_list):
    logger.info('Enter function %s' % autogen_regdef_h_union.__name__)

    for current_node in reserved_tree_leaf_list:
        if 'reserved' in current_node.name:
            continue
        filehandle.write("/** <#>brief %s(%s) */\n"%(current_node.long_name,current_node.short_name))
        filehandle.write("typedef union\n")
        filehandle.write("{\n")
        type_len = 32
        if current_node.size == 1:
            type_len = 8         
        if current_node.size == 2:
            type_len = 16         
        filehandle.write(f"\t{g_product_prefix}_UReg_{type_len}Bit U;\t\t\t/** <#>brief Unsigned access */\n")
        filehandle.write(f"\t{g_product_prefix}_SReg_{type_len}Bit I;\t\t\t/** <#>brief Signed access */\n")
        if type_len == 32:
            filehandle.write("\t%s_Bits B;\t\t/** <#>brief Bitfield access */\n"%(current_node.get_path()))
        filehandle.write("}%s;\n\n"%(current_node.get_path()))

def autogen_typesreg_h():
    # autogen Lnx_TypesReg.h. Define the type of variable.
    common_path = os.path.join(g_source_code_root_path, 'common') + os.sep
    if not os.path.exists(common_path):
        os.makedirs(common_path)
    fo=open(f"{common_path}{g_product_prefix}_TypesReg.h","w",encoding='utf-8')
    fo.write("#ifndef %s_TYPESREG_H\n"%(g_product_prefix.upper()))
    fo.write("#define %s_TYPESREG_H\n"%(g_product_prefix.upper()))
    fo.write("/******************************************************************************/\n")
    fo.write("#include <stdio.h>\n")
    fo.write("#include <stdbool.h>\n")
    fo.write("/******************************************************************************/\n")
    fo.write("typedef unsigned char     %s_UReg_8Bit;\n"%(g_product_prefix))
    fo.write("typedef unsigned short    %s_UReg_16Bit;\n"%(g_product_prefix))
    fo.write("typedef unsigned int      %s_UReg_32Bit;\n"%(g_product_prefix))
    fo.write("typedef double            %s_DoubleReg_64Bit;\n"%(g_product_prefix))
    fo.write("typedef unsigned long long int    %s_UReg_64Bit;\n"%(g_product_prefix))
    fo.write("typedef signed char       %s_SReg_8Bit;\n"%(g_product_prefix))
    fo.write("typedef signed short      %s_SReg_16Bit;\n"%(g_product_prefix))
    fo.write("typedef signed int        %s_SReg_32Bit;\n"%(g_product_prefix))
    fo.write("typedef signed long long int      %s_SReg_64Bit;\n"%(g_product_prefix))
    fo.write(f"typedef unsigned long long int   Ark_TickTime;\t/* Time in ticks */\n")
    fo.write(f"typedef unsigned short           Ark_SizeT;\n")
    fo.write(f"typedef bool                     {g_product_prefix}_Bool;\n\n")
    fo.write(f"#define {g_product_prefix.upper()}_NULL_PTR ((void *)0)\n")
    fo.write(f"#define {g_product_prefix.upper()}_INVALID_VALUE 0xFFFFFFFFu\n")
    fo.write(f"#define {g_product_prefix.upper()}_INVALID_ADDR {g_product_prefix.upper()}_INVALID_VALUE\n\n")

    upper_product_prefix = g_product_prefix.upper()
    assert_str = """
#define {product_prefix_up}_INLINE         static inline
#define {product_prefix_up}_TIME_NULL      (({product_prefix}_TickTime)0x0000000000000000LL)
#define {product_prefix_up}_TIME_INFINITE  (({product_prefix}_TickTime)0x7FFFFFFFFFFFFFFFLL)
#define {product_prefix_up}_SIZET_MAX      (0x7FFFu)

#define {product_prefix_up}_CUSTOM_ASSERT
#define {product_prefix_up}_BKPT_ASM __asm("BKPT #0\\n\\t")

#if defined({product_prefix_up}_CUSTOM_ASSERT)
static inline void {product_prefix}Assert(volatile bool x,const char *function,int lineNum,const char *fileName)
{{         
  
    if(x) {{ }} else {{
      (void)printf("Fatal Error function:%s,Number of rows with errors:%d\\n",function,lineNum);
      (void)printf("File name:%s\\n",fileName);
      
      {product_prefix_up}_BKPT_ASM; for(;;) {{ }} }}
}}
#define {product_prefix_up}_ASSERT(x, y, z, h) {product_prefix}Assert(x, y, z, h)
#else
/* Assert macro does nothing */
#define {product_prefix_up}_ASSERT(x, y, z, h) ((void)0)
#endif""".format(product_prefix = g_product_prefix,product_prefix_up = upper_product_prefix)
    fo.write(f"\n{assert_str}\n\n")
    fo.write("/******************************************************************************/\n")

    fo.write("#endif /* %s_TYPESREG_H*/\n"%(g_product_prefix.upper()))
    fo.write("\n")
    fo.close()
    logger.debug("Successfully generated %s_TypesReg.h"%(g_product_prefix))

def get_dictlist_from_ar_excel_sheet(filename, sheetname, keywords, start_keyword, mline_keyword = 'Range'):
    global g_general_types_h_flag
    try:
        # 读取Excel文件，同时指定不将"NULL"字符串解释为NaN
        df = pd.read_excel(filename, sheet_name=sheetname, header=None, na_values=['', '#N/A', 'N/A', 'null', 'NaN'], keep_default_na=False)
    except (FileNotFoundError, ValueError) as e:
        logger.error(f"Error reading the Excel file: {e}")
        return []
    
    tables_data = []
    current_table = {}
    range_buffer = []  # 用于临时存储Range字段的多行数据
    in_range_section = False  # 标记是否正在处理Range字段的数据

    for index, row in df.iterrows():
        keyword = row.iloc[0]

        # 检查是否为关键字行
        if pd.notna(keyword) and keyword in keywords:
            # 如果之前在处理Range字段且当前行不是Range关键字行
            if in_range_section and keyword != mline_keyword:
                # 处理Range数据
                if mline_keyword not in current_table:
                    current_table[mline_keyword] = [range_buffer]  # 如果Range字段不存在，则创建并添加range_buffer作为子列表
                elif len(range_buffer) != 0:
                    for range_buffer_entry in range_buffer:
                        current_table[mline_keyword].append(range_buffer_entry)  # 如果Range字段已存在，则添加range_buffer的每一项作为新的子列表
                range_buffer = []  # 重置range_buffer
                in_range_section = False  # 停止处理Range字段
                
            # 处理新的关键字行
            if keyword == start_keyword:
                if current_table:
                    tables_data.append(current_table)
                current_table = {keyword: row.iloc[1:].dropna().tolist()[0] if row.iloc[1:].size > 0 else []}
                in_range_section = False  # 重置Range处理状态
            elif keyword == mline_keyword:
                # 在这个假设下，我们直接将每行Range数据作为一个子列表添加到Range字段中
                current_table[mline_keyword] = current_table.get(mline_keyword, []) + [row.iloc[1:].dropna().tolist()]
            elif keyword in ['Parameters (in)', 'Parameters (out)', 'Parameters (in- out)', 'Parameters (in\n- out)', 'Return']:
                # 读取当前行的数据（跳过第一列的关键字）
                current_table[keyword] = row.iloc[1:].dropna().tolist() if len(row.iloc[1:].dropna()) > 1 else row.iloc[1:].dropna().tolist()[0]
            elif keyword == 'Multiplicity':
                current_table[keyword] = str(row.iloc[3]).replace(' ', '').replace('\n', '')
            elif keyword == 'Default value':
                current_table[keyword] = str(row.iloc[1])                
                if current_table['Multiplicity'] == 'EcucIntegerParamDef':
                    if str(row.iloc[1]).upper() == 'FALSE': current_table[keyword] = '0'
                    if str(row.iloc[1]).upper() == 'TRUE' : current_table[keyword] = '1'
                if current_table['Multiplicity'] == 'EcucBooleanParamDef':
                    if str(row.iloc[1]) == '0' or str(row.iloc[1]).upper() == 'FALSE': current_table[keyword] = 'FALSE'
                    if str(row.iloc[1]) == '1' or str(row.iloc[1]).upper() == 'TRUE' : current_table[keyword] = 'TRUE'
            else:
                current_table[keyword] = str(row.iloc[1])
                if keyword == 'File' and '_GeneralTypes.h' in current_table[keyword]: g_general_types_h_flag = True

            # 如果当前关键字是Range，则标记为正在处理Range字段
            if keyword == mline_keyword:
                in_range_section = True

        # 如果当前行是Range数据行（基于上下文判断）
        elif in_range_section and pd.notna(row.iloc[1]):
            # 收集Range行的数据，移除NaN值，并添加到range_buffer
            range_buffer.append(row.iloc[1:].dropna().tolist())

    # 处理最后一个表格（如果有的话）
    if current_table:
        if in_range_section and range_buffer:
            if mline_keyword in current_table:
                current_table[mline_keyword].append(range_buffer)
            else:
                current_table[mline_keyword] = [range_buffer]
        tables_data.append(current_table)

    return tables_data

def get_g_ar_cfg_interfaces_dictlist():
    global g_ar_cfg_interfaces_dictlist
    logger.info(f'get_g_ar_cfg_interfaces_dictlist:')
    keywords = ['Name', 'Description', 'Multiplicity', 'Range', 'Default value', 'Post-build variant value', 'Value configuration class', 'Origin', 'Dependency']
    g_ar_cfg_interfaces_dictlist = get_dictlist_from_ar_excel_sheet(g_input_excel_file_path, 'ArCfgInterfaces', keywords, 'Name', 'None')
    for dict in g_ar_cfg_interfaces_dictlist:
        logger.info(dict)
        for key,value in dict.items():
            logger.info(f'{key}: {value}')

def get_g_ar_type_definitions_dictlist():
    global g_ar_type_definitions_dictlist
    logger.info(f'get_g_ar_type_definitions_dictlist:')
    keywords = ['Syntax', 'Type', 'File', 'Range', 'Description', 'Source']
    g_ar_type_definitions_dictlist = get_dictlist_from_ar_excel_sheet(g_input_excel_file_path, 'ArTypeDefinitions', keywords, 'Syntax', 'Range')
    for dict in g_ar_type_definitions_dictlist:
        logger.info(dict)
        for key,value in dict.items():
            logger.info(f'{key}: {value}')
            
def shaping_g_apis_dictlist(apis_dictlist):
    logger.info(f'shaping_g_apis_dictlist:')
    shaped_apis_dictlist = copy.deepcopy(apis_dictlist)
    for entry in shaped_apis_dictlist:
        function_syntax = entry['Syntax'].replace('\n','')
        logger.info(function_syntax)
        # function_syntax的左括号前面如果有空格，则去掉
        # 格式的模式为 void I2c_Init (const I2c_ConfigType* const ConfigPtr)
        if function_syntax.split('(')[0].endswith(' '):
            function_syntax = function_syntax.replace(' (','(')   
               
        return_str = entry['Return'][0]
        func_name = function_syntax.split('(')[0].replace(return_str, '').strip()  # 提取函数名
            
        # 提取function_syntax中小括号内的parameters
        parameters = function_syntax.split('(')[1].split(')')[0]
        shaped_parameters = ''
        last_words = []
        # 用逗号分割parameters成多个部分，遍历这几个部分
        # parameters = "const I2c_ChannelType ChannelId, I2c_DataType *const DataPtr, const I2c_SizeType Size, const I2c_SlaveAddrType SlaveAddress"
        # parameters = "void"
        for parameter in parameters.split(','):            
            parameter = parameter.strip()
            # 提取最后一个单词
            parts = parameter.split(' ')
            last_word = parts[-1]  # 提取最后一个单词
            last_words.append(last_word)
        shaped_parameters = ', '.join(last_words)
        shaped_parameters = shaped_parameters.replace('\n','')
        if shaped_parameters == 'void':
            shaped_parameters = ''
        else:
            shaped_parameters += ', '

        # entry 新增key-value对 'func_name','retval','shaped_parameters'
        entry['func_name'] = func_name
        entry['shaped_parameters'] = shaped_parameters
    return shaped_apis_dictlist
    
def get_g_apis_dictlist(apis_sheet_name):
    logger.info(f'get_g_apis_dictlist:')
    keywords = ['Syntax', 'Service ID', 'Sync/Async', 'ASIL level', 'Re-entrancy', 'Parameters (in)', 'Parameters (out)', 'Parameters (in\n- out)', 'Return', 'Description', 'Source', 'Error handling', 'Configuration dependencies', 'User hints']
    apis_dictlist = get_dictlist_from_ar_excel_sheet(g_input_excel_file_path, apis_sheet_name, keywords, 'Syntax', 'Parameters (in)')
    shaped_apis_dictlist = shaping_g_apis_dictlist(apis_dictlist)
    for dict in shaped_apis_dictlist:
        logger.info(dict)
        for key,value in dict.items():
            logger.info(f'{key}: {value}')
    return shaped_apis_dictlist

def get_g_ar_error_codes_dictlist():
    global g_ar_error_codes_dictlist
    logger.info(f'get_g_ar_error_codes_dictlist:')
    excel_data = pd.read_excel(g_input_excel_file_path, sheet_name=None)
    if 'ArErrorCodes' not in excel_data: return
    g_ar_error_codes_dictlist = excel_data['ArErrorCodes'].to_dict(orient='records')
    for dict in g_ar_error_codes_dictlist:
        logger.info(dict)
        for key,value in dict.items():
            logger.info(f'{key}: {value}')
    
def get_T_dict_from_excel_sheet(filename, sheetname, indexstr):
    excel_data = pd.read_excel(filename, sheet_name=None)
    if sheetname in excel_data:
        return_dict = excel_data[sheetname].astype(str).set_index(indexstr).T.to_dict()
        return return_dict
        
def get_g_sysinfo_dict():
    global g_sysinfo_dict
    g_sysinfo_dict = get_T_dict_from_excel_sheet(g_input_excel_file_path, 'Baseinfo', 'Info Name')
    logger.info(g_sysinfo_dict)
        
def get_factory_version_info():
    # 从g_sysinfo_dict中获取版本信息
    # 注意：g_sysinfo_dict的键是Info Name，值是包含Info Value的字典
    vendor_id = g_sysinfo_dict.get('Vendor ID', {}).get('Info Value', '1')
    module_id = g_sysinfo_dict.get('Module ID', {}).get('Info Value', '1')
    ar_release = g_sysinfo_dict.get('AUTOSAR Release', {}).get('Info Value', '4.4.0')
    sw_version = g_sysinfo_dict.get('Module SW Version', {}).get('Info Value', '1.0.0')
    logger.info(f'get_factory_version_info: {vendor_id}, {ar_release}, {module_id}, {sw_version}')
    ar_release_split = ar_release.split('.')
    ar_release_major = ar_release_split[0]
    ar_release_minor = ar_release_split[1]
    ar_release_revision = ar_release_split[2]
    sw_version_split = sw_version.split('.')
    sw_version_major = sw_version_split[0]
    sw_version_minor = sw_version_split[1]
    sw_version_patch = sw_version_split[2]
    
    upper_modulename = g_module_name.upper()
    versioninfo_str = """
/*==================================================================================================
*                              SOURCE FILE VERSION INFORMATION
==================================================================================================*/
#define {modulename_up}_C_VENDOR_ID                    {vendorid}U
#define {modulename_up}_C_MODULE_ID                    {moduleid}U
#define {modulename_up}_C_AR_RELEASE_MAJOR_VERSION     {arreleasemajor}U
#define {modulename_up}_C_AR_RELEASE_MINOR_VERSION     {arreleaseminor}U
#define {modulename_up}_C_AR_RELEASE_REVISION_VERSION  {arreleaserevision}U
#define {modulename_up}_C_SW_MAJOR_VERSION             {swversionmajor}U
#define {modulename_up}_C_SW_MINOR_VERSION             {swversionminor}U
#define {modulename_up}_C_SW_PATCH_VERSION             {swversionpatch}U

/*==================================================================================================
*                                     FILE VERSION CHECKS
==================================================================================================*/
""".format(modulename_up = upper_modulename,vendorid = vendor_id,moduleid = module_id,arreleasemajor = ar_release_major,
           arreleaseminor = ar_release_minor,arreleaserevision = ar_release_revision,swversionmajor = sw_version_major,swversionminor = sw_version_minor,swversionpatch = sw_version_patch)
    return f"{versioninfo_str}"

def autogen_cfg_h():
    # autogen Adc_Cfg.h. Define the version.
    fo=open(f"{g_gen_file_path}{g_module_name}_Cfg.h","w",encoding='utf-8')
    fo.write("#ifndef %s_CFG_H\n"%(g_module_name.upper()))
    fo.write("#define %s_CFG_H\n"%(g_module_name.upper()))

    dem_report_en_disable_macro = f'''
/* Configuration Options for DEM Options for the enabling/disabling of DEM in {g_module_name.upper()} Driver */
#define {g_module_name.upper()}_DISABLE_DEM_REPORT  (0U)
#define {g_module_name.upper()}_ENABLE_DEM_REPORT   (1U)
\n'''
    fo.write(dem_report_en_disable_macro)

    # 从g_sysinfo_dict中获取AUTOSAR Release和Module SW Version
    # 注意：g_sysinfo_dict的键是Info Name，值是包含Info Value的字典
    ar_release = g_sysinfo_dict.get('AUTOSAR Release', {}).get('Info Value', '4.4.0')
    sw_version = g_sysinfo_dict.get('Module SW Version', {}).get('Info Value', '1.0.0')
    logger.info(f'autogen_cfg_h: {ar_release}, {sw_version}')

    # 生成published_information macro
    published_information_str = pick_published_information_outof_ar_cfg_interfaces()
    fo.write(f"{published_information_str}\n")
    
    # 生成pre-compiler macro
    pre_compiler_macro_str = pick_pre_compile_outof_ar_cfg_interfaces()
    fo.write(f"{pre_compiler_macro_str}\n")
        
    fo.write("\n#endif /* %s_CFG_H*/\n"%(g_module_name.upper()))
    fo.write("\n")
    fo.close()
    logger.debug("Successfully generated %s_Cfg.h"%(g_module_name))

def autogen_pbcfg_c():
    # autogen Adc_PBCfg.c
    fo=open(f"{g_gen_file_path}{g_module_name}_PBCfg.c","w",encoding='utf-8')
    fo.write(g_c_file_start_str)
    fo.write("/******************************************************************************/\n")
    fo.write(f'#include "{g_module_name}.h"\n')
    fo.write("/******************************************************************************/\n")
    
    generate_memmap_macros(fo, g_module_name, ["CONFIG_DATA"])
    fo.write("\n")
    
    fo.write(f"static const {g_module_name}_ConfigType {g_module_name}_Config =\n{{\n")
    # 生成post build macro
    postbuild_macro_str = pick_postbuild_outof_ar_cfg_interfaces()
    fo.write(f"{postbuild_macro_str}\n")
    fo.write(f"}};\n\n")
        
    fo.write(g_c_file_end_str)
    fo.close()
    logger.debug("Successfully generated %s_PBCfg.c"%(g_module_name))
    
# 当content_to_write发现具有相同 offset 的字典时，将它们组合起来，并在它们之前和之后插入特定的字典以形成“union”块。
def optimize_top_object_union_content(content_to_write):
    optimized_content = []
    current_group = []
    
    for child_dict in content_to_write:
        if current_group and child_dict['offset'] == current_group[0]['offset']:
            current_group.append(child_dict)
        else:
            if len(current_group) > 1:
                union_name = '_'.join(group_dict['name'] for group_dict in current_group)
                union_block = [{'write_string': '\tunion {\n'}]
                union_block.extend([{'write_string': '\t' + dict_in_group['write_string']} for dict_in_group in current_group])
                union_block.append({'write_string': f'\t}}{union_name};\n'})
                optimized_content.extend(union_block)
            else:
                optimized_content.extend(current_group)
            
            current_group = [child_dict]
    
    # 处理遍历完成后的最后一个组
    if len(current_group) > 1:
        union_name = '_'.join(group_dict['name'] for group_dict in current_group)
        union_block = [{'write_string': '\tunion {\n'}]
        union_block.extend([{'write_string': '\t' + dict_in_group['write_string']} for dict_in_group in current_group])
        union_block.append({'write_string': f'\t}}{union_name};\n'})
        optimized_content.extend(union_block)
    elif current_group:
        optimized_content.extend(current_group)
    
    return optimized_content

def tree_autogen_regdef_h(reserved_tree):
    logger.debug('Enter function %s' % tree_autogen_regdef_h.__name__)
    copyright_str = get_copyright_str(f'{g_product_prefix}{g_module_name}_regdef.h')

    fo=open(f"{g_gen_file_path}{g_product_prefix}{g_module_name}_regdef.h","w",encoding='utf-8')
    fo.write(copyright_str)
    fo.write("#ifndef %s%s_REGDEF_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("#define %s%s_REGDEF_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("\n")
    fo.write("/******************************************************************************/\n")
    fo.write(f"#include \"{g_product_prefix}_TypesReg.h\"\n")
    fo.write("/******************************************************************************/\n")
    fo.write("/******************************************************************************/\n")
    fo.write("/******************************************************************************/\n")
    fo.write("\n")

    #autogen bitfields struct (such as: Lnx_Evadc_CLC_Bits)
    reserved_tree_leaf_list = reserved_tree.get_leaf_nodes(reserved_tree)
    
    autogen_regdef_h_bits_struct(fo,reserved_tree_leaf_list)

    #autogen union (such as: Lnx_Evadc_CLC)
    autogen_regdef_h_union(fo,reserved_tree_leaf_list)
        
    #autogen GLOB/G_Q/G/FC/ROOT object
    non_leaf_nodes = reserved_tree.get_non_leaf_nodes_list()
        
    for current_node in non_leaf_nodes[::-1]:  #反向遍历列表
        #获取从根节点到当前节点的路径, 使用'_'连接所有的节点名称
        path_names_str = current_node.get_path()
        logger.info(path_names_str)              

        fo.write("/** <#>brief %s object */\n"%(current_node.name))
        fo.write("typedef volatile struct\n")
        fo.write("{\n")

        # 创建一个空字典列表，用于保存将要写入文件的内容
        content_to_write = []
        # 遍历当前节点的每一个子节点
        for child in current_node.children:
            # 构建要写入的字符串
            if 'reserved' in child.name:
                write_string = f"\t{g_product_prefix}_UReg_8Bit\t{child.name:>16};\t/** <#>brief {child.offset_start-current_node.offset_start:>4X}. {child.long_name} */\n"
            else:
                namestr = f'{child.name}[{child.number}]' if child.number > 1 else child.name
                write_string = f"\t{path_names_str}_{child.name:<8}\t{namestr+';':<12}/** <#>brief {child.offset_start-current_node.offset_start:>4X}. {child.long_name}({child.short_name}) */\n"
            
            # 创建一个字典来保存当前子节点的信息
            child_info = {
                'write_string': write_string,  # 要写入的字符串
                'offset': child.offset_start - current_node.offset_start,  # 子节点的偏移量
                'name': child.name  # 子节点的名称
            }            
            # 将字典添加到列表中
            content_to_write.append(child_info)

        # 插入“union”块
        optimized_content = optimize_top_object_union_content(content_to_write)  

        # 循环结束后，处理列表中的字典以写入文件
        for child_dict in optimized_content:
            fo.write(child_dict['write_string'])

        fo.write("}%s;\n\n"%(path_names_str))
        fo.write("\n") 


    #结束文件，并关闭文件
    fo.write(f"#endif /*{g_product_prefix.upper()}{g_module_name.upper()}_REGDEF_H */\n")
    fo.close()
    logger.info("Successfully generated %s%s_regdef.h"%(g_product_prefix,g_module_name))

def tree_autogen_bf_h(reg_tree):
    # autogen LnxEvadc_bf.h. Define the bitfield Mask.
    fo=open(f"{g_gen_file_path}{g_product_prefix}{g_module_name}_bf.h","w",encoding='utf-8')
    fo.write("#ifndef %s%s_BF_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("#define %s%s_BF_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("/******************************************************************************/\n")
    fo.write("#include \"%s%s_regdef.h\"\n"%(g_product_prefix,g_module_name))
    fo.write("/******************************************************************************/\n")
    fo.write("\n")

    reg_tree_leaf_list = reg_tree.get_leaf_nodes(reg_tree)

    for current_node in reg_tree_leaf_list:
        if 'reserved' in current_node.name.lower():
            continue
        if current_node.reg_sheet_dictlist == None:
            logger.error('No reg_sheet for %s'%(current_node.name))
        current_node.print_reg_sheet()
        for entry in current_node.reg_sheet_dictlist:
            if 'reserved' in entry['Field'].lower():
                continue
            if '' == entry['Field']:
                continue
            #按照格式#define IFX_EVADC_CLC_DISR_LEN (1u)  #define IFX_EVADC_CLC_DISR_MSK (0x1u) #define IFX_EVADC_CLC_DISR_OFF (0u)生成
            fo.write(f"/** <#>brief Length for {current_node.get_path().upper().replace(f'{g_product_prefix.upper()}',f'{g_product_prefix}')}_Bits.{entry['Field']} */\n")
            fo.write(f"#define {current_node.get_path().upper()}_{entry['Field'].upper()}_LEN ({entry['Width']}u)\n\n")
            fo.write(f"/** <#>brief Mask for {current_node.get_path().upper().replace(f'{g_product_prefix.upper()}',f'{g_product_prefix}')}_Bits.{entry['Field']} */\n")
            fo.write(f"#define {current_node.get_path().upper()}_{entry['Field'].upper()}_MSK ({entry['Mask']}u)\n\n")
            bits_n = parse_bits_mn(entry['Bits'])
            fo.write(f"/** <#>brief Offset for {current_node.get_path().upper().replace(f'{g_product_prefix.upper()}',f'{g_product_prefix}')}_Bits.{entry['Field']} */\n")
            fo.write(f"#define {current_node.get_path().upper()}_{entry['Field'].upper()}_OFF ({bits_n}u)\n\n")                                   

    if g_mask_style == 'nxp5777m':
        fo.write("\n/*************************** 8/16/32bits MASK ************************************/\n")
    for current_node in reg_tree_leaf_list:
        if g_mask_style != 'nxp5777m': continue
        if 'reserved' in current_node.name.lower():
            continue
        if current_node.reg_sheet_dictlist == None:
            logger.error('No reg_sheet for %s'%(current_node.name))
        current_node.print_reg_sheet()
        logger.info(f'{current_node.name},{current_node.size}')
        regsize_byte_1_2_4 = current_node.size
        for entry in current_node.reg_sheet_dictlist:
            if 'reserved' in entry['Field'].lower():
                continue
            if '' == entry['Field']:
                continue
            # NXP5777M style MASK
            bits_n = parse_bits_mn(entry['Bits'])
            nxp5777m_mask_int = int(entry['Mask'], 16) << bits_n
            nxp5777m_mask_hex_str = "0x{:08X}UL".format(nxp5777m_mask_int)  #缺省寄存器size为4bytes
            mask_length = 32
            if regsize_byte_1_2_4 == 1:
                nxp5777m_mask_hex_str = "0x{:02X}U".format(nxp5777m_mask_int)
                mask_length = 8
            if regsize_byte_1_2_4 == 2:
                nxp5777m_mask_hex_str = "0x{:04X}U".format(nxp5777m_mask_int)
                mask_length = 16
            nxp5777m_mask_prefix = f"{g_module_name.upper()}_{current_node.name.upper()}_{entry['Field'].upper()}"
            fo.write(f"/** <#>brief {current_node.name}.{entry['Field']} */\n")
            fo.write(f"#define {nxp5777m_mask_prefix}_MASK  \t({nxp5777m_mask_hex_str})\n")
            fo.write(f"#define {nxp5777m_mask_prefix}_SHIFT \t({bits_n}u)\n")
            fo.write(f"#define {nxp5777m_mask_prefix}_WIDTH \t({entry['Width']}u)\n")
            fo.write(f"#define {nxp5777m_mask_prefix}(x)    \t((({g_product_prefix}_UReg_{mask_length}Bit)((({g_product_prefix}_UReg_{mask_length}Bit)(x)) << {nxp5777m_mask_prefix}_SHIFT)) & {nxp5777m_mask_prefix}_MASK)\n")
            if entry['Width'] <=8:
                getval_length = 8
            elif entry['Width'] <=16:
                getval_length = 16
            elif entry['Width'] <=32:
                getval_length = 32
            else:
                logger.error(f"Width {entry['Width']} is too large for {nxp5777m_mask_prefix}_GET")
                getval_length = 32              
            fo.write(f"#define {nxp5777m_mask_prefix}_GET(y)\t(({g_product_prefix}_UReg_{getval_length}Bit)((({g_product_prefix}_UReg_{mask_length}Bit)(y) & {nxp5777m_mask_prefix}_MASK) >> {nxp5777m_mask_prefix}_SHIFT))\n\n")

    fo.write("\n")
    fo.write(f"#endif /*{g_product_prefix.upper()}{g_module_name.upper()}_BF_H */\n")
    fo.close()
    logger.debug("Successfully generated %s%s_bf.h"%(g_product_prefix,g_module_name))

#把字符串中的每个单词提出来，用下划线连成一个字符串
def connect_sentence_words(sentence):
    words = sentence.split(" ")
    # 去掉word开头的下划线和横杠
    words = [word.strip('_') for word in words]
    words = [word.strip('-') for word in words]
    
    # 使用列表推导式排除空字符和只包含空白字符的单词
    filtered_words = [word for word in words if word and word.strip()]
    return "_".join(filtered_words)

def enumitem_to_macro(typedef_name,typedef_type,typemacroitem):
    logger.info(f"enumitem_to_macro: {typedef_name} {typemacroitem}")
    
    if len(typedef_name) >31: logger.error('Length >31 Error: ',typedef_name)
    #截取type_name的长度小于255. c语言变量名长度的要求.但在这里为后面的enum_name预留64个字符
    if len(typedef_name) > 255-64:
        typedef_name = typedef_name[:254-64]
 
    enumitem_list  = typemacroitem.split('\n')

    enum_def = f"typedef {typedef_type} {typedef_name}_E;\n"
    enum_list_str = ''
    enum_size = 0
    prefix_check = f"#define {typedef_name}_CHECK(ITEMVALUE) ("
    prefix_check_space_str = ' ' * len(prefix_check)
    for line in enumitem_list:
        #按照'::'将line拆分成三部分，分别赋值给number, name,description
        parts = line.split('::')          
        # 确保拆分后有三个部分  
        if len(parts) == 3:  
            number = parts[0]  # 第一部分是number  
            name = parts[1]    # 第二部分是name  
            description = parts[2].lstrip()  # 第三部分是description
        else:  
            continue

        logger.info(f'enumitem_to_macro dispart: {number}-{name}-{description}')
        number = int(number, 16)
        enum_name = connect_sentence_words(name)
        #截取enum_name的长度小于255. c语言变量名长度的要求
        if len(enum_name) > 255:
            enum_name = enum_name[:254]
        #2进制数字转为十进制数字
        enum_def += f"#define {typedef_name}_{enum_name:<20}\t(({typedef_name}_E){number}u)\t/* {description} */\n"
        enum_list_str += f"((ITEMVALUE) == {typedef_name}_{enum_name}) || \\\n{prefix_check_space_str}"
        enum_size += 1

    # 去掉末尾的空格，换行符，或者空格+换行符,或者换行符+空格
    enum_list_str = enum_list_str.rstrip()
    enum_list_str = enum_list_str.rstrip('|| \\')
    enum_def += f"{prefix_check}{enum_list_str})\n"
    return enum_def

def enumitem_to_enum(type_name,enumitem,add_prefix = True,add_invalid_value = True):
    logger.info(f"enumitem_to_enum: {type_name} {enumitem}")
    #截取type_name的长度小于255. c语言变量名长度的要求.但在这里为后面的enum_name预留64个字符
    if len(type_name) > 255-64:
        type_name = type_name[:254-64]
 
    enumitem_list  = enumitem.split('\n')

    enum_def = "typedef enum {\n"
    enum_list_str = ''
    enum_size = 0
    enum_flag = ''
    if add_prefix is True:
        enum_flag = '_E'
    prefix_check = f"#define {type_name}{enum_flag}_CHECK(ENUMVALUE) ("
    prefix_check_space_str = ' ' * len(prefix_check)
    for line in enumitem_list:
        #按照'::'将line拆分成三部分，分别赋值给number, name,description
        parts = line.split('::')          
        # 确保拆分后有三个部分  
        if len(parts) == 3:  
            number = parts[0]  # 第一部分是number  
            name = parts[1]    # 第二部分是name  
            description = parts[2].lstrip()  # 第三部分是description
        else:  
            continue

        logger.info(f'enumitem_to_enum dispart: {number}-{name}-{description}')
        number = int(number, 16)
        if add_prefix is True:
            enum_name = type_name+"_" + connect_sentence_words(name)
        else:
            enum_name = connect_sentence_words(name)
        #截取enum_name的长度小于255. c语言变量名长度的要求
        if len(enum_name) > 255:
            enum_name = enum_name[:254]
        #2进制数字转为十进制数字
        enum_def += f"\t{enum_name:<40}\t={number:>4}u,\t/* {description} */\n"
        enum_list_str += f"((ENUMVALUE) == {enum_name}) || \\\n{prefix_check_space_str}"
        enum_size += 1

    # 定义无效值
    if add_invalid_value is True:
        invalid_value_str = "INVALID_VALUE"
        enum_invalid_value_name = type_name+"_" + invalid_value_str
        enum_def += f"\t{enum_invalid_value_name:<40}\t=\t{g_product_prefix.upper()}_{invalid_value_str}\n"
        
    # 去掉末尾的空格，换行符，或者空格+换行符,或者换行符+空格
    enum_def += "} " + type_name + f"{enum_flag};\n"
    enum_list_str = enum_list_str.rstrip()
    enum_list_str = enum_list_str.rstrip('|| \\')
    enum_def += f"{prefix_check}{enum_list_str})\n"
    return enum_def

# 将大驼峰字符串转为大写字母加下划线的字符串.例如 I2cVersionInfoApi-->I2C_VERSION_INFO_API
def camel_case_to_upper_snake_case(string):
    # 如果字符串已经是全大写且没有驼峰（即没有小写字母），则直接返回
    if string.isupper() and not any(c.islower() for c in string):
        return string
    
    # 定义替换规则
    replacements = {
        'DeInit': 'Deinit',
        'RW': 'Rw',
        'SW': 'Sw',
        'RunTime': 'Runtime',
        'API': 'Api',
        'LPdu': 'Lpdu',
        'TTCAN': 'TtCan'
    }
    # 使用循环应用替换规则
    for old, new in replacements.items(): string = string.replace(old, new)
    
    return_string = ''.join(['_' + i.upper() if i.isupper() else i.upper() for i in string])[1:]
    logger.info(return_string)
    return return_string

def pick_enummacro_outof_ar_cfg_interfaces():
    write_str = ''
    enumeration_list = pick_enumeration_list_outof_ar_type_definitions()
    # 生成 enum Macro
    for dict in g_ar_cfg_interfaces_dictlist:
        if dict['Multiplicity'] != 'EcucEnumerationParamDef': continue  # EcucBooleanParamDef EcucEnumerationParamDef EcucIntegerParamDef EcucReferenceDef EcucStringParamDef
        name_str = dict['Name']
        description_str = dict['Description'].replace('\n','')
        range_str = dict['Range'].strip()
        #write_str += f"\n"
        #write_str += f"/** <#>brief   \t\t{name_str}\n"
        #write_str += f" *  <#>details \t\t{description_str}\n"
        #write_str += f" */\n"
        
        #用换行将range_str分为多个部分。
        # ----注意：cfgif和 typedefinition两个表的Range格式不一样。一个是子表格，一个是格式字符串
        range_str_parts = range_str.split('\n')
        # 如果range的宏在type_definitions中已经定义为枚举，则跳过，不要重复定义
        more_definition = False
        for entry in range_str_parts:
            temp_format_str = entry
            #如果有冒号，去掉冒号后面的内容
            if ':' in temp_format_str:
                temp_format_str = temp_format_str.split(':')[0].strip()            
            if temp_format_str in enumeration_list:
                more_definition = True
            if temp_format_str in write_str:
                more_definition = True
                
        if more_definition is True: continue #跳过，不要重复定义
        notes_str = (f"/* {name_str} - {description_str} */\n")
        enum_str = ''
        for i in range(len(range_str_parts)):
            parts = range_str_parts[i].split(':')
            if len(parts) == 2:
                part_name, part_description = parts
                part_name = part_name.strip()
                part_description = part_description.strip()
                # 每个部分增加序号在结尾
                part_name = f"{part_name:<30} ({i}U)"
                enum_str += f"#define {part_name}\t /* {part_description} */\n"
            else:
                # 每个部分增加序号在结尾
                part_name = f"{parts[0]:<30} ({i}U)"
                enum_str += f"#define {part_name}\n"
        if enum_str not in write_str:
            write_str += notes_str
            write_str += enum_str

        write_str += '\n'
    return write_str

def pick_ecuctype_outof_ar_cfg_interfaces(ecuc_type, cfgclass = 'Pre-Compile'):
    write_str = ''
    # 生成 {ecuc_type} pre-compiler Macro
    for dict in g_ar_cfg_interfaces_dictlist:
        logger.info(dict['Name'])
        if dict['Value configuration class'] != cfgclass: continue
        logger.info(f"pick_ecuctype_outof_ar_cfg_interfaces {cfgclass}, {dict['Value configuration class']}: {dict['Name']}")
        if dict['Multiplicity'] != ecuc_type and ecuc_type != 'ALL': continue  # EcucBooleanParamDef EcucEnumerationParamDef EcucIntegerParamDef EcucReferenceDef EcucStringParamDef EcucSymbolicNameReferenceDef
        name_str = dict['Name']
        description_str = dict['Description'].replace('\n','')
        default_value_str = dict['Default value'].strip()
        if ecuc_type == 'EcucIntegerParamDef' and 'Published-Information' in cfgclass: default_value_str = f'({default_value_str}U)'
        if ecuc_type == 'EcucIntegerParamDef' and 'Pre-Compile' in cfgclass: default_value_str = f'({default_value_str}U)'
        if 'DemEventParameter' in dict['Range']: #and ecuc_type == 'EcucSymbolicNameReferenceDef' and 'Pre-Compile' in cfgclass:
            default_value_str = f'(DemConf_DemEventParameter_{name_str})'
            dem_report_string = dem_error_handling_to_dem_report_string(name_str)            
            if default_value_str == 'NULL':
                default_value_str += f'\n#define {dem_report_string:<28} ({g_module_name.upper()}_DISABLE_DEM_REPORT)'
            else:
                default_value_str += f'\n#define {dem_report_string:<28} ({g_module_name.upper()}_ENABLE_DEM_REPORT)'
        #write_str += f"\n"
        #write_str += f"/** <#>brief   \t\t{name_str}\n"
        #write_str += f" *  <#>details \t\t{description_str}\n"
        #write_str += f" */\n"
        if cfgclass == 'Pre-Compile':
            shaped_macro_name = camel_case_to_upper_snake_case(name_str)
            write_str += (f"/* {name_str}. Post-build variant value:{dict['Post-build variant value']} - {description_str} */\n")
            write_str += f"#define {shaped_macro_name:<28}  {default_value_str}\n\n"
        elif cfgclass == 'Published-Information':
            shaped_macro_name = camel_case_to_upper_snake_case(name_str)
            # 避免重复添加模块名前缀
            if not shaped_macro_name.startswith(f'{g_module_name.upper()}_'):
                shaped_macro_name = f'{g_module_name.upper()}_{shaped_macro_name}'
            shaped_macro_name = shaped_macro_name.replace('_AR_','_AR_RELEASE_') # _AR_RELEASE_特殊，作为一个整体转化
            write_str += (f"/* {name_str}. Post-build variant value:{dict['Post-build variant value']} - {description_str} */\n")
            write_str += f"#define {shaped_macro_name:<28}  {default_value_str}\n\n"
        elif cfgclass == 'Post-Build':
            # name_str 去掉开头的模块名前缀
            shaped_macro_name = name_str.replace(f'{g_module_name}','')
            shaped_default_value_str = default_value_str + ','
            write_str += f"\t.{shaped_macro_name:<28}\t= {shaped_default_value_str:<28}\t/* {name_str}. Post-build variant value:{dict['Post-build variant value']} */\n\n"
            logger.info(f"pick_ecuctype_outof_ar_cfg_interfaces {cfgclass}: {shaped_macro_name} = {default_value_str}")
        else:
            shaped_macro_name = name_str
            write_str += (f"/* {name_str}. Post-build variant value:{dict['Post-build variant value']} - {description_str} */\n")
            write_str += f"#define {shaped_macro_name:<28}  {default_value_str}\n\n"
    return write_str

def pick_published_information_outof_ar_cfg_interfaces():
    write_str = ''
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucBooleanParamDef', 'Published-Information')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucEnumerationParamDef', 'Published-Information')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucIntegerParamDef', 'Published-Information')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucStringParamDef', 'Published-Information')
    return write_str
    
def pick_pre_compile_outof_ar_cfg_interfaces():
    write_str = ''
    # EcucSymbolicNameReferenceDef  ????
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucBooleanParamDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucEnumerationParamDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucIntegerParamDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucFloatParamDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucStringParamDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucReferenceDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucSymbolicNameReferenceDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucParamConfContainerDef', 'Pre-Compile')
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('EcucFunctionNameDef', 'Pre-Compile')    

    return write_str

def pick_postbuild_outof_ar_cfg_interfaces():
    write_str = ''
    write_str += pick_ecuctype_outof_ar_cfg_interfaces('ALL', 'Post-Build')

    return write_str.rstrip()

def pick_typemacro_outof_ar_type_definitions(filter_file):
    write_str = ''
    # 生成TypeMacro
    for dict in g_ar_type_definitions_dictlist:
        if 'TypeMacro' not in dict['Type']: continue
        if dict['File'] != filter_file: continue

        for key,value in dict.items():
            if key != 'Range':continue
            name_str = dict['Syntax']
            type_str = dict['Type'].split('-')[1].strip()
            description_str = dict['Description']
            description_str = description_str.strip().replace('\n',g_notes_format_str)
            write_str += f"\n"
            write_str += f"/** <#>brief   \t\t{name_str}\n"
            write_str += f" *  <#>details \t\t{description_str}\n"
            write_str += f" */\n"
            write_str += f"typedef {type_str} {name_str};\n"
            for entry in value:
                logger.info(entry)
                number,remainder = entry[0].split('-')
                number = number.strip()
                remainder = remainder.strip()
                member_desc_str = entry[1].replace('\n',' ')
                # 如果number是数字,加u; 如果是mask宏，不加u
                if '_MASK' in number:
                    write_str += f"#define {remainder:<20}\t(({name_str}){number})\t\t/* {member_desc_str} */\n"
                else:
                    write_str += f"#define {remainder:<20}\t(({name_str}){number}u)\t\t/* {member_desc_str} */\n"
    return write_str

def pick_eventmacro_outof_ar_type_definitions(filter_file):
    write_str = ''
    # 生成EventMacro
    for dict in g_ar_type_definitions_dictlist:
        if 'EventMacro' not in dict['Type']: continue
        if dict['File'] != filter_file: continue

        for key,value in dict.items():
            if key != 'Range':continue
            write_str += f"\n"
            for entry in value:
                logger.info(entry)
                number,remainder = entry[0].split('-')
                number = number.strip()
                remainder = remainder.strip()
                member_desc_str = entry[1].replace('\n',' ')
                write_str += f"/* {member_desc_str} */\n"
                write_str += f"#define {remainder:<28}  ((uint8){number}U)\n\n"

    return write_str

def pick_errorcodesmacro_outof_ar_type_definitions(filter_file,safety_enable_flag):
    write_str = ''
    # 生成ErrorCodesMacro
    for dict in g_ar_type_definitions_dictlist:
        if 'ErrorCodesMacro' not in dict['Type']: continue
        if dict['File'] != filter_file: continue

        for key,value in dict.items():
            if key != 'Range':continue
            write_str += f"\n"
            for entry in value:
                logger.info(entry)
                number,remainder = entry[0].split('-')
                number = number.strip()
                remainder = remainder.strip()
                member_desc_str = entry[1].replace('\n',' ')
                write_str += f"/* {member_desc_str} */\n"
                write_str += f"#define {remainder:<28}  ((uint8){number}U)\n\n"

    if write_str != '':
        if safety_enable_flag is True:
            ar_precompiler_macros_string = f"#if (({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON) || ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON))\n"
        else:
            ar_precompiler_macros_string = f"#if ({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)\n"
            
        write_str = ar_precompiler_macros_string + write_str
        write_str += f"#endif\n\n"
    return write_str
    
def pick_int_typedef_outof_ar_type_definitions(filter_file):
    write_str = ''
    # 生成uint typedef
    for dict in g_ar_type_definitions_dictlist[::-1]:
        logger.debug(dict)
        if dict['Type'] not in ['uint8','uint16','uint32']: continue
        if dict['File'] != filter_file: continue

        name_str = dict['Syntax'] + ';'
        type_str = dict['Type']
        description_str = dict['Description']
        #write_str += f"\n"
        #write_str += f"/** <#>brief   \t\t{name_str}\n"
        #write_str += f" *  <#>details \t\t{description_str}\n"
        #write_str += f" */\n"
        write_str += f"typedef {type_str:<8} {name_str:<20}\t/* {description_str} */\n\n"
    return write_str

def pick_enumeration_outof_ar_type_definitions(filter_file):
    write_str = ''
    # 生成enum
    for dict in g_ar_type_definitions_dictlist:
        if dict['Type'] != 'Enumeration': continue
        if dict['File'] != filter_file: continue
        logger.info(f"pick_enumeration_outof_ar_type_definitions:{dict['Syntax']}, {filter_file}, {dict['Range']}")
        for key,value in dict.items():
            if key != 'Range':continue
            name_str = dict['Syntax']
            description_str = dict['Description']
            write_str += f"\n"
            write_str += f"/** <#>brief   \t\t{name_str}\n"
            write_str += f" *  <#>details \t\t{description_str}\n"
            write_str += f" */\n"
            write_str += f"typedef enum {{\n"
            for entry in value:
                if '-' in entry[0]:
                    number,remainder = entry[0].split('-')
                    number = number.strip()
                    remainder = remainder.strip()
                else:
                    number = -1
                    remainder = entry[0].strip()

                if number == -1:
                    enum_member = f'{entry[0]},'
                    enum_member = f'{enum_member:<40}'
                else:
                    enum_member = f'{remainder:<32}\t={number:>8}u,'
                member_desc_str = entry[1].replace('\n',' ')
                write_str += f"\t{enum_member:<40}\t/* {member_desc_str} */\n"
            write_str += f"}} {name_str};\n"
    return write_str

def pick_enumeration_list_outof_ar_type_definitions():
    logger.info(f"pick_enumeration_list_outof_ar_type_definitions:")
    enumeration_list = []
    # 生成enum
    for dict in g_ar_type_definitions_dictlist:
        if (dict['Type'] != 'Enumeration' and 'TypeMacro' not in dict['Type']): continue
        logger.info(f"pick_enumeration_list_outof_ar_type_definitions:{dict['Syntax']}, {dict['Range']}")
        for key,value in dict.items():
            if key != 'Range':continue
            for entry in value:
                if '-' in entry[0]:
                    number,remainder = entry[0].split('-')
                    number = number.strip()
                    remainder = remainder.strip()
                else:
                    number = -1
                    remainder = entry[0].strip()

                enumeration_list.append(remainder)
    logger.info(enumeration_list)
    return enumeration_list

def pick_fnptr_outof_ar_type_definitions(filter_file):
    write_str = '\n'
    # 生成函数指针定义
    for dict in g_ar_type_definitions_dictlist:
        if ('FnPtrType' not in dict['Syntax']) and ('Pointer to a function' not in dict['Type']) and ('typedef' not in dict['Type']): continue  # 如何判断是函数指针????
        if dict['File'] != filter_file: continue

        name_str = dict['Syntax']
        type_str = dict['Type']
        description_str = dict['Description']
        #write_str += f"\n"
        #write_str += f"/** <#>brief   \t\t{name_str}\n"
        #write_str += f" *  <#>details \t\t{description_str}\n"
        #write_str += f" */\n"
        prefix = 'Pointer to a function of type '
        if type_str.startswith(prefix):
            #去掉字符串typedef_type前缀
            type_str = type_str[len(prefix):]
            # 字符串typedef_type格式为：type **** Function_name (****)， 用typedef_name的值替换'Function_Name', 'typedef'替换'type'
            type_str = 'typedef '+ re.sub(r'Function_Name', f'(*{name_str})', type_str) + ';'
        write_str += f"{type_str:<80}\t/* {description_str} */\n\n"        
    return write_str

def pick_struct_outof_ar_type_definitions(filter_file):
    write_str = ''
    # 生成struct
    for dict in g_ar_type_definitions_dictlist:
        if dict['Type'] != 'Structure': continue
        logger.info(f"pick_struct_outof_ar_type_definitions:{dict['Syntax']}, {filter_file}, {dict['Range']}")
        if dict['File'] != filter_file: continue

        for key,value in dict.items():
            if key != 'Range':continue
            if value[0][0] == '--': continue
            if value[0][0] == '-': continue
            name_str = dict['Syntax']
            description_str = dict['Description']
            description_str = description_str.strip().replace('\n',g_notes_format_str)
            write_str += f"\n"
            write_str += f"/** <#>brief   \t\t{name_str}\n"
            write_str += f" *  <#>details \t\t{description_str}\n"
            write_str += f" */\n"
            write_str += "typedef struct\n{\n"
            for entry in value:
                struct_member = entry[0]  + ';'   
                member_desc_str = entry[1].replace('\n',' ')
                write_str += f"\t{struct_member:<30}\t/* {member_desc_str} */\n"
            write_str += f"}} {name_str};\n"            
    return write_str

def pick_filelist_outof_ar_type_definitions():
    filelist = []
    # 生成struct
    for dict in g_ar_type_definitions_dictlist:
        if dict['File'] == f'{g_module_name}.h': continue
        if dict['File'] == f'{g_product_prefix}{g_module_name}.h': continue

        for key,value in dict.items():
            filelist.append(dict['File'])
    filelist = list(set(filelist))  # 去重          
    return filelist
        
def build_g_ar_type_definitions_to_h(filename,fo,safety_enable_flag):
    # 生成 ErrorCodesMacro, EventMacro
    write_str = pick_errorcodesmacro_outof_ar_type_definitions(filename,safety_enable_flag)
    write_str += pick_eventmacro_outof_ar_type_definitions(filename)
    
    # 生成uint typedef
    write_str += pick_int_typedef_outof_ar_type_definitions(filename)

    # 生成TypeMacro    
    write_str += pick_typemacro_outof_ar_type_definitions(filename) 
    
    # 生成enum
    write_str += pick_enumeration_outof_ar_type_definitions(filename)

    # 生成函数指针定义
    write_str += pick_fnptr_outof_ar_type_definitions(filename)
    
    # 生成struct
    write_str += pick_struct_outof_ar_type_definitions(filename)
    fo.write(f"{write_str}\n") 

def add_api_dependency_macro_switch(entry):
    # 存在多依赖的情况，比如 CanControllerActivation,CanSetBaudrateApi
    config_dependencies_list_len = 0  # 存在['-']，所以这里不能用len[config_dependencies_parts]
    return_string = ''
    config_dependencies = entry['Configuration dependencies']
    config_dependencies_parts = config_dependencies.split(',')
    config_dependencies_precompile_str = ''
    logger.info(f'{config_dependencies_parts},{config_dependencies_list_len}')
    for config_dependency in config_dependencies_parts:
        config_dependency = config_dependency.strip()
        if not pd.isna(config_dependency) and config_dependency != '-':
            shaped_config_dependency = camel_case_to_upper_snake_case(config_dependency)
            config_dependencies_precompile_str += f"({shaped_config_dependency} == STD_ON) && "
            config_dependencies_list_len +=1
    if config_dependencies_list_len == 1:
        return_string = f'\n#if {config_dependencies_precompile_str[:-4]}'
    if config_dependencies_list_len > 1:
        return_string = f'\n#if ({config_dependencies_precompile_str[:-4]})'
        
    return config_dependencies_list_len, return_string

def build_g_apis_to_h(fo,dictlist):
    logger.info('enter build_g_apis_to_h')
    format_str_1 = g_notes_format_str + '\t'
    for entry in dictlist:
        function_syntax = entry['Syntax'].replace('\n','')
        logger.info(function_syntax)
        if function_syntax.startswith('LOCAL_INLINE'): continue
        # function_syntax的左括号前面如果有空格，则去掉
        # 格式的模式为 void I2c_Init (const I2c_ConfigType* const ConfigPtr)
        if function_syntax.split('(')[0].endswith(' '):
            function_syntax = function_syntax.replace(' (','(')      
        description_str = entry['Description'].strip().replace('\n',g_notes_format_str)
        return_str = entry['Return']
        if '-' == return_str[1].strip():
            return_str1 = ''
        else:
            return_str1 = format_str_1 + return_str[1].strip().replace('\n',format_str_1)
        det_str = entry['Error handling'].strip()
        #if not det_str.startswith('DET:'): continue
        det_str = det_str.replace('\n ','\n')
        det_str = g_notes_format_str + det_str.strip().replace('\n',g_notes_format_str)
        
        config_dependencies_list_len,  return_string = add_api_dependency_macro_switch(entry)
        fo.write(return_string)
        
        fo.write('\n')
        fo.write(f"/** <#>brief   \t\t{function_syntax}\n")
        fo.write(f" *  <#>details \t\t{description_str}\n")
        fo.write(f" *\n")
        if entry['Parameters (in)'][0][0] != '-':
            for pin_entry in entry['Parameters (in)']:
                pin_entry_1 = pin_entry[1].strip()
                if '\n' in pin_entry[1].strip():
                    pin_entry_1 = '\n' + pin_entry[1].strip()
                pin_entry_1 = pin_entry_1.replace('\n',format_str_1)
                fo.write(f" *  <#>param[in]     \t{pin_entry[0]} - {pin_entry_1}\n")
        if entry['Parameters (out)'][0] != '-':
            fo.write(f" *  <#>param[out]    \t{entry['Parameters (out)'][0]} - {entry['Parameters (out)'][1]}\n")
        inout_param = entry['Parameters (in\n- out)']
        logger.info(f'{function_syntax}. inout_param: {inout_param}')
        if inout_param[0] != '-':        
            fo.write(f" *  <#>param[in,out] \t{inout_param[0]} - {inout_param[1]}\n")
        fo.write(f" *  <#>return \t\t{return_str[0]}\t{return_str1}\n")
        fo.write(f" *  <#>note   \t\tService ID:  {entry['Service ID']}\n")
        fo.write(f" *  <#>note   \t\tSync/Async:  {entry['Sync/Async']}\n")
        fo.write(f" *  <#>note   \t\tRe-entrancy: {entry['Re-entrancy']}\n")
        fo.write(f" *  <#>note   \t\tError handling{det_str}\n")

        fo.write(f" */\n")
        fo.write(f"extern {function_syntax};\n")
        if config_dependencies_list_len >0:
            fo.write(f"#endif\n")

def build_g_apis_local_line_to_c(fo,dictlist):
    logger.info('enter build_g_apis_local_line_to_c')
    for entry in dictlist:
        function_syntax = entry['Syntax'].replace('\n','')
        if not function_syntax.startswith('LOCAL_INLINE'): continue
        # function_syntax的左括号前面如果有空格，则去掉
        # 格式的模式为 void I2c_Init (const I2c_ConfigType* const ConfigPtr)
        if function_syntax.split('(')[0].endswith(' '):
            function_syntax = function_syntax.replace(' (','(')      

        fo.write(f"\n{function_syntax};\n")
                
# 在文件开头定义所有可能的section类型
MEMMAP_SECTIONS = {
    "CONFIG_DATA": [
        "CONFIG_DATA_8",
        "CONFIG_DATA_16",
        "CONFIG_DATA_32",
        "CONFIG_DATA_UNSPECIFIED",
        "CONFIG_DATA_8_NO_CACHEABLE",
        "CONFIG_DATA_16_NO_CACHEABLE",
        "CONFIG_DATA_32_NO_CACHEABLE",
        "CONFIG_DATA_UNSPECIFIED_NO_CACHEABLE",
    ],
    "CONST": [
        "CONST_BOOLEAN",
        "CONST_8",
        "CONST_16",
        "CONST_32",
        "CONST_UNSPECIFIED",
    ],
    "CODE": [
        "CODE",
        "RAMCODE",
        "CODE_AC",
    ],
    "VAR_CLEARED": [
        "VAR_CLEARED_BOOLEAN",
        "VAR_CLEARED_8",
        "VAR_CLEARED_16",
        "VAR_CLEARED_32",
        "VAR_CLEARED_UNSPECIFIED",
        "VAR_CLEARED_BOOLEAN_NO_CACHEABLE",
        "VAR_CLEARED_8_NO_CACHEABLE",
        "VAR_CLEARED_16_NO_CACHEABLE",
        "VAR_CLEARED_32_NO_CACHEABLE",
        "VAR_CLEARED_UNSPECIFIED_NO_CACHEABLE",
        "VAR_CLEARED_UNSPECIFIED_AE_NO_CACHEABLE",
    ],
    "VAR_INIT": [
        "VAR_INIT_BOOLEAN",
        "VAR_INIT_8",
        "VAR_INIT_16",
        "VAR_INIT_32",
        "VAR_INIT_UNSPECIFIED",
        "VAR_INIT_BOOLEAN_NO_CACHEABLE",
        "VAR_INIT_8_NO_CACHEABLE",
        "VAR_INIT_16_NO_CACHEABLE",
        "VAR_INIT_32_NO_CACHEABLE",
        "VAR_INIT_UNSPECIFIED_NO_CACHEABLE",
        "VAR_INIT_UNSPECIFIED_AE_NO_CACHEABLE",
    ],
    "VAR_SHARED": [
        "VAR_SHARED_INIT_UNSPECIFIED_NO_CACHEABLE",
        "VAR_SHARED_CLEARED_UNSPECIFIED_NO_CACHEABLE",
    ],
    "QM": [
        "CONST_QM_LOCAL_32",
        "CONST_QM_GLOBAL_32"
    ]
}

# 示例用法：
# 只生成CONFIG_DATA相关的section
#generate_memmap_macros(fo, "Adc", "Lnx", ["CONFIG_DATA"])
# 只生成CODE和CONST相关的section
#generate_memmap_macros(fo, "Can", "Lnx", ["CODE", "CONST"])
# 生成特定的几个section
#generate_memmap_macros(fo, "Uart", "Lnx", ["CODE", "VAR_CLEARED_8_NO_CACHEABLE"])
# 生成所有section
#generate_memmap_macros(fo, "Spi", "Lnx")
def generate_memmap_macros(fo, module_name, section_types=None):
    """
    生成内存映射宏定义，包括START_SEC和STOP_SEC部分
    
    Args:
        fo: 文件对象，用于写入生成的代码
        module_name: 模块名称
        section_types: 需要生成的section类型列表，可以是MEMMAP_SECTIONS中的key，
                      也可以是具体的section名称列表。如果为None，则生成所有类型
    """
    logger.debug('Enter function %s' % generate_memmap_macros.__name__)
    module_upper = module_name.upper()
    
    sections_to_generate = []
    
    # 如果没有指定section_types，使用所有section
    if section_types is None:
        for section_group in MEMMAP_SECTIONS.values():
            sections_to_generate.extend(section_group)
    else:
        # 处理传入的section_types
        for section_type in section_types:
            if section_type in MEMMAP_SECTIONS:
                # 如果是预定义组的key，添加该组所有section
                sections_to_generate.extend(MEMMAP_SECTIONS[section_type])
            else:
                # 否则认为是具体的section名称
                sections_to_generate.append(section_type)
    
    # 生成指定的宏组合
    for section in sections_to_generate:
        # START_SEC宏
        start_macro = f"""
#define {module_upper}_START_SEC_{section}
#include "{module_name}_MemMap.h"
"""
        fo.write(start_macro)
        fo.write(f"// TODO: Implement")
        
        # STOP_SEC宏
        stop_macro = f"""
#define {module_upper}_STOP_SEC_{section}
#include "{module_name}_MemMap.h"
"""
        fo.write(stop_macro)
        
        # 如果是通信模块，添加特殊的缓冲区section
        if (module_name in ["Spi", "I2c", "Uart"] and 
            section == "VAR_CLEARED_8_NO_CACHEABLE"):
            comm_buffer_content = f"static uint8 {module_name}_Buffer[{module_upper}_BUFFER_SIZE];"
            fo.write(f"\n{comm_buffer_content}\n")
    
    logger.debug("Successfully generated MemMap macros for %s" % module_name)

def build_g_ar_error_codes_to_autosar_h(fo):
    # 检查 ArErrorCodes 列表是否不为空
    if g_ar_error_codes_dictlist:
        fo.write(f"#if ({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)\n\n")
        module_sysinfo = g_sysinfo_dict[g_module_name]
        swver = module_sysinfo['AUTOSAR Release']
                
        for entry in g_ar_error_codes_dictlist:
            if swver == '4.4.0':
                macro_name = entry['Error Name: Description'].split(':')[0].replace(' ','').replace('\n','')
                description_str = entry['Error Name: Description'].split(':')[1].strip()
                macro_value = entry['Error ID (AS422)']
            else: #swver == '4.2.2':
                description_str = entry['Description']
                error_codes_and_value = entry['Error code and value'].replace(' ','').replace('\n','')
                macro_name, macro_value = error_codes_and_value.split('=', 1)  # 使用1作为maxsplit参数，确保只分割一次                
            fo.write(f"/* {description_str} */\n")
            fo.write(f"#define {macro_name:<28}  ((uint8){macro_value}U)\n\n")
            
        for entry in g_ar_apis_dictlist:
            service_id = entry['Service ID']
            if service_id == 'nan': continue
            if service_id == 'NA': continue
            if service_id == '-': continue
            function_name =  entry['func_name'].upper()
            if f'{g_module_name.upper()}_' in function_name:
                function_name = function_name.replace(f'{g_module_name.upper()}_','')
                function_name = f'{g_module_name.upper()}_SID_' + function_name.replace('_','')
            description_str = f"API Service ID for {entry['func_name']}"
            fo.write(f"/* {description_str} */\n")
            fo.write(f"#define {function_name:<28}  ((uint8){service_id}U)\n\n")
        
        fo.write(f"#endif /* ({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON) */\n\n")

def build_sid_macro_to_h(fo,safety_enable_flag):
    write_str = ''
    if g_ar_apis_dictlist:
        for entry in g_ar_apis_dictlist:
            service_id = entry['Service ID']
            if service_id == 'nan': continue
            if service_id == 'NA': continue
            if service_id == '-': continue
            function_name =  entry['func_name'].upper()
            if f'{g_module_name.upper()}_' in function_name:
                function_name = function_name.replace(f'{g_module_name.upper()}_','')
                function_name = f'{g_module_name.upper()}_SID_' + function_name.replace('_','')
            description_str = f"API Service ID for {entry['func_name']}"
            write_str += f"/* {description_str} */\n"
            write_str += f"#define {function_name:<28}  ((uint8){service_id}U)\n\n"
            
    if write_str != '':
        if safety_enable_flag is True:
            ar_precompiler_macros_string = f"#if (({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON) || ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON))\n\n"
        else:
            ar_precompiler_macros_string = f"#if ({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)\n\n"
        write_str = ar_precompiler_macros_string + write_str
        write_str += f"#endif\n\n"
        
    fo.write(write_str) 

def build_g_apis_to_c(fo,dictlist,safety_enable_flag):
    logger.info('enter build_g_apis_to_c')
    start_info = """
/*******************************************************************************
**                      Global Function Definitions                           **
*******************************************************************************/
"""
    fo.write(f"{start_info}")
    
    fo.write(f"#define {g_module_name.upper()}_START_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    
    format_str = '\n *\t\t\t\t\t'
    format_str_1 = format_str + '\t'

    for entry in dictlist:
        function_syntax = entry['Syntax'].replace('\n','')
        # function_syntax的左括号前面如果有空格，则去掉
        # 格式的模式为 void I2c_Init (const I2c_ConfigType* const ConfigPtr)
        if function_syntax.split('(')[0].endswith(' '):
            function_syntax = function_syntax.replace(' (','(')      
        description_str = entry['Description'].strip().replace('\n',format_str)
        return_str = entry['Return']
        if '-' == return_str[1].strip():
            return_str1 = ''
        else:
            return_str1 = format_str_1 + return_str[1].strip().replace('\n',format_str_1)
        det_str = entry['Error handling'].strip()
        if not det_str.startswith('DET:'): continue
        det_str = det_str.replace('\n ','\n')
        det_str = format_str + det_str.strip().replace('\n',format_str)
        
        config_dependencies_list_len,  return_string = add_api_dependency_macro_switch(entry)
        fo.write(return_string)

        fo.write(f"\n")
        function_notes = ''
        function_notes += f"/** <#>brief   \t\t{function_syntax}\n"
        function_notes += f" *  <#>details \t\t{description_str}\n"
        function_notes += f" *\n"
        if entry['Parameters (in)'][0][0] != '-':
            for pin_entry in entry['Parameters (in)']:
                pin_entry_1 = pin_entry[1].strip()
                if '\n' in pin_entry[1].strip():
                    pin_entry_1 = '\n' + pin_entry[1].strip()
                pin_entry_1 = pin_entry_1.replace('\n',format_str_1)
                function_notes += f" *  <#>param[in]     \t{pin_entry[0]} - {pin_entry_1}\n"
        if entry['Parameters (out)'][0] != '-':
            function_notes += f" *  <#>param[out]    \t{entry['Parameters (out)'][0]} - {entry['Parameters (out)'][1]}\n"
        inout_param = entry['Parameters (in\n- out)']
        logger.info(f'{function_syntax}. inout_param: {inout_param}')
        if inout_param[0] != '-':        
            function_notes += f" *  <#>param[in,out] \t{inout_param[0]} - {inout_param[1]}\n"
        function_notes += f" *  <#>return \t\t{return_str[0]}\t{return_str1}\n"
        function_notes += f" *  <#>note   \t\tService ID:  {entry['Service ID']}\n"
        function_notes += f" *  <#>note   \t\tSync/Async:  {entry['Sync/Async']}\n"
        function_notes += f" *  <#>note   \t\tRe-entrancy: {entry['Re-entrancy']}\n"
        function_notes += f" *  <#>note   \t\tError handling{det_str}\n"
        function_notes += f" */\n"
        fo.write(function_notes)
        
        function_content = f"{function_syntax}\n"        
        function_content += f'{{\n'  
        # 获取retval,func_name,shaped_parameters
        retval = entry['Return'][0]
        func_name = entry['func_name']
        shaped_parameters = entry['shaped_parameters'].strip()
        #shaped_parameters 去掉结尾的逗号
        if shaped_parameters.endswith(','):
            shaped_parameters = shaped_parameters[:-1]
        ptr_space = ' '
        if retval.endswith('*'):  # 如果retval以*结尾，则ptr_space赋值为' '
            ptr_space = ''
        if retval != 'void':
            function_content += f"\t{retval}{ptr_space}retVal;\n"
        
        det_check_flag = True
        if entry['Service ID'].strip() in ['-','NA','None']: det_check_flag = False
        if entry['Error handling'].strip() in ['-','None']: det_check_flag = False
        logger.info(f"function build_g_apis_to_c entry lines: {entry}")
        error_handling_str = parse_ar_apis_error_handling(entry['Error handling'])
        det_str_items = error_handling_str['DET']
        
        # 如果det内容为None，则不要生成detcheck函数
        if det_str_items == 'None': det_check_flag = False
        
        if det_check_flag:
            module_name_up = g_module_name.upper()
            ar_precompiler_macros_string = f""  
            if safety_enable_flag is True:
                ar_precompiler_macros_string += f"\n\t#if (({module_name_up}_DEV_ERROR_DETECT == STD_ON) || ({module_name_up}_SAFETY_ENABLE == STD_ON))"
            else:
                ar_precompiler_macros_string += f"\n\t#if ({module_name_up}_DEV_ERROR_DETECT == STD_ON)"
            ar_precompiler_macros_string += """
\tuint8 lDetVal;
\tlDetVal = {function_name}DetCheck({parameter_name});
\tif(E_OK == lDetVal)
\t#endif
\t{{
\t\t// TODO: Implement\n\n\n
\t}}
""".format(modulename_up = module_name_up, function_name = func_name, parameter_name = shaped_parameters)
            function_content += ar_precompiler_macros_string
            
        # 解析dem部分多行内容
        error_handling_str_items = error_handling_str['DEM']
        if error_handling_str_items.strip() != 'None' and error_handling_str_items != '':
            error_handling_block_string = build_dem_code_block_to_string(error_handling_str_items)
            function_content += f"{error_handling_block_string}\n"
            
        # 解析Safety Errors部分多行内容
        error_handling_str_items = error_handling_str['Safety Errors']
        if error_handling_str_items.strip() != 'None' and error_handling_str_items != '':
            error_handling_block_string = build_safety_errors_code_block_to_string(error_handling_str_items,entry)
            function_content += f"{error_handling_block_string}\n"

        # 解析Runtime Errors部分多行内容
        error_handling_str_items = error_handling_str['Runtime Errors']
        if error_handling_str_items.strip() != 'None' and error_handling_str_items != '':
            error_handling_block_string = build_runtime_errors_code_block_to_string(error_handling_str_items,entry)
            function_content += f"{error_handling_block_string}\n"

        if retval != 'void':
            function_content += f"\treturn retVal;\n"
        function_content += f'}}'   
        
        fo.write(f"{function_content}\n")      
                  
        if config_dependencies_list_len >0:
            fo.write(f"#endif\n")
        
def tree_autogen_module_h(reg_tree):
    # autogen LnxEvadc.h. 
    file_name = f'{g_product_prefix}{g_module_name}.h'
    fo=open(f"{g_gen_file_path}{file_name}","w",encoding='utf-8')
    fo.write("#ifndef %s%s_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("#define %s%s_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("/******************************************************************************/\n")
    fo.write("#include \"%s%s_bf.h\"\n"%(g_product_prefix,g_module_name))
    fo.write("#include \"%s%s_reg.h\"\n"%(g_product_prefix,g_module_name))
    fo.write('#include "Std_Types.h"\n')
    fo.write("/******************************************************************************/\n")
    fo.write("\n")

    reg_tree_leaf_list = reg_tree.get_leaf_nodes(reg_tree)

    for entry in g_reglist_dictlist:
        if 'reserved' in entry['Field']:
            continue
        #EnumItem
        if entry.get('EnumItem') is None:  # 如果'EnumItem'不存在或其值为None
            continue            
        #如果entry['EnumItem']内容为nan，则continue
        if entry['EnumItem'] == 'nan':
            continue
        if pd.isnull(entry['EnumItem']):
            continue
            
        short_name = entry['Short Name']
        long_name = get_longname_from_global_dict(short_name)

        description_str = entry['Description']
        #截取'Value After Reset:'之前的部分
        if 'Value After Reset:' in description_str:
            description_str = description_str[:description_str.find('Value After Reset:')]
        #如果g_trans_flag为True，并且entry['Description Chinese']存在，则用entry['Description Chinese']替换description_str
        if g_trans_flag is True and 'Description Chinese' in entry:
            description_str = entry['Description Chinese']
            #截取'复位后值:''重置后值:''复位后的值:''重置后的值:'之前的部分
            reset_value_keywords = ['复位后值:', '重置后值:', '复位后的值:', '重置后的值:']
            # 检查描述字符串中是否包含关键词列表中的任意一个关键词
            for keyword in reset_value_keywords:
                if keyword in description_str:
                    # 如果找到关键词，则截取该关键词之前的部分
                    description_str = description_str[:description_str.find(keyword)]
                    break  # 找到后跳出循环，避免处理多个关键词

        logger.info(f"{entry['Short Name']}{entry['Field']}, {entry['EnumItem']}")
        fo.write(f"/** <#>brief {long_name}({short_name}): {entry['Field']}\n")
        fo.write(f" * Notes: \n")
        fo.write(f" *        {description_str}\n".replace('\n','\n *        '))
        short_name_size,short_name_path = get_values_from_tree_leaf_list(short_name,reg_tree_leaf_list)
        path_str = short_name_path.replace(f'{g_product_prefix}_{g_module_name.upper()}_','')
        path_str = f"{g_product_prefix}{g_module_name}_{path_str}"
        type_name = f"{path_str}_{entry['Field']}"
        #fo.write(f"Definition in {type_name}\n")
        fo.write(f" */\n")
        if g_trans_enum_item_description is True and 'EnumItem Chinese' in entry:
            enumitem_chose = entry['EnumItem Chinese']
        else:
            enumitem_chose = entry['EnumItem']
        resultstr = enumitem_to_macro(f'{type_name}',f'{g_product_prefix}_UReg_32Bit',enumitem_chose)
        fo.write(f"{resultstr}\n")

    build_g_ar_type_definitions_to_h(file_name,fo,False)
    
    fo.write(f"\n")

    # 生成moudule.c函数的声明
    fo.write(f"#define {g_module_name.upper()}_START_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    build_g_apis_to_h(fo,g_reg_apis_dictlist)
    fo.write(f"\n#define {g_module_name.upper()}_STOP_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")

    fo.write(f"\n#endif /*{g_product_prefix.upper()}{g_module_name.upper()}_H */\n")
    fo.close()
    logger.debug("Successfully generated %s%s.h"%(g_product_prefix,g_module_name))

def tree_autogen_module_c():
    # autogen LnxEvadc.h. Define the bitfield Value Enum.
    fo=open(f"{g_gen_file_path}{g_product_prefix}{g_module_name}.c","w",encoding='utf-8')        
    file_header_str = f"/******************************************************************************/\n"
    file_header_str += f"#include \"%s%s.h\"\n"%(g_product_prefix,g_module_name)
    file_header_str += f"/******************************************************************************/\n\n"
    fo.write(file_header_str)

# 增加header注释
    header_notes_str = """
/*******************************************************************************
**                        GLOBAL CONSTANTS/VARIABLES                          **
*******************************************************************************/"""
    fo.write(header_notes_str)
    generate_memmap_macros(fo, g_module_name, ["CONST", "VAR_CLEARED", "VAR_INIT", "VAR_SHARED"])
    
    build_g_apis_to_c(fo,g_reg_apis_dictlist,False)
    fo.write(f"\n#define {g_module_name.upper()}_STOP_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    fo.close()

    logger.debug("Successfully generated %s%s.c"%(g_product_prefix,g_module_name))

def build_switch_string(entry,retval):
    if 'Member List' not in entry:return None
    member_list_entry = entry['Member List'][0]
    if pd.isna(member_list_entry['Union']):
        reg_str = member_list_entry['Register']
    else:
        reg_str = f"{member_list_entry['Union']}.{member_list_entry['Register']}"
    field = member_list_entry['Field']
    switch_str = f"\tswitch ({g_module_name.lower()}->{reg_str}.B.{field}) /* {member_list_entry['Notes']}*/\n\t{{\n"
    if retval.endswith('_E'):
        type_name = retval[:-2]  # 移除末尾的 '_E'
    else:
        type_name = retval
    for entry in g_reglist_dictlist:
        if entry['Short Name'] not in type_name:continue
        if entry['Field'] not in type_name:continue
        break
    if g_trans_flag is True:
        enumitem_str = entry['EnumItem Chinese']
    else:
        enumitem_str = entry['EnumItem']
    enumitem_list = enumitem_str.split('\n')

    for line in enumitem_list:
        #按照'::'将line拆分成三部分，分别赋值给number, name,description
        parts = line.split('::')          
        # 确保拆分后有三个部分  
        if len(parts) == 3:  
            number = parts[0]  # 第一部分是number  
            name = parts[1]    # 第二部分是name  
            description = parts[2].lstrip()  # 第三部分是description
        else:  
            continue

        logger.info(f'build_switch_string dispart: {number}-{name}-{description}')
        enum_name = type_name+"_" + connect_sentence_words(name)
        #截取enum_name的长度小于255. c语言变量名长度的要求
        if len(enum_name) > 255:
            enum_name = enum_name[:254]
        #2进制数字转为十进制数字
        switch_str += f"\tcase {number}u:\n\t\tretVal = {enum_name}; /* {description}*/\n\t\tbreak;\n\n"

    switch_str += f"\tdefault:\n\t\tretVal = {type_name}_INVALID_VALUE;\n\t\tbreak;\n"
    switch_str += f'\t}}\n'
    return switch_str

sample_parse_string = """ 
DET:
FEE_E_UNINIT : API service is called when the module is not initialized
FEE_E_INVALID_CANCEL: API service is called while no job is pending
Runtime Errors: None
DEM: None
Safety Errors:
CRC_E_PARAM_LENGTH: Error ID for zero length check. This is implemented as MCAL safety error.
CRC_E_PARAM_POINTER: Error ID for NULLPTR check. This is implemented as MCAL safety error.
CRC_E_CHANNEL_TIMEOUT: Error ID for timeout check in hardware mode when waiting for free channels. This is implemented as MCAL safety error.
Note: All DET IDs are also reported as safety errors.
"""

result_str = '''{
    'DET': 'FEE_E_UNINIT : API service is called when the module is not initialized FEE_E_INVALID_CANCEL: API service is called while no job is pending',
    'Runtime Errors': 'None',
    'DEM': 'None',
    'Safety Errors': 'CRC_E_PARAM_LENGTH: Error ID for zero length check. This is implemented as MCAL safety error. CRC_E_PARAM_POINTER: Error ID for NULLPTR check. This is implemented as MCAL safety error. CRC_E_CHANNEL_TIMEOUT: Error ID for timeout check in hardware mode when waiting for free channels. This is implemented as MCAL safety error.',
    'Note': 'All DET IDs are also reported as safety errors.'
}'''
def parse_ar_apis_error_handling(input_string):
    logger.info(f"Enter function {parse_ar_apis_error_handling.__name__}")

    # 匹配以关键字（可能带有前导空格）后跟冒号开始的内容，直到下一个关键字或字符串结束
    pattern = r'(?P<section>(?:(?:DET|Runtime Errors|DEM|Safety Errors|Note)\s*:.*?)(?:(?=\b(?:DET|Runtime Errors|DEM|Safety Errors|Note)\s*:)|$))'
    # 使用正则表达式查找所有匹配项
    matches = re.finditer(pattern, input_string, re.DOTALL)
    result = {
        'DET': '',
        'Runtime Errors': '',
        'DEM': '',
        'Safety Errors': '',
        'Note': ''
    }
    # 遍历匹配项，并更新结果字典
    for match in matches:
        # 提取整个匹配的内容（包括关键字和冒号）
        section_content = match.group().strip()

        # 提取关键字（去除前导空格和冒号）
        section_name = re.match(r'\s*(?P<name>DET|Runtime Errors|DEM|Safety Errors|Note)\s*:', section_content).group('name')

        # 去除内容中的关键字和冒号（如果需要）
        # 这里简单处理，仅去除开头的关键字和冒号
        section_content = section_content[len(section_name) + 2:].strip()  # +2 是为了去除冒号和可能存在的空格

        # 更新结果字典
        result[section_name] = section_content

    return result

def dem_error_handling_to_dem_report_string(dem_error_handling_str_line):
    dem_report_str = dem_error_handling_str_line
    if '_E_' in dem_error_handling_str_line:
            dem_report_str = dem_error_handling_str_line.replace('_E_', '_')
    dem_report_str = f'{dem_report_str}_DEM_REPORT'
    return dem_report_str

def build_dem_code_block_to_string(str_items):
    logger.info(f"Enter function {build_dem_code_block_to_string.__name__}.parameter: {str_items}")
    block_dem_passed = ''
    block_dem_failed = ''
    block_content = ''
    
    # 如果dem内容为None，则返回''
    if str_items == 'None': return ''
            
    # 解析dem部分多行内容
    str_items_list  = str_items.split('\n')
    for str_line in str_items_list:
        # str_line如果只是空格换行，则跳过
        str_line = str_line.strip()
        if str_line == '': continue
        if not ':' in str_line: continue
        if str_line.strip().lower() == 'none': continue
        # str_line如果有冒号，取冒号前面的部分作为参数
        if ':' in str_line:
            str_line = str_line.split(':')[0].strip()
            
        dem_report_str_line = dem_error_handling_to_dem_report_string(str_line)
        block_dem_passed = f"""
    #if ({dem_report_str_line} == {g_module_name.upper()}_ENABLE_DEM_REPORT)
    /* Raise a DEM {str_line} PASSED */
    Dem_ReportErrorStatus({str_line}, DEM_EVENT_STATUS_PASSED);
    #endif
"""
        block_content += block_dem_passed

        block_dem_failed = f"""
    #if ({dem_report_str_line} == {g_module_name.upper()}_ENABLE_DEM_REPORT)
    /* Raise a DEM {str_line} FAILED */
    Dem_ReportErrorStatus({str_line}, DEM_EVENT_STATUS_FAILED);
    #endif
"""
        block_content += block_dem_failed

    return block_content

def build_safety_errors_code_block_to_string(str_items,entry):
    logger.info(f"Enter function {build_safety_errors_code_block_to_string.__name__}.parameter: {str_items}")
    block_string = ''
    block_content = ''
    
    # 如果safety_errors内容为None，则返回''
    if str_items == 'None': return ''
    
    function_name =  entry['func_name'].upper()
    if f'{g_module_name.upper()}_' in function_name:
        function_name = function_name.replace(f'{g_module_name.upper()}_','')
        serviceid_macro_name = f'{g_module_name.upper()}_SID_' + function_name.replace('_','')            
            
    # 解析safety_errors部分多行内容
    str_items_list  = str_items.split('\n')
    for str_line in str_items_list:
        # str_line如果只是空格换行，则跳过
        str_line = str_line.strip()
        if str_line == '': continue
        if not ':' in str_line: continue
        if str_line.strip().lower() == 'none': continue
        # str_line如果有冒号，取冒号前面的部分作为参数
        if ':' in str_line:
            str_line = str_line.split(':')[0].strip()
        block_string = f"""
\t#if ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON)
\tMcal_ReportSafetyError({g_module_name.upper()}_MODULE_ID, {g_module_name.upper()}_INSTANCE_ID, {serviceid_macro_name}, {str_line});
\t#endif
"""
        block_content += block_string

    return block_content
        
def build_runtime_errors_code_block_to_string(str_items,entry):
    logger.info(f"Enter function {build_runtime_errors_code_block_to_string.__name__}.parameter: {str_items}")
    block_string = ''
    block_content = ''
    
    # 如果内容为None，则返回''
    if str_items == 'None': return ''
            
    function_name =  entry['func_name'].upper()
    if f'{g_module_name.upper()}_' in function_name:
        function_name = function_name.replace(f'{g_module_name.upper()}_','')
        serviceid_macro_name = f'{g_module_name.upper()}_SID_' + function_name.replace('_','')            

    # 解析多行内容
    str_items_list  = str_items.split('\n')
    for str_line in str_items_list:
        # str_line如果只是空格换行，则跳过
        str_line = str_line.strip()
        if str_line == '': continue
        if not ':' in str_line: continue
        if str_line.strip().lower() == 'none': continue
        # str_line如果有冒号，取冒号前面的部分作为参数
        if ':' in str_line:
            str_line = str_line.split(':')[0].strip()
        block_string = f"""
\t#if ({g_module_name.upper()}_RUNTIME_ERROR_DETECT == STD_ON)
\tDet_ReportRuntimeError({g_module_name.upper()}_MODULE_ID, {g_module_name.upper()}_INSTANCE_ID, {serviceid_macro_name}, {str_line});
\t#endif
"""
        block_content += block_string

    return block_content

def build_detcheck_functions_to_autosar_c(filehandler,safety_enable_flag):
    logger.info(f"Enter function {build_detcheck_functions_to_autosar_c.__name__}")
    file_content = ''
    format_str = '\n *\t\t\t\t\t'
    format_str_1 = format_str + '\t'

    for entry in g_ar_apis_dictlist:
        if entry['Service ID'].strip() in ['-','NA','None']: continue
        if entry['Error handling'].strip() in ['-','None']: continue
        if not entry['Error handling'].strip().startswith('DET:'): continue
        logger.info(f"function build_detcheck_functions_to_autosar_c entry lines: {entry}")
        error_handling_str = parse_ar_apis_error_handling(entry['Error handling'])
        det_str_items = error_handling_str['DET']
        
        # 如果det内容为None，则不要生成detcheck函数
        if det_str_items == 'None': continue
        
        det_str = det_str_items.replace('\n ','\n')
        det_str = format_str_1 + det_str.strip().replace('\n',format_str_1)
        
        return_type = entry['Return'][0]
        # 去掉entry['Syntax']字符串 开头的return_type字符串
        pattern = r'^\s*' + re.escape(return_type)        
        func_name_temp = re.sub(pattern, '', entry['Syntax']).lstrip()
        current_function = func_name_temp.replace('\n','').strip()
        parameters = current_function.replace(entry['func_name'],'').strip().lstrip('(').rstrip(')')
        logger.info(f"detcheck fucntion parameters: {parameters}")
        
        description_str = entry['Description']     
        retval =  'LOCAL_INLINE Std_ReturnType'

        config_dependencies_list_len,  return_string = add_api_dependency_macro_switch(entry)
        if config_dependencies_list_len == 0: 
            return_string += f"\n#if "
            end_string = '\n'
        else:
            return_string += ' && \\\n\t'
            return_string = return_string.replace('\n#if ','\n#if (')
            end_string = ')\n'

        if safety_enable_flag is True:
            function_precompile_header = return_string + f"(({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON) || ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON))" + end_string
        else:
            function_precompile_header = return_string + f"({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)" + end_string

        # 拼接字符串
        function_header = f"/**\n"
        if g_trans_flag is True and 'Function Description Chinese' in entry:
            function_header += f" *  <#>brief  \t\t函数{current_function} 对应的DET检测函数\n"
        else:
            function_header += f" *  <#>brief  \t\tThis function is to check all the DETs for {entry['func_name']} API\n"

        function_header += f" *  <#>note   \t\tDET:         {det_str}\n"
        
        function_header += f" */\n"
        ptr_space = ' '
        if retval.endswith('*') and retval.endswith(' *'):
            ptr_space = ''
        function_content = f"{retval}{ptr_space}{entry['func_name']}DetCheck({parameters})\n"

        function_content += "{\n"
        function_content += f"\tStd_ReturnType ErrorStatus = E_OK;\n"
        
        # 解析det部分多行内容
        det_str_items_list  = det_str_items.split('\n')
        for det_str_line in det_str_items_list:
            # det_str_line如果只是空格换行，则跳过
            det_str_line = det_str_line.strip()
            if det_str_line == '': continue
            if det_str_line.strip().lower() == 'none': continue
            # det_str_line如果有冒号，取冒号前面的部分作为参数
            if ':' in det_str_line:
                det_str_line = det_str_line.split(':')[0].strip()
            function_name =  entry['func_name'].upper()
            if f'{g_module_name.upper()}_' in function_name:
                function_name = function_name.replace(f'{g_module_name.upper()}_','')
                serviceid_macro_name = f'{g_module_name.upper()}_SID_' + function_name.replace('_','')  
            function_content += f"\n"
            if safety_enable_flag is True: function_content += f"\t#if ({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)\n"          
            function_content += f"\tDet_ReportError({g_module_name.upper()}_MODULE_ID, {g_module_name.upper()}_INSTANCE_ID, {serviceid_macro_name}, {det_str_line});\n"
            if safety_enable_flag is True: function_content += f"\t#endif\n"
            
            if safety_enable_flag is True:
                function_content += f"\t#if ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON)\n"          
                function_content += f"\tMcal_ReportSafetyError({g_module_name.upper()}_MODULE_ID, {g_module_name.upper()}_INSTANCE_ID, {serviceid_macro_name}, {det_str_line});\n"
                function_content += f"\t#endif\n"
                
            function_content += f"\tErrorStatus = E_NOT_OK;\n"
        function_content += f"\n\t// TODO: Implement\n"
        function_content += f"\treturn ErrorStatus;\n"
        # 添加函数尾
        function_content += "}\n"
        function_precompile_tailer = f"#endif\n"
        
        file_content += function_precompile_header + function_header + function_content + function_precompile_tailer

    # 一次性写入整个函数内容到文件
    filehandler.write(file_content)
    
def build_detcheck_extern_to_autosar_c(filehandler,safety_enable_flag):
    logger.info(f"Enter function {build_detcheck_extern_to_autosar_c.__name__}")
    file_content = '/************** Local APIs for DET checks - Start**************/'

    for entry in g_ar_apis_dictlist:
        if entry['Service ID'].strip() in ['-','NA','None']: continue
        if entry['Error handling'].strip() in ['-','None']: continue
        if not entry['Error handling'].strip().startswith('DET:'): continue
        logger.info(f"function build_detcheck_extern_to_autosar_c entry lines: {entry}")
        error_handling_str = parse_ar_apis_error_handling(entry['Error handling'])
        det_str_items = error_handling_str['DET']
        
        # 如果det内容为None，则不要生成detcheck函数
        if det_str_items == 'None': continue        
        
        return_type = entry['Return'][0]
        # 去掉entry['Syntax']字符串 开头的return_type字符串
        pattern = r'^\s*' + re.escape(return_type)        
        func_name_temp = re.sub(pattern, '', entry['Syntax']).lstrip()
        
        current_function = func_name_temp.replace('\n','').strip()
        
        parameters = current_function.replace(entry['func_name'],'').strip().lstrip('(').rstrip(')')
        retval =  'LOCAL_INLINE Std_ReturnType'
        
        config_dependencies_list_len,  return_string = add_api_dependency_macro_switch(entry)
        if config_dependencies_list_len == 0: 
            return_string += f"\n#if "
            end_string = '\n'
        else:
            return_string += ' && \\\n\t'
            return_string = return_string.replace('\n#if ','\n#if (')
            end_string = ')\n'

        if safety_enable_flag is True:
            function_precompile_header = return_string + f"(({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON) || ({g_module_name.upper()}_SAFETY_ENABLE == STD_ON))" + end_string
        else:
            function_precompile_header = return_string + f"({g_module_name.upper()}_DEV_ERROR_DETECT == STD_ON)" + end_string

        # 拼接字符串
        function_header = f""
        ptr_space = ' '
        if retval.endswith('*') and retval.endswith(' *'):
            ptr_space = ''
        function_content = f"{retval}{ptr_space}{entry['func_name']}DetCheck({parameters});\n"
        function_precompile_tailer = f"#endif\n"
        
        file_content += function_precompile_header + function_header + function_content + function_precompile_tailer

    # 一次性写入整个函数内容到文件
    filehandler.write(file_content)
    
def get_retval(row):
    retval_str = 'void'
    if '<#>retval' in row['Function Description']:
        retval_str = row['Function Description'].split('<#>retval ')[1].split(':')[0].strip()
    return retval_str

def tree_autogen_GeneralTypes_h():
    # autogen Can_GeneralTypes.h
    # 创建common/Platform目录
    platform_dir = os.path.join(g_gen_file_path, 'common', 'Platform')
    os.makedirs(platform_dir, exist_ok=True)
    
    file_name = f'{g_module_name}_GeneralTypes.h'
    file_path = os.path.join(platform_dir, file_name)
    fo=open(file_path,"w",encoding='utf-8')
    fo.write("#ifndef %s_GENERALTYPES_H\n"%(g_module_name.upper()))
    fo.write("#define %s_GENERALTYPES_H\n"%(g_module_name.upper()))
    fo.write("""\n/*******************************************************************************
**                      Global Type Definitions                               **
*******************************************************************************/\n""")
    build_g_ar_type_definitions_to_h(file_name,fo,False)
       
    fo.write(f"#endif /*{g_module_name.upper()}_GENERALTYPES_H */\n")
    fo.close()
    logger.debug("Successfully generated %s_GeneralTypes.h"%(g_module_name))
    
def tree_autogen_autosar_h():
    # autogen LnxEvadc.h. Define the bitfield Value Enum.
    file_name = f'{g_module_name}.h'
    fo=open(f"{g_gen_file_path}{file_name}","w",encoding='utf-8')
    fo.write("#ifndef %s_H\n"%(g_module_name.upper()))
    fo.write("#define %s_H\n"%(g_module_name.upper()))
    fo.write("/******************************************************************************/\n")
    fo.write('#include "Std_Types.h"\n')
    fo.write("#include \"%s%s_bf.h\"\n"%(g_product_prefix,g_module_name))
    fo.write("#include \"%s%s.h\"\n"%(g_product_prefix,g_module_name))

    filelist = pick_filelist_outof_ar_type_definitions()
    if filelist != '':
        for file in filelist:
            fo.write(f'#include "{file}"\n')
    fo.write(f'#include "{g_module_name}_Cfg.h"\n')
    fo.write("/******************************************************************************/\n")
    fo.write("\n")

    pick_precompile_str = pick_ecuctype_outof_ar_cfg_interfaces('EcucBooleanParamDef', 'Pre-Compile')
    logger.info(pick_precompile_str)
    safety_enable_flag = False
    # {g_module_name}SafetyEnable 配置项是否存在于cfgInterfaces表中
    if f'{g_module_name}SafetyEnable' in pick_precompile_str: safety_enable_flag = True
        
    pre_compiler_string = pick_pre_compile_outof_ar_cfg_interfaces()
    if 'INSTANCEID' not in pre_compiler_string.replace(' ','').upper():
        fo.write(f"#define {g_module_name.upper()}_INSTANCE_ID              ((uint8)0U)\n\n")
    
    enum_macro_str = pick_enummacro_outof_ar_cfg_interfaces()
    fo.write(enum_macro_str)
    #build_g_ar_error_codes_to_autosar_h(fo)  
    build_sid_macro_to_h(fo,safety_enable_flag)    
    build_g_ar_type_definitions_to_h(file_name,fo,safety_enable_flag)
    
    fo.write(f"#define {g_module_name.upper()}_START_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    build_g_apis_to_h(fo,g_ar_apis_dictlist)
    fo.write(f"\n#define {g_module_name.upper()}_STOP_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
       
    fo.write(f"\n#endif /*{g_module_name.upper()}_H */\n")
    fo.close()
    logger.debug("Successfully generated %s%s.h"%(g_product_prefix,g_module_name))

def tree_autogen_autosar_c():
    # autogen LnxEvadc.h. Define the bitfield Value Enum.
    fo=open(f"{g_gen_file_path}{g_module_name}.c","w",encoding='utf-8')

    pick_precompile_str = pick_ecuctype_outof_ar_cfg_interfaces('EcucBooleanParamDef', 'Pre-Compile')
    logger.info(pick_precompile_str)
    safety_enable_flag = False
    # {g_module_name}SafetyEnable 配置项是否存在于cfgInterfaces表中
    if f'{g_module_name}SafetyEnable' in pick_precompile_str: safety_enable_flag = True
    
    module_name_up = g_module_name.upper()
    including_string = """/*******************************************************************************
**                      Includes                                              **
*******************************************************************************/  
#include "{modulename}.h"

#if({modulename_up}_INIT_DEINIT_API_MODE != {modulename_up}_MCAL_SUPERVISOR)
#include "McalLib_OsStub.h"
#endif

#if ({modulename_up}_DEV_ERROR_DETECT == STD_ON)
/* Include default error detection header file if DET is Enabled */
#include "Det.h"
#endif
""".format(modulename_up = module_name_up, modulename = g_module_name)
    fo.write(f"{including_string}")

    including_safety_string = """
#if ({modulename_up}_SAFETY_ENABLE == STD_ON)
#include "Mcal_SafetyError.h"
#endif
""".format(modulename_up = module_name_up, modulename = g_module_name)
    if safety_enable_flag is True: fo.write(f"{including_safety_string}")

    dem_report_include_string = ''
    for dict in g_ar_cfg_interfaces_dictlist:
        if 'DemEventParameter' not in dict['Range']: continue
        name_str = dict['Name']
        dem_report_string = dem_error_handling_to_dem_report_string(name_str)
        dem_report_include_string += f'({dem_report_string}  == {g_module_name.upper()}_ENABLE_DEM_REPORT)||'

    if dem_report_include_string != '':
        dem_report_include_string = dem_report_include_string[:-2]
        
    dem_h_include = """#if ({dem_report_string})
#include "Dem.h"
#endif

""".format(dem_report_string = dem_report_include_string)
    
    if dem_report_include_string != '':
        fo.write(f"\n{dem_h_include}")

    factory_version_info_string = get_factory_version_info()
    fo.write(factory_version_info_string)

    version_check = """#ifndef {modulename_up}_AR_RELEASE_MAJOR_VERSION
  #error "{modulename_up}_AR_RELEASE_MAJOR_VERSION is not defined."
#endif

#if ( {modulename_up}_AR_RELEASE_MAJOR_VERSION != 4U)
  #error "{modulename_up}_AR_RELEASE_MAJOR_VERSION does not match."
#endif

#ifndef {modulename_up}_SW_MAJOR_VERSION
  #error "{modulename_up}_SW_MAJOR_VERSION is not defined."
#endif

#ifndef {modulename_up}_SW_MINOR_VERSION
  #error "{modulename_up}_SW_MINOR_VERSION is not defined."
#endif

#ifndef {modulename_up}_SW_PATCH_VERSION
  #error "{modulename_up}_SW_PATCH_VERSION is not defined."
#endif

#if ( {modulename_up}_SW_MAJOR_VERSION != {modulename_up}_C_SW_MAJOR_VERSION )
  #error "{modulename_up}_SW_MAJOR_VERSION does not match."
#endif

#if ( {modulename_up}_SW_MINOR_VERSION != {modulename_up}_C_SW_MINOR_VERSION )
  #error "{modulename_up}_SW_MINOR_VERSION does not match."
#endif

#if ( {modulename_up}_SW_PATCH_VERSION != {modulename_up}_C_SW_PATCH_VERSION )
  #error "{modulename_up}_SW_PATCH_VERSION does not match."
#endif
 
#if ({modulename_up}_DEV_ERROR_DETECT == STD_ON)

  #ifndef DET_AR_RELEASE_MAJOR_VERSION
    #error "DET_AR_RELEASE_MAJOR_VERSION is not defined."
  #endif

  #if (DET_AR_RELEASE_MAJOR_VERSION != 4U)
    #error "DET_AR_RELEASE_MAJOR_VERSION does not match."
  #endif

#endif /* End for {modulename_up}_DEV_ERROR_DETECT */
""".format(modulename_up = module_name_up, modulename = g_module_name)
    fo.write(f"{version_check}")
    
    dem_report_cfg_string = ''
    for dict in g_ar_cfg_interfaces_dictlist:
        if 'DemEventParameter' not in dict['Range']: continue
        name_str = dict['Name']
        dem_report_string = dem_error_handling_to_dem_report_string(name_str)
        dem_report_cfg_string += f'({dem_report_string}  == {g_module_name.upper()}_ENABLE_DEM_REPORT)||'

    if dem_report_cfg_string != '':
        dem_report_cfg_string = dem_report_cfg_string[:-2]
        
    dem_version_check = """
#if ({dem_report_string})

  #ifndef DEM_AR_RELEASE_MAJOR_VERSION
    #error "DEM_AR_RELEASE_MAJOR_VERSION is not defined."
  #endif

  #if ( DEM_AR_RELEASE_MAJOR_VERSION != 4U)
    #error "DEM_AR_RELEASE_MAJOR_VERSION does not match."
  #endif

#endif /*End for DEM Checks*/
""".format(dem_report_string = dem_report_cfg_string)
    
    if dem_report_cfg_string != '':
        fo.write(f"{dem_version_check}")

    user_mode_macros = """
/*******************************************************************************
**                         User Mode Macros                                   **
*******************************************************************************/
#if({modulename_up}_INIT_DEINIT_API_MODE == {modulename_up}_MCAL_SUPERVISOR)
  #define {modulename_up}_INIT_DEINIT_WRITE_PERIP_ENDINIT_PROTREG(RegAdd,Data)   \
    Mcal_WritePeripEndInitProtReg(RegAdd,Data)
  #define {modulename_up}_INIT_DEINIT_WRITE_SAFETY_ENDINIT_PROTREGMASK(RegAdd,Data,Mask)   \
    Mcal_WriteSafetyEndInitProtRegMask(RegAdd,Data,Mask)
#else
  #define {modulename_up}_INIT_DEINIT_WRITE_PERIP_ENDINIT_PROTREG(RegAdd,Data)   \
    MCAL_LIB_WRITEPERIPENDINITPROTREG(RegAdd,Data)
  #define {modulename_up}_INIT_DEINIT_WRITE_SAFETY_ENDINIT_PROTREGMASK(RegAdd,Data,Mask)   \
    MCAL_LIB_WRITESAFETYENDINITPROTREGMASK(RegAdd,Data,Mask)
#endif
""".format(modulename_up = module_name_up, modulename = g_module_name)
    fo.write(f"{user_mode_macros}\n")                    

    user_mode_macros_runtime = """#if ({modulename_up}_RUNTIME_API_MODE == {modulename_up}_MCAL_SUPERVISOR)
  #define {modulename_up}_RUNTIME_WRITE_SAFETY_ENDINIT_PROTREGMASK(RegAdd,Data,Mask)    \
  Mcal_WriteSafetyEndInitProtRegMask(RegAdd,Data,Mask)
  #define {modulename_up}_RUNTIME_OS_MODIFY32(ModuleId, RegAdd, ClearMask, SetMask)        \
                         ((RegAdd->U) = ((RegAdd->U)|(SetMask)))
#else
  #define {modulename_up}_RUNTIME_WRITE_SAFETY_ENDINIT_PROTREGMASK(RegAdd,Data,Mask)    \
  MCAL_LIB_WRITESAFETYENDINITPROTREGMASK(RegAdd,Data,Mask)
  #define {modulename_up}_RUNTIME_OS_MODIFY32(ModuleId,RegAdd,ClearMask,SetMask)  \
  MCAL_SFR_OS_MODIFY32(ModuleId, RegAdd, ClearMask, SetMask)
#endif
""".format(modulename_up = module_name_up, modulename = g_module_name)
    pre_compile_str = pick_ecuctype_outof_ar_cfg_interfaces('EcucEnumerationParamDef', 'Pre-Compile')
    if '_RUNTIME_API_MODE' in pre_compile_str:
        fo.write(f"{user_mode_macros_runtime}\n")                    
    
    # 增加private function header注释
    private_function_header_str = """
/*******************************************************************************
**                        Private Function Declarations                       **
*******************************************************************************/"""
    fo.write(private_function_header_str)
    fo.write(f"\n#define {g_module_name.upper()}_START_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    build_detcheck_extern_to_autosar_c(fo,safety_enable_flag)
    build_g_apis_local_line_to_c(fo,g_ar_apis_dictlist)
    fo.write(f"\n#define {g_module_name.upper()}_STOP_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    
    # 增加header注释
    header_notes_str = """
/*******************************************************************************
**                        GLOBAL CONSTANTS/VARIABLES                          **
*******************************************************************************/"""
    fo.write(header_notes_str)
    generate_memmap_macros(fo, g_module_name, ["CONST", "VAR_CLEARED", "VAR_INIT", "VAR_SHARED"])
    
    build_g_apis_to_c(fo,g_ar_apis_dictlist,safety_enable_flag)
    
    build_detcheck_functions_to_autosar_c(fo,safety_enable_flag)
    fo.write(f"\n#define {g_module_name.upper()}_STOP_SEC_CODE\n")
    fo.write(f"#include \"{g_module_name}_MemMap.h\"\n")
    
    fo.close()
    logger.debug("Successfully generated %s.c"%(g_module_name))
    
def tree_autogen_reg_h(reg_tree):
    logger.debug('Enter function %s' % tree_autogen_reg_h.__name__)
    copyright_str = get_copyright_str(f'{g_product_prefix}{g_module_name}_reg.h')

    fo=open(f"{g_gen_file_path}{g_product_prefix}{g_module_name}_reg.h","w",encoding='utf-8')
    fo.write(copyright_str)
    fo.write("#ifndef %s%s_REG_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("#define %s%s_REG_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("/******************************************************************************/\n")
    fo.write(f"#include \"{g_product_prefix}_TypesReg.h\"\n")
    fo.write("\n")

    reg_tree_leaf_list = reg_tree.get_leaf_nodes(reg_tree)
    if g_l1_node_number == 1:
        fo.write("#define %s ((*(%s_%s *)0x%Xu))\n\n"%('MODULE_'+g_module_name.upper(),g_product_prefix,g_module_name.upper(),int(g_base_address,16)))
    else:
        for i in range(g_l1_node_number):
            if str(i) not in g_module_no_set:
                continue
            fo.write(f"#define MODULE_{g_module_name.upper()}{i} ((*({g_product_prefix}_{g_module_name.upper()} *)0x{(int(g_base_address,16)+i*reg_tree.children[0].size):X}u))\n\n")

    fo.write(f"#define {g_module_name.upper()}_CONTROLLERS_NO {g_l1_node_number}\n")
    if len(g_module_no_set) == 0 :
        fo.write(f"#define MODULE_{g_module_name.upper()}0 MODULE_{g_module_name.upper()}\n\n")
    fo.write("extern %s_UReg_32Bit %s_BASE_ADDRS[%s_CONTROLLERS_NO];\n\n"%(g_product_prefix,g_module_name.upper(),g_module_name.upper()))
    fo.write(f"#define {g_module_name.upper()}_GET_BASE_ADDRESS(offset) ((uint32)({g_module_name.upper()}_BASE_ADDRS[(offset)]))\n\n")

    for current_node in reg_tree_leaf_list:        
        fo.write(f"/** <#>brief {current_node.offset_start:X}, {current_node.long_name}({current_node.short_name}) */\n")
        get_path_without_root = current_node.get_path().replace(g_product_prefix+'_','')
        fo.write(f"#define {get_path_without_root}(offset) (*(volatile {current_node.get_original_values_path()} *)({g_module_name.upper()}_GET_BASE_ADDRESS(offset) + 0x{current_node.offset_start:X}u))\n")
        fo.write("\n")
                
    fo.write(f"#endif /*{g_product_prefix.upper()}{g_module_name.upper()}_REG_H */\n")
    fo.close()
    logger.debug("Successfully generated %s%s_reg.h"%(g_product_prefix,g_module_name))

# 递归查找第一个非Nan行。
def fill_nan_recursively(df, row, i, key='Ctrl.', default_value='I'):
    # 检查当前行的指定键是否为 NaN
    if pd.isna(row.get(key, np.nan)):
        # 如果不是第一行，则递归检查上一行
        if i > 0:
            # 获取上一行的值
            prev_value = df[i-1].get(key, np.nan)
            # 如果上一行的值也是 NaN，则继续递归
            if pd.isna(prev_value):
                # 递归调用函数
                fill_nan_recursively(df, row, i-1, key, default_value)
            else:
                # 如果上一行的值不是 NaN，则将其赋给当前行
                row[key] = prev_value
        # 如果已经是第一行，并且值仍然是 NaN，则赋默认值
        elif pd.isna(row.get(key, np.nan)):
            row[key] = default_value

#判断row['Symbol']和temp_row['Symbol']两个字符串相似，其中一个是另一个的子字符串，并且只多一个末尾字符
# 例如：i2c_scl 和 i2c_sclA
def string_is_like(symbol1, symbol2):
    if (symbol1 in symbol2 and len(symbol2) - len(symbol1) == 1) or \
       (symbol2 in symbol1 and len(symbol1) - len(symbol2) == 1):
        # 执行需要的操作，例如：
        # print("The symbols are similar.")
        return True
    else:
        return False

# 合并具有相同row['Ball']，row['Ctrl In Out']不同的行
def merge_similar_rows(sheet_dict_list,product_prefix,module_name):
    # 首先，根据 'Port' 和 'Direction' 排序
    #sheet_dict_list.sort(key=lambda x: (x['Ball'], x['Ctrl In Out']))
    
    merged_list = []
    temp_row = None
    
    for row in sheet_dict_list:
        if temp_row and row['Port'] == temp_row['Port'] and row['Direction'] != temp_row['Direction'] and string_is_like(row['Module'],temp_row['Module']):
            # 如果满足合并条件，则更新temp_row的相关字段
            if row['Function'] != '':
                temp_row['Function'] += '. ' + row['Function']
            temp_row['Direction'] = 'INOUT'
            # 其他需要合并的字段可以在这里继续添加
            # ...
            if len(row['Module']) < len(temp_row['Module']):
                temp_row['Module'] = row['Module']
            temp_row['Port Type'] = f"{product_prefix}{module_name}_InOut" #'IfxEvadc_Emux_inOut'
        else:
            if temp_row:
                merged_list.append(temp_row)
            temp_row = row.copy()

    
    # 添加最后一行（可能是未被添加的最后一组中的唯一行）
    if temp_row:
        merged_list.append(temp_row)
    
    return merged_list

def autogen_pinmap_h():
    logger.info('Enter function %s' % autogen_pinmap_h.__name__)

    # 打开excel表格
    excel_data = pd.read_excel(g_input_excel_file_path, sheet_name=None)

    fo=open(f"{g_gen_file_path}{g_product_prefix}{g_module_name}_PinMap.h","w",encoding='utf-8')
    fo.write("#ifndef %s%s_PINMAP_H\n"%(g_product_prefix.upper(),g_module_name.upper()))
    fo.write("#define %s%s_PINMAP_H\n\n"%(g_product_prefix.upper(),g_module_name.upper()))

    fo.write("#include \"%s%s_reg.h\"\n"%(g_product_prefix,g_module_name))
    module_name_upper = g_module_name.upper()

    type_define_str = """
/** <#>addtogroup {product_prefix}{module_name}_pinmap
* */

/** <#>brief {module_name_upper} input pin mapping structure */
typedef const struct
{{
    {product_prefix}_{module_name_upper}*        module;    /** <#>brief Base address */
    IfxPort_Pin       pin;       /** <#>brief Port pin */
    Ifx_RxSel         select;    /** <#>brief Input multiplexer value */
}} {product_prefix}{module_name}_In;

/** <#>brief {module_name_upper} output pin mapping structure */
typedef const struct
{{
    {product_prefix}_{module_name_upper}*        module;    /** <#>brief Base address */
    IfxPort_Pin       pin;       /** <#>brief Port pin */
    IfxPort_OutputIdx select;    /** <#>brief Port control code */
}} {product_prefix}{module_name}_Out;

/** <#>brief {module_name_upper} input/output pin mapping structure */
typedef const struct
{{
    {product_prefix}_{module_name_upper}*        module;    /** <#>brief Base address */
    IfxPort_Pin       pin;       /** <#>brief Port pin */
    Ifx_RxSel         inSelect;  /** <#>brief Input multiplexer value */
    IfxPort_OutputIdx outSelect; /** <#>brief Port control code */
}} {product_prefix}{module_name}_InOut;
""".format(product_prefix=g_product_prefix, module_name=g_module_name, module_name_upper=module_name_upper)
       
    fo.write(f"{type_define_str}\n")

    # 遍历'PinMap'sheet，构建词典列表
    all_sheet_dictlist = excel_data['PinMap'].to_dict(orient='records')

    for entry in all_sheet_dictlist:
        logger.info(entry)
        spread_module_name = entry['Module'].replace(f'{g_alias.upper()}',f'{g_module_name}')
        fo.write(f"{g_product_prefix.upper()}_EXTERN {entry['Port Type']} {g_product_prefix}{spread_module_name}_{entry['Port']}_{entry['Direction']};  /** <#>brief {entry['Function']}.{entry['Description']}*/\n")

    fo.write("\n")
    fo.write(f"#endif /*{g_product_prefix.upper()}{g_module_name.upper()}_PINMAP_H */\n")
    fo.close()
    logger.debug("Successfully generated %s%s_PinMap.h"%(g_product_prefix,g_module_name))

def init_autocoder(config: AutocoderConfig):
    """
    初始化autocoder，设置全局变量和日志
    
    Args:
        config: 配置对象
    """
    global logger, g_debug_level, g_mask_style, g_language, g_input_excel_file_path
    global g_trans_flag, g_source_code_root_path, g_trans_enum_item_description
    global g_reg_short_description, g_gen_file_path, g_sysinfo_dict
    global g_pinmap_filename, g_register_style, g_alias, g_variable_name_set
    global g_baseinfo_dict, g_bits_variable_dict, g_treeview_dict, g_reglist_shortname_key_dict
    global g_cfg_macro_dict, g_ar_macro_dict
    
    # 设置日志
    numeric_level = getattr(logging, config.debug_level.upper(), None)
    if not isinstance(numeric_level, int):
        raise ValueError('Invalid log level: %s' % config.debug_level)
    
    logger = logging.getLogger('autocoder')
    logger.setLevel(numeric_level)
    
    # 清除现有的handlers
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # 添加新的handler
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    # 设置全局变量
    g_debug_level = config.debug_level
    g_mask_style = config.mask_style
    g_language = config.language
    g_input_excel_file_path = config.input_file
    g_reg_short_description = config.reg_short_description
    g_gen_file_path = config.output_dir
    
    # 初始化其他全局变量
    g_pinmap_filename = ''
    g_register_style = ''
    g_alias = ''
    g_variable_name_set = {'n', 'i', 'x', 'y'}  # 默认变量名集合
    g_baseinfo_dict = {}
    g_bits_variable_dict = {}
    g_treeview_dict = {}
    g_reglist_shortname_key_dict = {}
    g_cfg_macro_dict = {}
    g_ar_macro_dict = {}
    
    # 语言相关设置
    if g_language.lower() == 'chinese':
        g_trans_flag = True
        g_source_code_root_path = os.path.join(config.output_dir, 'Chinese') + os.sep
    else:
        g_trans_flag = False
        g_source_code_root_path = os.path.join(config.output_dir, 'English') + os.sep
    
    g_trans_enum_item_description = g_trans_flag
    
    # 加载系统信息
    if config.sysinfo_json and os.path.exists(config.sysinfo_json):
        with open(config.sysinfo_json, 'r', encoding='utf-8') as f:
            sysinfo_data = json.load(f)
            g_sysinfo_dict = sysinfo_data.get('baseinfo', {})
    
    # 确保输出目录存在
    os.makedirs(g_gen_file_path, exist_ok=True)
    
    logger.info(f"Autocoder initialized with config: {config}")

def reset_autocoder():
    """
    重置autocoder的全局状态，用于多次运行
    """
    global g_reserved_reg_tree, g_spread_reg_tree, g_reg_tree
    global g_product_prefix, g_module_name, g_base_address, g_module_no_set
    global g_space_size, g_l1_node_number
    global g_functions_dictlist, g_dev_functions_dictlist, g_mcal_functions_dictlist
    global g_ar_functions_dictlist, g_dev_enum_dictlist, g_mcal_enum_dictlist
    global g_ar_enum_dictlist, g_mcal_type_macro_dictlist, g_reglist_dictlist
    global g_dev_structure_dictlist, g_mcal_structure_dictlist, g_ar_structure_dictlist
    global g_ar_cfg_interfaces_dictlist, g_ar_type_definitions_dictlist
    global g_ar_apis_dictlist, g_reg_apis_dictlist, g_ar_error_codes_dictlist
    global g_general_types_h_flag
    
    # 重置树结构
    g_reserved_reg_tree = []
    g_spread_reg_tree = []
    g_reg_tree = []
    
    # 重置模块信息
    g_product_prefix = []
    g_module_name = []
    g_base_address = []
    g_module_no_set = set()
    g_space_size = 0
    g_l1_node_number = 0
    
    # 重置所有字典列表
    g_functions_dictlist = []
    g_dev_functions_dictlist = []
    g_mcal_functions_dictlist = []
    g_ar_functions_dictlist = []
    g_dev_enum_dictlist = []
    g_mcal_enum_dictlist = []
    g_ar_enum_dictlist = []
    g_mcal_type_macro_dictlist = []
    g_reglist_dictlist = []
    g_dev_structure_dictlist = []
    g_mcal_structure_dictlist = []
    g_ar_structure_dictlist = []
    g_ar_cfg_interfaces_dictlist = []
    g_ar_type_definitions_dictlist = []
    g_ar_apis_dictlist = []
    g_reg_apis_dictlist = []
    g_ar_error_codes_dictlist = []
    
    # 重置标志
    g_general_types_h_flag = False
    
    if logger:
        logger.info("Autocoder state reset completed")

def run_autocoder() -> List[str]:
    """
    运行autocoder核心逻辑，生成代码文件
    
    Returns:
        List[str]: 生成的文件路径列表
    """
    global g_reg_apis_dictlist, g_ar_apis_dictlist, g_reg_tree, g_reserved_reg_tree, g_spread_reg_tree
    global g_input_excel_file_path, g_gen_file_path, g_pinmap_filename
    
    if logger is None:
        raise RuntimeError("Autocoder not initialized. Please call init_autocoder() first.")
    
    # 重置状态以确保干净的运行环境
    reset_autocoder()
    
    logger.info(f"autocoder v{g_version}. Start...")
    logger.info("Use " + g_input_excel_file_path)
    
    # 如果g_input_excel_file_path没有.xlsx后缀，则自动加上.xlsx
    if not g_input_excel_file_path.endswith(".xlsx"):
        g_input_excel_file_path += ".xlsx"
    logger.info("Use " + g_input_excel_file_path)

    if not os.path.exists(g_input_excel_file_path):
        raise FileNotFoundError(f"Not such file {g_input_excel_file_path}!")

    # 执行原始的主要逻辑
    get_g_ar_cfg_interfaces_dictlist()
    get_g_ar_type_definitions_dictlist()
    g_reg_apis_dictlist = get_g_apis_dictlist('RegApis')
    g_ar_apis_dictlist = get_g_apis_dictlist('ArApis')
    get_g_ar_error_codes_dictlist()

    read_global_setting_from_excel()
    autogen_reg_list_excel()
    fill_long_name_in_reglist_table()
    fill_width_mask_in_reglist_table()
    fill_enum_item_in_reglist_table()
    fill_enum_item_in_reglist_table('DevEnum')
    fill_enum_item_in_reglist_table('McalEnum')
    fill_enum_item_in_reglist_table('ArEnum')
    fill_enum_item_in_reglist_table('McalTypeMacro', 'TypeMacroItem')
    read_global_setting_from_excel()

    if len(g_pinmap_filename) > 0:
        rebuild_pinmap_sheet_in_excel()
        autogen_pinmap_h()

    # 从excel中读取表单，初始化一些全局词典变量
    get_g_sysinfo_dict()

    # 这里又两种实现路径：1.直接从excel读取人工写好的树  2.从excel读取数据，然后根据生成树。目前暂时采用路径1
    g_reg_tree = build_tree_from_excel()

    generated_files = []

    if g_reg_tree != []:
        g_reg_tree.traverse_and_save_original_values()

        # 复制一个新的树，用reserved填满所有空隙  
        g_reserved_reg_tree = g_reg_tree.gen_a_reserved_tree()
        g_reserved_reg_tree.traverse_and_save_original_values()
        print_tree(g_reg_tree)

        print_tree(g_reserved_reg_tree)
        # 生成regdef.h
        tree_autogen_regdef_h(g_reserved_reg_tree)

        # 复制一个新的树，把G[12]这样的所有object展开成G[0],G[1]...G[11]这样的结构
        g_spread_reg_tree = g_reg_tree.gen_a_spread_tree()
        print_tree(g_spread_reg_tree)
        # 生成reg.h
        tree_autogen_reg_h(g_spread_reg_tree)

        # 生成bf.h
        tree_autogen_bf_h(g_reg_tree)

        # 生成TypesReg.h
        autogen_typesreg_h()

        # 生成{module}.h
        tree_autogen_module_h(g_reg_tree)
        tree_autogen_module_c()

    # 生成Cfg.h PBCfg.c
    autogen_cfg_h()
    autogen_pbcfg_c()

    # 生成GeneralTypes.h
    if g_general_types_h_flag:
        tree_autogen_GeneralTypes_h()

    # 生成autosar   .h .c
    tree_autogen_autosar_h()
    tree_autogen_autosar_c()

    # 收集生成的文件
    generated_files = []
    output_path = Path(g_gen_file_path)
    
    # 收集所有可能生成的文件类型
    file_patterns = [
        '*.h',   # 头文件
        '*.c',   # C源文件
        '*.hpp', # C++头文件
        '*.cpp', # C++源文件
    ]
    
    for pattern in file_patterns:
        for file_path in output_path.glob(pattern):
            if file_path.is_file():
                generated_files.append(str(file_path.absolute()))
    
    # 如果没有找到生成的文件，至少返回输出目录
    if not generated_files:
        generated_files.append(str(output_path.absolute()))

    from datetime import datetime
    now = datetime.now()
    current_time_str = now.strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"Time: {current_time_str}. Generate Success!")
    logger.info(f"Generated {len(generated_files)} files: {generated_files}")
    
    return generated_files

def convert_excel_to_code(input_file: str, output_dir: str = './converted_markdown_files/', 
                         debug_level: str = 'info', language: str = 'english',
                         reg_short_description: bool = True, mask_style: str = 'nxp',
                         sysinfo_json: str = '') -> List[str]:
    """
    便利函数：直接从参数运行Excel到代码的转换
    
    Args:
        input_file: 输入Excel文件路径
        output_dir: 输出目录
        debug_level: 调试级别 ('debug', 'info', 'warning', 'error')
        language: 语言 ('english', 'chinese')
        reg_short_description: 是否使用寄存器短描述
        mask_style: 掩码样式 ('nxp', 'infineon', 'arkuart')
        sysinfo_json: 系统信息JSON文件路径
    
    Returns:
        List[str]: 生成的文件路径列表
    """
    config = AutocoderConfig(
        debug_level=debug_level,
        language=language,
        reg_short_description=reg_short_description,
        mask_style=mask_style,
        input_file=input_file,
        output_dir=output_dir,
        sysinfo_json=sysinfo_json
    )
    
    init_autocoder(config)
    return run_autocoder()

